#!/usr/bin/env python3
"""Build the blank, public FTE roster template shipped with WFMHub."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TARGET = ROOT / "templates" / "FTE Count.xlsx"
AGENT_HEADERS = [
    "Client ID",
    "Status",
    "Name",
    "Team leader",
    "Ops Manager",
    "LOB",
    "Market",
    "Language",
    "Location",
    "City",
    "FTE",
    "End date if leaver",
]
AGENT_ROW_FIELDS = [
    "agent_id", "employment_status", "agent_name", "team_leader", "ops_manager",
    "lob", "market", "language", "location", "city", "fte", "end_date",
]

PTO_HEADERS = [
    "Client ID",
    "Name",
    "Start date",
    "End date",
    "Day coverage",
    "Start time",
    "End time",
    "PTO type",
    "Approval status",
    "Comment",
]
PTO_ROW_FIELDS = [
    "agent_id", "agent_name", "start_date", "end_date", "day_coverage",
    "start_time", "end_time", "pto_type", "approval_status", "comment",
]

AWAY_HEADERS = [
    "Client ID",
    "Name",
    "Start date",
    "End date",
    "Away type",
    "Case status",
    "Comment",
]
AWAY_ROW_FIELDS = [
    "agent_id", "agent_name", "start_date", "end_date", "away_type",
    "case_status", "comment",
]


HEADER_FILL = PatternFill("solid", fgColor="0F3B42")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
THIN = Side(style="thin", color="D5DADD")
MISSING_FILL = PatternFill("solid", fgColor="FCE8E6")


def _add_input_sheet(
    workbook: Workbook,
    name: str,
    headers: Sequence[str],
    row_fields: Sequence[str],
    table_name: str,
    widths: Sequence[int],
    rows: Iterable[dict[str, Any]] = (),
):
    """Create one governed, filterable user-input table."""

    rows = list(rows)
    sheet = workbook.create_sheet(name)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(1, column, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=THIN)
    sheet.row_dimensions[1].height = 32
    if rows:
        for row_number, row in enumerate(rows, 2):
            for column, field in enumerate(row_fields, 1):
                sheet.cell(row_number, column, row.get(field))
    else:
        # Keep one blank table row so filters and structured formatting exist
        # the first time the workbook is opened.
        for column in range(1, len(headers) + 1):
            sheet.cell(2, column, None)
    last_table_row = max(2, len(rows) + 1)
    last_column = sheet.cell(1, len(headers)).column_letter
    table = Table(displayName=table_name, ref=f"A1:{last_column}{last_table_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[sheet.cell(1, column).column_letter].width = width
    return sheet


def build_template(
    target: Path = DEFAULT_TARGET,
    agent_rows: Iterable[dict[str, Any]] = (),
    pto_rows: Iterable[dict[str, Any]] = (),
    away_rows: Iterable[dict[str, Any]] = (),
) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    agent_rows = list(agent_rows)
    workbook = Workbook()
    workbook.properties.title = "WFMHub standard FTE roster template"
    workbook.properties.subject = "Authoritative agent scope for WFMHub"
    workbook.properties.creator = "WFMHub Portable"
    workbook.properties.company = "WFM"

    guide = workbook.active
    guide.title = "START_HERE"
    guide.sheet_view.showGridLines = False
    guide.merge_cells("A1:F1")
    guide["A1"] = "WFMHub standard FTE roster"
    guide["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F2933")
    guide["A1"].alignment = Alignment(vertical="center")
    guide.row_dimensions[1].height = 28
    instructions = [
        "1. Agent: maintain one row per employee. Client ID is the Verint Agent ID and must stay as text.",
        "2. PTO: add approved vacation/PTO dates. Use one row per continuous period; End date is inclusive.",
        "3. For part-day PTO, choose Partial day and enter both Start time and End time. Leave times blank for Full day.",
        "4. Away: add long absences such as long sickness or maternity. Leave End date blank while the case is open.",
        "5. Use the dropdown statuses. Pending/cancelled records must never be treated as approved operational facts.",
        "6. Keep the sheet names and header cells unchanged. Do not add formulas inside the three input tables.",
        "7. Save this file as FTE Count.xlsx inside the configured FTE folder.",
        "8. Never paste worldwide LILO, Agent Status, schedule, or call data into this workbook.",
        "9. WFMHub scopes extracts through the Agent sheet by Agent ID or one unique normalized-name match.",
        "10. PTO/Away are population-ready registers; this release does not apply them to attendance or staffing calculations yet.",
    ]
    guide["A3"] = "How to use it"
    guide["A3"].font = Font(name="Calibri", size=12, bold=True, color="0F3B42")
    for row, instruction in enumerate(instructions, 4):
        guide.cell(row, 1, instruction)
        guide.cell(row, 1).font = Font(name="Calibri", size=11, color="1F2933")
        guide.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        guide.row_dimensions[row].height = 26
    guide["A16"] = "Agent required"
    guide["B16"] = "Client ID and Name"
    guide["A17"] = "PTO required"
    guide["B17"] = "Client ID, Start date, End date, Day coverage, PTO type, Approval status"
    guide["A18"] = "Away required"
    guide["B18"] = "Client ID, Start date, Away type, Case status; End date may be blank while active"
    guide["A19"] = "Operational rule"
    guide["B19"] = "Verint Activities remain the final corrected/payroll record after re-export"
    for row in range(16, 20):
        guide.cell(row, 1).font = Font(bold=True, color="0F3B42")
    guide.column_dimensions["A"].width = 105
    guide.column_dimensions["B"].width = 42

    sheet = _add_input_sheet(
        workbook, "Agent", AGENT_HEADERS, AGENT_ROW_FIELDS, "tblFTEAgents",
        [16, 14, 28, 24, 24, 18, 16, 14, 18, 18, 10, 20], agent_rows,
    )
    for row in range(2, 5001):
        sheet.cell(row, 1).number_format = "@"
        sheet.cell(row, 11).number_format = "0.00"
        sheet.cell(row, 12).number_format = "yyyy-mm-dd"

    status_validation = DataValidation(
        type="list",
        formula1='"Active,Inactive,Leaver"',
        allow_blank=True,
    )
    status_validation.promptTitle = "Employment status"
    status_validation.prompt = "Choose Active, Inactive, or Leaver."
    sheet.add_data_validation(status_validation)
    status_validation.add("B2:B5000")

    fte_validation = DataValidation(
        type="decimal",
        operator="between",
        formula1="0",
        formula2="1",
        allow_blank=True,
    )
    fte_validation.promptTitle = "FTE"
    fte_validation.prompt = "Enter a value from 0 to 1, for example 1 or 0.5."
    sheet.add_data_validation(fte_validation)
    fte_validation.add("K2:K5000")

    sheet.conditional_formatting.add(
        "A2:A5000",
        FormulaRule(formula=['AND(COUNTA($A2:$L2)>0,$A2="")'], fill=MISSING_FILL),
    )
    sheet.conditional_formatting.add(
        "C2:C5000",
        FormulaRule(formula=['AND(COUNTA($A2:$L2)>0,$C2="")'], fill=MISSING_FILL),
    )
    sheet["A1"].comment = Comment("Required. Store as text; keep leading zeros.", "WFMHub")
    sheet["C1"].comment = Comment("Required. Use the agent name from operational exports when possible.", "WFMHub")
    sheet["L1"].comment = Comment("Leave blank for active employees.", "WFMHub")

    pto = _add_input_sheet(
        workbook, "PTO", PTO_HEADERS, PTO_ROW_FIELDS, "tblFTEPTO",
        [16, 28, 14, 14, 16, 13, 13, 20, 18, 38], pto_rows,
    )
    for row in range(2, 5001):
        pto.cell(row, 1).number_format = "@"
        pto.cell(row, 3).number_format = "yyyy-mm-dd"
        pto.cell(row, 4).number_format = "yyyy-mm-dd"
        pto.cell(row, 6).number_format = "hh:mm"
        pto.cell(row, 7).number_format = "hh:mm"
    day_validation = DataValidation(type="list", formula1='"Full day,Partial day"', allow_blank=False)
    pto.add_data_validation(day_validation)
    day_validation.add("E2:E5000")
    pto_type_validation = DataValidation(
        type="list", formula1='"Vacation,Paid leave,Personal leave,Other"', allow_blank=False,
    )
    pto.add_data_validation(pto_type_validation)
    pto_type_validation.add("H2:H5000")
    approval_validation = DataValidation(
        type="list", formula1='"Approved,Pending,Cancelled"', allow_blank=False,
    )
    pto.add_data_validation(approval_validation)
    approval_validation.add("I2:I5000")
    for column in ("A", "C", "D", "E", "H", "I"):
        pto.conditional_formatting.add(
            f"{column}2:{column}5000",
            FormulaRule(formula=[f'AND(COUNTA($A2:$J2)>0,${column}2="")'], fill=MISSING_FILL),
        )
    pto["A1"].comment = Comment("Required. Must match the Agent sheet Client ID / Verint Agent ID.", "WFMHub")
    pto["D1"].comment = Comment("Inclusive. For one day, repeat the Start date.", "WFMHub")
    pto["F1"].comment = Comment("Required only when Day coverage is Partial day.", "WFMHub")
    pto["G1"].comment = Comment("Required only when Day coverage is Partial day.", "WFMHub")
    pto["I1"].comment = Comment("Only Approved rows are eligible for operational use.", "WFMHub")

    away = _add_input_sheet(
        workbook, "Away", AWAY_HEADERS, AWAY_ROW_FIELDS, "tblFTEAway",
        [16, 28, 14, 14, 24, 18, 42], away_rows,
    )
    for row in range(2, 5001):
        away.cell(row, 1).number_format = "@"
        away.cell(row, 3).number_format = "yyyy-mm-dd"
        away.cell(row, 4).number_format = "yyyy-mm-dd"
    away_type_validation = DataValidation(
        type="list",
        formula1='"Long sickness,Maternity/Parental,Administrative leave,Other"',
        allow_blank=False,
    )
    away.add_data_validation(away_type_validation)
    away_type_validation.add("E2:E5000")
    case_validation = DataValidation(
        type="list", formula1='"Active,Planned,Closed,Cancelled"', allow_blank=False,
    )
    away.add_data_validation(case_validation)
    case_validation.add("F2:F5000")
    for column in ("A", "C", "E", "F"):
        away.conditional_formatting.add(
            f"{column}2:{column}5000",
            FormulaRule(formula=[f'AND(COUNTA($A2:$G2)>0,${column}2="")'], fill=MISSING_FILL),
        )
    away["A1"].comment = Comment("Required. Must match the Agent sheet Client ID / Verint Agent ID.", "WFMHub")
    away["D1"].comment = Comment("Inclusive. Leave blank while an Active case is open.", "WFMHub")
    away["F1"].comment = Comment("Cancelled rows are never eligible for operational use.", "WFMHub")

    workbook.active = 0
    workbook.save(target)
    return target


def standardize_source(source: Path, target: Path) -> Path:
    sys.path.insert(0, str(ROOT / "src"))
    from wfmhub.ingestion import parse_fte

    rows = parse_fte(source.resolve(), "standardized-template").tables["raw.fte_agent"]
    return build_template(target.resolve(), rows)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, help="Optional populated FTE workbook to standardize")
    parser.add_argument("--output", type=Path, default=DEFAULT_TARGET)
    args = parser.parse_args()
    print(standardize_source(args.source, args.output) if args.source else build_template(args.output))
