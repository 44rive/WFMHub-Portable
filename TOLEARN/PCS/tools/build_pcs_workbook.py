#!/usr/bin/env python3
"""Build a ready-to-send PCS daily workbook from a WFMHub pcs_agent_day export.

Standalone. Imports nothing from wfmhub and writes nothing into the hub folder.
It reads one exported file and writes one new workbook.

This is the fallback path for anyone without Power Pivot, and a way to produce a
static pack for distribution. The Power Query + Data Model build described in
PCS_Setup_ClickByClick.md is the primary route.

Every ratio here is computed as a sum divided by a sum, never as an average of
the pre-computed per-agent-day ratio columns. A zero denominator yields None,
which is written as a blank cell -- never 0.00.

Usage
-----
    python build_pcs_workbook.py data/current/pcs_agent_day.csv
    python build_pcs_workbook.py <export> --out output/PCS_Daily.xlsx --min-responses 5

Requires xlsxwriter. openpyxl is needed only to read an .xlsx export.
"""

from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    import xlsxwriter
except ImportError:
    sys.exit("xlsxwriter is required.  Run this with WFMHub's runtime python.exe, "
             "which already has it:  runtime\\python.exe build_pcs_workbook.py <export>")


# --- Columns that are ratios already. Never summed, never averaged. -----------
FORBIDDEN = {
    "average_talk_seconds", "average_hold_seconds", "average_wrap_seconds",
    "average_handle_seconds", "response_rate", "q1_average", "q2_average",
    "pcs_average", "top_box_percent", "low_score_percent",
}

# --- Counters that are safe to sum -------------------------------------------
SUMMABLE = [
    "call_legs", "handled_calls", "inbound_calls", "outbound_calls",
    "talk_seconds", "hold_seconds", "wrap_seconds", "handle_seconds",
    "pcs_enabled_calls", "survey_responses", "q1_response_count", "q1_score_sum",
    "q2_response_count", "q2_score_sum", "pcs_score_count", "pcs_score_sum",
    "top_box_responses", "low_score_responses", "comments_count",
]

REQUIRED = {"business_date", "agent_id", "pcs_enabled_calls", "survey_responses",
            "pcs_score_sum", "pcs_score_count"}


def die(message: str) -> None:
    sys.exit(f"ERROR: {message}")


def to_number(value) -> float:
    """Blank, None and unparseable all mean 'nothing to add'."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            continue
    return None


def read_export(path: Path) -> list[dict]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
    elif path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError:
            die("openpyxl is needed to read an .xlsx export. Export as CSV instead.")
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook["DATA"] if "DATA" in workbook.sheetnames else workbook.worksheets[0]
            iterator = sheet.iter_rows(values_only=True)
            headers = [str(cell or "").strip() for cell in next(iterator)]
            rows = [dict(zip(headers, values)) for values in iterator]
        finally:
            workbook.close()
    else:
        die(f"Unsupported file type {path.suffix!r}. Use the CSV or XLSX export.")

    if not rows:
        die(f"{path.name} contains no data rows.")

    missing = REQUIRED - set(rows[0])
    if missing:
        die(f"{path.name} is missing expected columns: {', '.join(sorted(missing))}. "
            "Is this really the pcs_agent_day export?")
    return rows


def prepare(rows: list[dict]) -> list[dict]:
    """Type the rows and drop the pre-computed ratio columns."""
    out, dropped, keys = [], 0, set()
    for raw in rows:
        business_date = to_date(raw.get("business_date"))
        if business_date is None:
            dropped += 1
            continue
        # Newest-file-wins would need file metadata we do not have here, so a
        # repeated key means the caller concatenated exports. Keep the first and
        # report it rather than silently doubling every count.
        key = raw.get("agent_day_key") or f"{business_date}-{raw.get('agent_id')}"
        if key in keys:
            dropped += 1
            continue
        keys.add(key)

        row = {name: (str(raw.get(name) or "").strip() or None) for name in (
            "agent_day_key", "agent_id", "agent_name", "team_leader",
            "ops_manager", "lob", "market", "language", "location")}
        row["business_date"] = business_date
        for name in SUMMABLE:
            row[name] = to_number(raw.get(name))
        out.append(row)

    if dropped:
        print(f"  note: skipped {dropped} row(s) - unparseable date or duplicate agent-day key")
    return out


def ratio(numerator: float, denominator: float):
    """Blank, not zero, when there is nothing to divide by."""
    return numerator / denominator if denominator else None


def aggregate(rows: list[dict]) -> dict[str, float]:
    total = defaultdict(float)
    for row in rows:
        for name in SUMMABLE:
            total[name] += row[name]
    return total


def measures(total: dict[str, float], agents: int | None = None) -> dict:
    """Every ratio is a sum over a sum. This is the whole point of the file."""
    return {
        "Agents": agents,
        "Handled calls": total["handled_calls"],
        "PCS-enabled calls": total["pcs_enabled_calls"],
        "Survey responses": total["survey_responses"],
        "Response rate": ratio(total["survey_responses"], total["pcs_enabled_calls"]),
        "PCS average": ratio(total["pcs_score_sum"], total["pcs_score_count"]),
        "Q1 average": ratio(total["q1_score_sum"], total["q1_response_count"]),
        "Q2 average": ratio(total["q2_score_sum"], total["q2_response_count"]),
        "Top box %": ratio(total["top_box_responses"], total["survey_responses"]),
        "Low score %": ratio(total["low_score_responses"], total["survey_responses"]),
        "AHT (sec)": ratio(total["handle_seconds"], total["handled_calls"]),
        "Comments": total["comments_count"],
    }


MEASURE_ORDER = ["Agents", "Handled calls", "PCS-enabled calls", "Survey responses",
                 "Response rate", "PCS average", "Q1 average", "Q2 average",
                 "Top box %", "Low score %", "AHT (sec)", "Comments"]

FIELD_LABELS = {
    "ops_manager": "Ops Manager", "team_leader": "Team Leader", "lob": "LOB",
    "agent_name": "Agent", "agent_id": "Agent ID", "language": "Language",
    "market": "Site", "location": "City", "business_date": "Date",
}


def label_of(field: str) -> str:
    return FIELD_LABELS.get(field, field.replace("_", " ").title())


PERCENT_MEASURES = {"Response rate", "Top box %", "Low score %"}
DECIMAL_MEASURES = {"PCS average", "Q1 average", "Q2 average", "AHT (sec)"}


def anchor_date(rows: list[dict]) -> tuple[date, bool]:
    """Anchor the periods on today, unless the data does not reach today.

    A stale extract would otherwise leave Today and Yesterday empty and make the
    whole report look broken. Anchoring on the newest business date keeps the
    views meaningful, and the caller shows a staleness banner so nobody mistakes
    an old extract for a fresh one.
    """
    today = date.today()
    newest = max(r["business_date"] for r in rows)
    return (today, False) if newest >= today else (newest, True)


def periods(rows: list[dict]) -> list[tuple[str, list[dict]]]:
    """Latest day (partial) / prior day / rolling 7 / month to date."""
    anchor, stale = anchor_date(rows)
    prior = anchor - timedelta(days=1)
    week_start = anchor - timedelta(days=7)
    month_start = anchor.replace(day=1)
    latest_label = f"{anchor:%Y-%m-%d}" + ("" if stale else " (partial)")
    return [
        (f"Latest day {latest_label}", [r for r in rows if r["business_date"] == anchor]),
        (f"Prior day {prior:%Y-%m-%d}", [r for r in rows if r["business_date"] == prior]),
        ("Rolling 7 days", [r for r in rows if week_start <= r["business_date"] < anchor]),
        (f"Month of {anchor:%b %Y}", [r for r in rows if r["business_date"] >= month_start]),
        ("Full export window", rows),
    ]


def group_by(rows: list[dict], *fields: str) -> dict[tuple, list[dict]]:
    grouped = defaultdict(list)
    for row in rows:
        grouped[tuple(row.get(f) or "(blank)" for f in fields)].append(row)
    return grouped


# --- Workbook -----------------------------------------------------------------

class Book:
    def __init__(self, path: Path):
        self.wb = xlsxwriter.Workbook(str(path), {"default_date_format": "yyyy-mm-dd"})
        f = self.wb.add_format
        self.title = f({"bold": True, "font_size": 15, "font_color": "#1F3864"})
        self.sub = f({"font_size": 9, "italic": True, "font_color": "#666666"})
        self.head = f({"bold": True, "bg_color": "#1F3864", "font_color": "white",
                       "border": 1, "text_wrap": True, "valign": "vcenter"})
        self.body = f({"border": 1})
        self.int_ = f({"border": 1, "num_format": "#,##0"})
        self.dec = f({"border": 1, "num_format": "0.00"})
        self.pct = f({"border": 1, "num_format": "0.0%"})
        self.date = f({"border": 1, "num_format": "yyyy-mm-dd"})
        self.warn = f({"border": 1, "bg_color": "#FCE4E4", "font_color": "#9C0006"})
        self.note = f({"font_size": 9, "italic": True, "font_color": "#9C0006"})
        self.wrap = f({"text_wrap": True, "valign": "top", "border": 1})

    def fmt_for(self, name: str):
        if name in PERCENT_MEASURES:
            return self.pct
        if name in DECIMAL_MEASURES:
            return self.dec
        return self.int_

    def write_measure(self, ws, row, col, name, value):
        # None means no data. Leave the cell genuinely empty.
        if value is None:
            ws.write_blank(row, col, None, self.fmt_for(name))
        else:
            ws.write_number(row, col, value, self.fmt_for(name))


def sheet_header(book: Book, ws, title: str, subtitle: str, stamp: str, width: int = 8):
    ws.hide_gridlines(2)
    ws.merge_range(0, 0, 0, width, title, book.title)
    ws.merge_range(1, 0, 1, width, subtitle, book.sub)
    ws.write(0, width + 1, "Data as at", book.sub)
    ws.write(1, width + 1, stamp, book.sub)


def build(rows: list[dict], out: Path, min_responses: int, source_name: str,
          refresh_stamp: str, rollup_by: list[str], scorecard_by: list[str]) -> None:
    book = Book(out)
    dates = sorted({r["business_date"] for r in rows})
    window = f"{dates[0]:%Y-%m-%d} to {dates[-1]:%Y-%m-%d}" if dates else "no dates"
    stamp = refresh_stamp or datetime.now().strftime("%Y-%m-%d %H:%M")

    # ---- SUMMARY -------------------------------------------------------------
    ws = book.wb.add_worksheet("SUMMARY")
    sheet_header(book, ws, "PCS daily summary",
                 f"Source: {source_name}   |   Window: {window}   |   "
                 f"Every ratio is a sum over a sum. A blank cell means no data, not a zero score.",
                 stamp, 6)
    ws.write(3, 0, "Measure", book.head)
    period_list = periods(rows)
    for index, (label, _) in enumerate(period_list):
        ws.write(3, index + 1, label, book.head)
    for r, name in enumerate(MEASURE_ORDER, 4):
        ws.write(r, 0, name, book.body)
        for index, (_, subset) in enumerate(period_list):
            agents = len({row["agent_id"] for row in subset}) if subset else 0
            values = measures(aggregate(subset), agents) if subset else {}
            book.write_measure(ws, r, index + 1, name, values.get(name))
    anchor, stale = anchor_date(rows)
    row_note = len(MEASURE_ORDER) + 5
    if stale:
        ws.write(row_note, 0,
                 f"STALE EXTRACT: the newest business date in this export is "
                 f"{anchor:%Y-%m-%d}, which is {(date.today() - anchor).days} day(s) before "
                 f"today ({date.today():%Y-%m-%d}). All periods are anchored on the newest "
                 f"data day, NOT on today. Refresh the source extract before sending.",
                 book.note)
        row_note += 1
    ws.write(row_note, 0,
             "The most recent day is PARTIAL - surveys arrive after the call ends, so its "
             "response rate is always understated and rises through the day. Do not compare "
             "a partial day against a complete one.", book.note)
    ws.set_column(0, 0, 22)
    ws.set_column(1, len(period_list), 17)
    ws.freeze_panes(4, 1)

    # ---- ROLLUP / TL / AGENT -------------------------------------------------
    def breakdown(sheet_name, title, subtitle, fields, headers, subset,
                  rank=False, sort_key=None):
        ws = book.wb.add_worksheet(sheet_name)
        sheet_header(book, ws, title, subtitle, stamp, len(headers) + len(fields) - 1)
        for c, name in enumerate(headers + MEASURE_ORDER[1:]):
            ws.write(3, c, name, book.head)
        grouped = group_by(subset, *fields)
        records = []
        for key, group in grouped.items():
            total = aggregate(group)
            values = measures(total)
            values["Agents"] = len({row["agent_id"] for row in group})
            rankable = total["survey_responses"] >= min_responses
            records.append((key, values, rankable, total))
        if sort_key:
            records.sort(key=sort_key)
        if rank:
            # Rankable first, ordered worst-first: that is where coaching goes.
            records.sort(key=lambda item: (
                not item[2],
                item[1]["PCS average"] if item[1]["PCS average"] is not None else 99,
            ))
        r = 4
        for key, values, rankable, total in records:
            for c, part in enumerate(key):
                ws.write(r, c, part, book.body if rankable or not rank else book.warn)
            for c, name in enumerate(MEASURE_ORDER[1:], len(key)):
                value = values.get(name)
                # Below the threshold the average is not shown at all: an
                # unreliable number on a coaching sheet is worse than a gap.
                if rank and not rankable and name in {"PCS average", "Q1 average", "Q2 average"}:
                    value = None
                book.write_measure(ws, r, c, name, value)
            r += 1
        if rank:
            ws.write(r + 1, 0,
                     f"Rows shaded red have fewer than {min_responses} responses in this "
                     "period. They are shown for completeness, with their response count, "
                     "but their average is withheld because it is not reliable.", book.note)
        for c in range(len(headers)):
            ws.set_column(c, c, 24)
        ws.set_column(len(headers), len(headers) + len(MEASURE_ORDER), 15)
        ws.freeze_panes(4, len(headers))
        ws.autofilter(3, 0, max(4, r - 1), len(headers) + len(MEASURE_ORDER) - 2)
        return records

    mtd = [r for r in rows if r["business_date"] >= date.today().replace(day=1)]
    rolling = [r for r in rows if
               date.today() - timedelta(days=7) <= r["business_date"] < date.today()]
    scope = mtd if mtd else rows

    breakdown("ROLLUP", "PCS by " + " and ".join(label_of(f) for f in rollup_by),
              f"Most recent month in the export. Window {window}. No agent names here.",
              tuple(rollup_by), [label_of(f) for f in rollup_by], scope,
              # Volume first: an alphabetical rollup buries the groups that
              # actually carry the feedback behind rows of zeros.
              sort_key=lambda item: (-item[3]["survey_responses"],
                                     -item[3]["pcs_enabled_calls"], item[0]))

    tl_scope = rolling if rolling else scope
    breakdown("TL_SCORECARD", "Agent PCS by " + label_of(scorecard_by[0]),
              f"Rolling 7 days. Response count sits beside every average, deliberately. "
              f"Minimum {min_responses} responses to be ranked.",
              tuple(scorecard_by), [label_of(f) for f in scorecard_by], tl_scope,
              rank=True)

    breakdown("BY_AGENT", "Agent ranking",
              f"Rolling 7 days, worst first. Minimum {min_responses} responses to be ranked.",
              ("agent_name", "market", "lob"), ["Agent", "Site", "LOB"],
              tl_scope, rank=True)

    # ---- DAILY_TREND ---------------------------------------------------------
    ws = book.wb.add_worksheet("DAILY_TREND")
    sheet_header(book, ws, "Daily PCS trend",
                 "Volume sits behind the average, so a swing on three responses reads "
                 "as a swing on three responses.", stamp, 7)
    trend_headers = ["Date", "PCS average", "Response rate", "Survey responses",
                     "PCS-enabled calls", "Top box %", "Low score %", "AHT (sec)"]
    for c, name in enumerate(trend_headers):
        ws.write(3, c, name, book.head)
    by_date = group_by(rows, "business_date")
    r = 4
    for (day,), group in sorted(by_date.items(), key=lambda item: item[0]):
        values = measures(aggregate(group))
        ws.write_datetime(r, 0, datetime.combine(day, datetime.min.time()), book.date)
        for c, name in enumerate(trend_headers[1:], 1):
            book.write_measure(ws, r, c, name, values.get(name))
        r += 1
    if r > 5:
        chart = book.wb.add_chart({"type": "column"})
        chart.add_series({"name": "Survey responses",
                          "categories": ["DAILY_TREND", 4, 0, r - 1, 0],
                          "values": ["DAILY_TREND", 4, 3, r - 1, 3],
                          "fill": {"color": "#BDD7EE"}})
        line = book.wb.add_chart({"type": "line"})
        line.add_series({"name": "PCS average",
                         "categories": ["DAILY_TREND", 4, 0, r - 1, 0],
                         "values": ["DAILY_TREND", 4, 1, r - 1, 1],
                         "y2_axis": True,
                         "line": {"color": "#1F3864", "width": 2.25},
                         "marker": {"type": "circle", "size": 5}})
        chart.combine(line)
        chart.set_title({"name": "PCS average against response volume"})
        chart.set_y_axis({"name": "Survey responses"})
        chart.set_y2_axis({"name": "PCS average", "min": 1, "max": 5})
        chart.set_size({"width": 900, "height": 340})
        chart.set_legend({"position": "bottom"})
        ws.insert_chart(r + 2, 0, chart)
    ws.set_column(0, 0, 13)
    ws.set_column(1, len(trend_headers) - 1, 17)
    ws.freeze_panes(4, 1)

    # ---- EXCEPTIONS ----------------------------------------------------------
    ws = book.wb.add_worksheet("EXCEPTIONS")
    sheet_header(book, ws, "Exceptions - action list",
                 "Rolling 7 complete days. Keep this to one screen; an exception list "
                 "that needs scrolling gets ignored.", stamp, 6)
    r = 3

    def block(heading, records, columns):
        nonlocal r
        ws.write(r, 0, heading, book.head)
        for c in range(1, len(columns)):
            ws.write(r, c, "", book.head)
        r += 1
        for c, name in enumerate(columns):
            ws.write(r, c, name, book.head)
        r += 1
        if not records:
            ws.write(r, 0, "Nothing to action.", book.body)
            for c in range(1, len(columns)):
                ws.write_blank(r, c, None, book.body)
            r += 2
            return
        for record in records:
            for c, value in enumerate(record):
                if isinstance(value, float) and columns[c].endswith("%"):
                    ws.write_number(r, c, value, book.pct)
                elif isinstance(value, (int, float)):
                    ws.write_number(r, c, value, book.int_)
                else:
                    ws.write(r, c, value, book.body)
            r += 1
        r += 1

    silent, low, thin = [], [], []
    for (agent, tl, lob), group in group_by(tl_scope, "agent_name", "market", "lob").items():
        total = aggregate(group)
        enabled = total["pcs_enabled_calls"]
        responses = total["survey_responses"]
        if enabled > 0 and responses == 0:
            silent.append((agent, tl, lob, enabled))
        if responses >= min_responses:
            low_pct = ratio(total["low_score_responses"], responses)
            if low_pct is not None and low_pct > 0.20:
                low.append((agent, tl, lob, responses, low_pct))
        if 0 < responses < min_responses:
            thin.append((agent, tl, lob, responses, enabled))

    silent.sort(key=lambda item: -item[3])
    low.sort(key=lambda item: -item[4])
    thin.sort(key=lambda item: item[3])

    block("ENABLED BUT SILENT - PCS-enabled calls, zero responses",
          silent[:25], ["Agent", "Site", "LOB", "PCS-enabled calls"])
    block(f"LOW-SCORE CONCENTRATION - over 20% low scores, minimum {min_responses} responses",
          low[:25], ["Agent", "Site", "LOB", "Responses", "Low score %"])
    block(f"INSUFFICIENT DATA - fewer than {min_responses} responses, not ranked",
          thin[:25], ["Agent", "Site", "LOB", "Responses", "PCS-enabled calls"])
    ws.set_column(0, 2, 26)
    ws.set_column(3, 5, 18)

    # ---- DEFINITIONS ---------------------------------------------------------
    ws = book.wb.add_worksheet("DEFINITIONS")
    sheet_header(book, ws, "Definitions", "Agree these before the first send.", stamp, 2)
    ws.write(3, 0, "Term", book.head)
    ws.write(3, 1, "Definition", book.head)
    definitions = [
        ("PCS-enabled call",
         "An inbound call leg whose Storm PostCallSurveyMode field reads 2. It means the call "
         "was in scope to be surveyed. It does NOT mean a survey was delivered, and it does "
         "NOT require the call to have been handled."),
        ("Leg, not interaction",
         "A transferred call is two legs. Both count as PCS-enabled, but only one survey can "
         "come back. Response rate is therefore a floor, not a precise collection rate. Use it "
         "to compare teams and track direction, not to state what share of customers replied."),
        ("Survey response",
         "A PCS-enabled leg where at least one scored question returned a valid score on the "
         "1 to 5 scale. A call answering only one question still counts as a full response."),
        ("Response rate", "Survey responses / PCS-enabled calls."),
        ("Call score",
         "The mean of that one call's valid answers across the scored questions (Q1 and Q2). "
         "This single number is what feeds top box, low score and the PCS average."),
        ("PCS average",
         "Sum of call scores / number of scored responses. Every response weighs equally, "
         "whatever its own question count. Never an average of agent averages."),
        ("Top box %",
         "Share of responses whose CALL score is 4.0 or higher. Applied to the call's average, "
         "not to any single question: Q1=5 with Q2=3 averages 4.0 and counts as top box."),
        ("Low score %", "Share of responses whose call score is 2.0 or lower."),
        ("AHT", "(Talk + hold + wrap) / handled calls. Handled means any of the three exceeded zero."),
        ("Blank is not zero",
         "A blank average means there were no valid responses. It is NOT a score of zero. A "
         "zero would misread as poor performance and would wrongly drag the team average down."),
        ("...but 0% response rate is real",
         "An agent with 40 PCS-enabled calls and zero responses shows a BLANK average but a 0.0% "
         "response rate. The average has nothing to average; the response rate has a real "
         "denominator, so 0% is a true and actionable fact. An agent with no enabled calls at "
         "all shows blank for both. A ratio appears whenever its denominator exists."),
        (f"Minimum {min_responses} responses",
         "Agents below the threshold are shown with their response count but their average is "
         "withheld. On a 1 to 5 scale a single low score among four responses moves the mean by "
         "about 0.75, so thin-data rankings are noise and damage trust in the report."),
        ("Today is partial",
         "Surveys arrive after the call ends, so today's response rate is always understated "
         "and rises through the day. Never compare a partial day against a complete one."),
        ("Provenance",
         "Source extracts are never modified. Every export carries a manifest with the period, "
         "row count, rule version and the SHA-256 of the rulebook that produced it."),
    ]
    for r, (term, text) in enumerate(definitions, 4):
        ws.write(r, 0, term, book.body)
        ws.write(r, 1, text, book.wrap)
        ws.set_row(r, 44)
    ws.set_column(0, 0, 26)
    ws.set_column(1, 1, 118)

    book.wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build a PCS daily workbook from a WFMHub pcs_agent_day export.")
    parser.add_argument("export", type=Path, help="Exported pcs_agent_day CSV or XLSX")
    parser.add_argument("--out", type=Path, help="Output workbook (default: output/PCS_Daily_<date>.xlsx)")
    parser.add_argument("--min-responses", type=int, default=5,
                        help="Responses required before an agent is ranked (default 5)")
    parser.add_argument("--rollup-by", default="ops_manager,lob",
                        help="Comma-separated fields for the ROLLUP sheet")
    parser.add_argument("--scorecard-by", default="team_leader,agent_name",
                        help="Comma-separated fields for the TL_SCORECARD sheet")
    args = parser.parse_args()

    export = args.export.expanduser().resolve()
    if not export.is_file():
        die(f"Export not found: {export}")

    here = Path(__file__).resolve().parent
    out = (args.out or here / "output" /
           f"PCS_Daily_{datetime.now():%Y-%m-%d_%H%M}.xlsx").expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)

    stamp_file = export.parent / "last_refresh.txt"
    refresh_stamp = ""
    if stamp_file.is_file():
        first = stamp_file.read_text(encoding="utf-8", errors="replace").splitlines()
        if first:
            refresh_stamp = first[0].replace("Data as at :", "").strip()

    print(f"  Reading  {export}")
    rows = prepare(read_export(export))
    print(f"  Rows     {len(rows):,}")
    if not rows:
        die("Nothing left to report after cleaning.")

    build(rows, out, args.min_responses, export.name, refresh_stamp,
          [f.strip() for f in args.rollup_by.split(",") if f.strip()],
          [f.strip() for f in args.scorecard_by.split(",") if f.strip()])
    print(f"  Written  {out}")
    print("\n  Reconcile the SUMMARY totals against the hub's own PCS workbook "
          "before the first send. They must tie exactly.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
