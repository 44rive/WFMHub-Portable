#!/usr/bin/env python3
"""Produce a pcs_agent_day export directly from a raw 'PCS Report.xlsx'.

Use this when the hub has not yet ingested the Call-by-Call extract but you need
the report now. It imports WFMHub's OWN parsing and scoping functions read-only,
so the output is identical to what `WFMHub.cmd export pcs_agent_day` produces.

Nothing in the hub is modified. The source workbook is only read.

    python prepare_from_pcs_report.py "TOLEARN/PCS Report.xlsx" \
        --hub ~/WFMHub-Portable --out data/current/pcs_agent_day.csv

Reads sheet RDATA (call legs) and sheet AGENT LIST (the FTE roster used for scope).
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

# --- Rulebook settings (config/wfm_rules.toml [pcs]) --------------------------
SCORED_QUESTIONS = [1, 2]
COMMENT_QUESTIONS = [3]
SURVEY_MODE = "2"
MIN_SCORE, MAX_SCORE = 1.0, 5.0
TOP_BOX_MIN, LOW_SCORE_MAX = 4.0, 2.0

OUT_COLUMNS = [
    "agent_day_key", "business_date", "agent_id", "agent_name", "team_leader",
    "ops_manager", "lob", "market", "language", "location", "call_legs",
    "handled_calls", "inbound_calls", "outbound_calls", "talk_seconds",
    "hold_seconds", "wrap_seconds", "handle_seconds", "average_talk_seconds",
    "average_hold_seconds", "average_wrap_seconds", "average_handle_seconds",
    "pcs_enabled_calls", "survey_responses", "response_rate",
    "q1_response_count", "q1_score_sum", "q1_average",
    "q2_response_count", "q2_score_sum", "q2_average",
    "pcs_score_count", "pcs_score_sum", "pcs_average",
    "top_box_responses", "low_score_responses", "top_box_percent",
    "low_score_percent", "comments_count",
]


def load_hub(hub: Path):
    """Import the hub's real parsing code. Read-only; nothing is executed against it."""
    src = hub / "src"
    if not (src / "wfmhub" / "ingestion.py").is_file():
        sys.exit(f"ERROR: no wfmhub source under {src}. Pass --hub <WFMHub-Portable folder>.")
    sys.path.insert(0, str(src))
    from wfmhub.ingestion import AgentScope, _call_record, _normalize_agent_name
    from wfmhub.utils import normalize_id
    return AgentScope, _call_record, _normalize_agent_name, normalize_id


SITE_RE = re.compile(r"^(AP[A-Z]{2})_([A-Z]{3})_")

DURATION_FIELDS = ("Talk Time", "Hold Time", "Total Wrap Time",
                   "Total Queue Wait Time", "Ringing Duration",
                   "Call Duration Until Conferenced")


def excel_duration_to_seconds(value):
    """Excel stores a duration cell as a fraction of a day. openpyxl hands that
    back as a float, and 0.0010416 is 90 seconds, not 0 seconds.

    The hub's own duration_seconds() rounds a bare number straight to an int,
    which is correct for the CSV extract it normally reads (where durations
    arrive as 'HH:MM:SS' text) but silently produces 0 for every call when the
    source is this XLSX. Convert here, before the hub's parser sees it.
    """
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        seconds = float(value)
        if seconds == 0:
            return 0
        # A call duration can never be a whole day, so a value below 1 is always
        # a day fraction. Anything larger is already expressed in seconds.
        return round(seconds * 86400) if 0 < seconds < 1 else round(seconds)
    return value


def site_of(service, queue):
    """Storm service/queue names start AP<country>_<city>_. That prefix is the site."""
    for value in (service, queue):
        match = SITE_RE.match(str(value or "").strip().upper())
        if match:
            return f"{match.group(1)}/{match.group(2)}", match.group(2)
    return None, None


def repair_mojibake(text):
    """Recover text that was UTF-8 but got decoded as Latin-1 upstream.

    The source workbook already contains 'Lara LourenAo'-style damage: it arrives
    that way from Storm, so the defect is not introduced here. The transformation
    is exactly reversible, and it is only applied when the round trip succeeds
    cleanly, so text that was never damaged is left untouched.

    Repairing matters because these names go in front of Team Leaders.
    """
    if not text or not isinstance(text, str):
        return text, False
    if not any(marker in text for marker in ("Ã", "Å", "Â", "â€", "Ä", "Ð")):
        return text, False
    # cp1252 first: Windows is what mangled it, and cp1252 maps 0x80-0x9F to
    # characters that latin-1 leaves undefined. Polish names such as
    # 'SoAowiow' only round-trip through cp1252.
    for codec in ("cp1252", "latin-1"):
        try:
            fixed = text.encode(codec).decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            continue
        if fixed != text and "\ufffd" not in fixed:
            return fixed, True
    return text, False


def commonest(counter):
    return counter.most_common(1)[0][0] if counter else None


def ratio(numerator, denominator):
    """Blank, not zero, when the denominator does not exist."""
    return numerator / denominator if denominator else None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("source", type=Path, help="PCS Report.xlsx")
    ap.add_argument("--hub", type=Path, default=Path.home() / "WFMHub-Portable")
    ap.add_argument("--out", type=Path, default=Path("data/current/pcs_agent_day.csv"))
    ap.add_argument("--calls-sheet", default="RDATA")
    ap.add_argument("--roster-sheet", default="AGENT LIST")
    ap.add_argument("--keep-mojibake", action="store_true",
                    help="Do not repair upstream UTF-8/Latin-1 name damage.")
    ap.add_argument("--scope", choices=("roster", "none"), default="roster",
                    help="roster = FTE-scoped exactly as the hub does. "
                         "none = keep every leg, deriving dimensions from the call itself. "
                         "Use 'none' when the roster does not yet cover this extract.")
    args = ap.parse_args()

    AgentScope, _call_record, _normalize_agent_name, normalize_id = load_hub(args.hub.expanduser())

    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl is required to read the source workbook.")

    source = args.source.expanduser().resolve()
    if not source.is_file():
        sys.exit(f"ERROR: source not found: {source}")

    print(f"  Reading {source.name}")
    wb = load_workbook(source, read_only=True, data_only=True)

    # ---- 1. Roster -> AgentScope, exactly as _load_agent_scope builds it ------
    ws = wb[args.roster_sheet]
    it = ws.iter_rows(values_only=True)
    rhdr = [str(c).strip() if c else "" for c in next(it)]
    roster = [dict(zip(rhdr, row)) for row in it if row and row[0]]

    def pick(record, *names):
        for name in names:
            for key in record:
                if key.strip().casefold() == name.casefold():
                    return record[key]
        return None

    ids, names_map, profile = set(), defaultdict(set), {}
    for record in roster:
        agent_id = normalize_id(pick(record, "Client ID", "Agent ID"))
        if not agent_id:
            continue
        ids.add(agent_id)
        name = pick(record, "Name", "Agent Name")
        key = _normalize_agent_name(name)
        if key:
            names_map[key].add(agent_id)
        profile[agent_id] = {
            "agent_name": str(name).strip() if name else None,
            "team_leader": (str(pick(record, "Team leader", "Team Leader") or "").strip() or None),
            "ops_manager": (str(pick(record, "Ops Manager") or "").strip() or None),
            "lob": (str(pick(record, "LOB") or "").strip() or None),
            "market": None, "language": None, "location": None,
        }
    unique = {k: next(iter(v)) for k, v in names_map.items() if len(v) == 1}
    scope = AgentScope(frozenset(ids), unique, hashlib.sha256(b"local").hexdigest())
    print(f"  Roster  {len(ids)} agents ({len(unique)} unique-name fallbacks)")
    if args.scope == "none":
        print("  Scope   DISABLED - keeping every leg; dimensions come from the call record")
        scope = None

    # ---- 2. Parse call legs with the hub's own _call_record -------------------
    ws = wb[args.calls_sheet]
    it = ws.iter_rows(values_only=True)
    raw_headers = list(next(it))
    header_map = {str(h or "").strip().strip("[]"): h for h in raw_headers}

    records, scoped_out, rejected, total = [], 0, 0, 0
    repaired_names = set()
    for source_row, row in enumerate(it, 2):
        if row is None or all(v is None for v in row):
            continue
        total += 1
        values = dict(zip(raw_headers, row))
        for field in DURATION_FIELDS:
            key = header_map.get(field)
            if key is not None:
                values[key] = excel_duration_to_seconds(values.get(key))
        record, reason = _call_record(values, header_map, "local", source_row, scope)
        if reason == "outside roster":
            scoped_out += 1
        elif reason:
            rejected += 1
        elif record:
            if not args.keep_mojibake:
                for field in ("agent_name", "lob", "queue", "service"):
                    fixed, changed = repair_mojibake(record.get(field))
                    if changed:
                        if field == "agent_name":
                            repaired_names.add(record[field])
                        record[field] = fixed
            records.append(record)
    wb.close()

    print(f"  Legs    {total:,} read | {len(records):,} in scope | "
          f"{scoped_out:,} outside roster | {rejected:,} rejected")
    if repaired_names:
        print(f"  Encoding {len(repaired_names)} damaged name(s) repaired "
              f"(UTF-8 mis-decoded as Latin-1 UPSTREAM, in the source workbook)")
        for name in sorted(repaired_names)[:3]:
            fixed, _ = repair_mojibake(name)
            print(f"           {name!r} -> {fixed!r}")

    # ---- 3. Deduplicate on call_key, newest source_row wins -------------------
    best = {}
    for record in records:
        key = record["call_key"]
        if key not in best or record["source_row"] > best[key]["source_row"]:
            best[key] = record
    legs = list(best.values())
    print(f"  Dedup   {len(legs):,} distinct call legs "
          f"({len(records) - len(legs):,} duplicate keys collapsed)")

    # ---- 4. Aggregate to agent/day, mirroring _build_pcs ----------------------
    def valid(score):
        return score is not None and MIN_SCORE <= score <= MAX_SCORE

    groups = defaultdict(list)
    for leg in legs:
        if leg.get("agent_id") and leg.get("business_date"):
            groups[(leg["business_date"], leg["agent_id"])].append(leg)

    out_rows = []
    for (business_date, agent_id), group in sorted(groups.items()):
        agg = defaultdict(float)
        names = defaultdict(set)
        dims = defaultdict(Counter)
        for leg in group:
            talk = leg.get("talk_seconds") or 0
            hold = leg.get("hold_seconds") or 0
            wrap = leg.get("wrap_seconds") or 0
            handle = talk + hold + wrap
            direction = str(leg.get("call_direction") or "").upper()
            eligible = direction == "I" and str(leg.get("post_call_survey_mode") or "") == SURVEY_MODE

            agg["call_legs"] += 1
            agg["handled_calls"] += 1 if handle > 0 else 0
            agg["inbound_calls"] += 1 if direction == "I" else 0
            agg["outbound_calls"] += 1 if direction == "O" else 0
            agg["talk_seconds"] += talk
            agg["hold_seconds"] += hold
            agg["wrap_seconds"] += wrap
            agg["handle_seconds"] += handle
            agg["pcs_enabled_calls"] += 1 if eligible else 0
            if leg.get("agent_name"):
                names["agent_name"].add(str(leg["agent_name"]).strip())
            if leg.get("lob"):
                dims["lob"][str(leg["lob"]).strip()] += 1
            if leg.get("language"):
                dims["language"][str(leg["language"]).strip()] += 1
            site, city = site_of(leg.get("service"), leg.get("queue"))
            if site:
                dims["site"][site] += 1
                dims["city"][city] += 1

            if not eligible:
                continue
            scores = [leg.get(f"question_{n}_score") for n in SCORED_QUESTIONS]
            good = [s for s in scores if valid(s)]
            if good:
                call_score = sum(good) / len(good)
                agg["survey_responses"] += 1
                agg["pcs_score_count"] += 1
                agg["pcs_score_sum"] += call_score
                agg["top_box_responses"] += 1 if call_score >= TOP_BOX_MIN else 0
                agg["low_score_responses"] += 1 if call_score <= LOW_SCORE_MAX else 0
            q1, q2 = leg.get("question_1_score"), leg.get("question_2_score")
            if valid(q1):
                agg["q1_response_count"] += 1
                agg["q1_score_sum"] += q1
            if valid(q2):
                agg["q2_response_count"] += 1
                agg["q2_score_sum"] += q2
            if any(str(leg.get(f"question_{n}") or "").strip() for n in COMMENT_QUESTIONS):
                agg["comments_count"] += 1

        person = profile.get(agent_id, {})
        handled = agg["handled_calls"]
        responses = agg["survey_responses"]
        out_rows.append({
            "agent_day_key": f"{business_date:%Y%m%d}-{agent_id}",
            "business_date": business_date.isoformat(),
            "agent_id": agent_id,
            "agent_name": person.get("agent_name") or next(iter(names["agent_name"]), None),
            "team_leader": person.get("team_leader"),
            "ops_manager": person.get("ops_manager"),
            "lob": person.get("lob") or commonest(dims["lob"]),
            "market": person.get("market") or commonest(dims["site"]),
            "language": person.get("language") or commonest(dims["language"]),
            "location": person.get("location") or commonest(dims["city"]),
            **{k: int(agg[k]) for k in (
                "call_legs", "handled_calls", "inbound_calls", "outbound_calls",
                "talk_seconds", "hold_seconds", "wrap_seconds", "handle_seconds",
                "pcs_enabled_calls", "survey_responses", "q1_response_count",
                "q2_response_count", "pcs_score_count", "top_box_responses",
                "low_score_responses", "comments_count")},
            "average_talk_seconds": ratio(agg["talk_seconds"], handled),
            "average_hold_seconds": ratio(agg["hold_seconds"], handled),
            "average_wrap_seconds": ratio(agg["wrap_seconds"], handled),
            "average_handle_seconds": ratio(agg["handle_seconds"], handled),
            "response_rate": ratio(responses, agg["pcs_enabled_calls"]),
            "q1_score_sum": agg["q1_score_sum"],
            "q1_average": ratio(agg["q1_score_sum"], agg["q1_response_count"]),
            "q2_score_sum": agg["q2_score_sum"],
            "q2_average": ratio(agg["q2_score_sum"], agg["q2_response_count"]),
            "pcs_score_sum": agg["pcs_score_sum"],
            "pcs_average": ratio(agg["pcs_score_sum"], agg["pcs_score_count"]),
            "top_box_percent": ratio(agg["top_box_responses"], responses),
            "low_score_percent": ratio(agg["low_score_responses"], responses),
        })

    out = args.out.expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUT_COLUMNS)
        writer.writeheader()
        for row in out_rows:
            writer.writerow({k: ("" if row.get(k) is None else row.get(k)) for k in OUT_COLUMNS})

    enabled = sum(r["pcs_enabled_calls"] for r in out_rows)
    responses = sum(r["survey_responses"] for r in out_rows)
    score_sum = sum(r["pcs_score_sum"] for r in out_rows)
    score_count = sum(r["pcs_score_count"] for r in out_rows)
    print(f"\n  Agent-days        {len(out_rows):,}")
    print(f"  PCS-enabled calls {enabled:,}")
    print(f"  Survey responses  {responses:,}")
    print(f"  Response rate     {responses/enabled:.2%}" if enabled else "  Response rate     n/a")
    print(f"  PCS average       {score_sum/score_count:.4f}" if score_count else "  PCS average       n/a")
    print(f"\n  Written {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
