"""Read-only source discovery, parsing and idempotent raw ingestion."""

from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from .config import Config
from .database import DatabaseConnection
from .mapping import QueueMapping, load_queue_mapping
from .progress import ProgressCallback
from .utils import (
    classify_assignment,
    classify_event,
    classify_status,
    duration_seconds,
    file_sha256,
    normalize_header,
    normalize_id,
    parse_date,
    parse_datetime,
    parse_time,
    parse_verint_interval,
)


class SourceSchemaError(RuntimeError):
    pass


@dataclass
class ParseResult:
    tables: dict[str, list[dict[str, Any]]] = field(default_factory=dict)
    rejected: list[str] = field(default_factory=list)
    scoped_out: int = 0
    source_variant: str | None = None

    @property
    def row_count(self) -> int:
        primary = [name for name in self.tables if name != "raw.schedule_event"]
        return sum(len(self.tables[name]) for name in primary)


@dataclass(frozen=True)
class SourceCandidate:
    family: str
    path: Path


@dataclass
class IngestSummary:
    loaded: int = 0
    skipped: int = 0
    failed: int = 0
    rows: int = 0
    scoped_out: int = 0
    errors: list[str] = field(default_factory=list)


def _normalize_agent_name(value: Any) -> str | None:
    """Return a conservative, accent-insensitive key for exact name fallback."""
    text = str(value or "").strip()
    if not text:
        return None
    decomposed = unicodedata.normalize("NFKD", text).casefold()
    words = re.findall(r"[a-z0-9]+", "".join(char for char in decomposed if not unicodedata.combining(char)))
    return " ".join(words) or None


@dataclass(frozen=True)
class AgentScope:
    """Date-aware agent scope derived from Active and dated Leaver FTE rows."""

    agent_ids: frozenset[str]
    unique_names: dict[str, str]
    fingerprint: str
    eligibility: dict[str, tuple[str, date | None]] = field(default_factory=dict)

    def active_on(self, roster_id: str, business_date: date | None) -> bool:
        rule = self.eligibility.get(roster_id)
        if rule is None:
            # Backwards-compatible programmatic scopes used by parser clients.
            return roster_id in self.agent_ids
        status, leave_date = rule
        normalized_status = " ".join(status.upper().replace("_", " ").split())
        if normalized_status == "ACTIVE":
            return True
        return (
            normalized_status == "LEAVER"
            and leave_date is not None
            and business_date is not None
            and business_date <= leave_date
        )

    def resolve(
        self,
        agent_id: Any,
        agent_name: Any,
        business_date: date | None = None,
    ) -> str | None:
        normalized_id = normalize_id(agent_id)
        if normalized_id in self.agent_ids and self.active_on(normalized_id, business_date):
            return normalized_id
        name_key = _normalize_agent_name(agent_name)
        roster_id = self.unique_names.get(name_key) if name_key else None
        if roster_id is None or not self.active_on(roster_id, business_date):
            return None
        # Preserve a populated operational source ID (especially Verint Data
        # Source IDs) while using the unique roster name only as the scope gate.
        return normalized_id or roster_id


AGENT_SCOPED_FAMILIES = {"schedule", "lilo", "agent_status", "calls"}
AGENT_SCOPE_POLICY_VERSION = "v3-active-or-leaver-id-or-unique-name"
SCHEDULE_PARSER_POLICY_VERSION = "v2-explicit-start-end-vs-activities"
FTE_PARSER_POLICY_VERSION = "v2-time-off-registers"
CALL_PARSER_POLICY_VERSION = "v4-active-roster-or-mapped-queue"


FILENAME_DATE_RE = re.compile(r"(?<!\d)(\d{4}-\d{2}-\d{2})(?!\d)")


TABLE_COLUMNS: dict[str, list[str]] = {
    "raw.fte_agent": ["source_file_id", "source_row", "agent_id", "employment_status", "agent_name", "team_leader", "ops_manager", "lob", "market", "language", "location", "city", "fte", "end_date"],
    "raw.fte_time_off": [
        "source_file_id", "source_sheet", "source_row", "source_kind",
        "agent_id", "agent_name", "start_date", "end_date", "day_coverage",
        "start_time", "end_time", "absence_type", "record_status", "comment",
    ],
    "raw.schedule_shift": ["source_file_id", "source_row", "schedule_date", "agent_id_raw", "agent_id", "agent_name", "scheduling_period", "shift_assignment", "assignment", "assignment_type", "scheduled_start", "scheduled_end", "shift_events", "parse_ok"],
    "raw.schedule_event": ["source_file_id", "source_row", "event_index", "schedule_date", "agent_id", "agent_name", "activity", "activity_type", "event_start", "event_end", "parse_ok"],
    "raw.lilo": ["source_file_id", "source_row", "extract_date", "agent_id", "agent_name", "first_login", "raw_last_logout", "last_logout", "overnight_adjusted"],
    "raw.agent_status": ["source_file_id", "source_row", "serial_number", "extract_date", "agent_id", "agent_name", "status", "actual_category", "status_start", "status_end", "duration_seconds", "queue"],
    "raw.forecast_interval": ["source_file_id", "source_row", "queue_name", "business_date", "interval_time", "interval_minutes", "interval_start", "volume_forecast", "abandons_forecast", "sl_forecast", "sl_required", "aht_forecast_seconds", "headcount_forecast", "net_staffing_forecast", "fte_forecast", "fte_required"],
    "raw.queue_actual": ["source_file_id", "source_row", "source_system", "business_date", "interval_time", "interval_start", "hour_start", "language", "queue_id", "queue", "business_partner", "lob", "offered", "answered", "abandoned", "short_calls", "answered_15s", "answered_20s", "answered_30s", "asa_seconds", "aht_seconds", "abandoned_20s"],
    "raw.call_leg": [
        "source_file_id", "source_row", "call_key", "interaction_key", "business_date",
        "call_start", "call_end", "communication_type", "call_direction",
        "originating_address", "business_partner_id", "lob", "destination_address",
        "service", "call_reference_number", "call_id", "call_progress",
        "queue_wait_seconds", "queue_id", "queue", "call_treatment_id",
        "call_treatment", "agent_group_id", "agent_group", "called_user_group",
        "agent_id", "agent_name", "clearing_party", "talk_seconds", "hold_seconds",
        "wrap_seconds", "completion_code", "transferred", "conference_at",
        "conference_duration_seconds", "shared_call_reference", "ringing_seconds",
        "internal", "direct", "menu_progress", "recording_mode", "recording_consent",
        "language_ivr", "language", "voicebot_id", "voicebot_destination",
        "voicebot_percentage_split", "voicebot_name", "routing_intent", "agent_intent",
        "licence_plate", "ani", "post_call_survey_mode", "pcs_status",
        "question_1", "question_2", "question_3", "question_4", "question_5",
        "question_6", "question_7", "question_8", "question_9", "question_10",
        "question_1_score", "question_2_score", "question_3_score", "question_4_score",
        "question_5_score", "question_6_score", "question_7_score", "question_8_score",
        "question_9_score", "question_10_score", "session_identifier", "callback_id",
        "callback_at", "callback_status", "callback_offered",
    ],
}


FTE_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "agent_id": ("CLIENTID", "AGENTID", "VERINTID", "DATASOURCEID", "DATASOURCEIDS"),
    "agent_name": ("NAME", "AGENTNAME", "AGENT", "EMPLOYEENAME", "FULLNAME"),
    "employment_status": ("STATUS", "EMPLOYMENTSTATUS", "AGENTSTATUS"),
    "team_leader": ("TEAMLEADER", "TEAMLEADERNAME", "TL"),
    "ops_manager": ("OPSMANAGER", "OPERATIONSMANAGER", "OPERATIONSMANAGERNAME"),
    "lob": ("LOB", "LINEOFBUSINESS"),
    "market": ("MARKET",),
    "language": ("LANGUAGE",),
    "location": ("LOCATION", "SITE"),
    "city": ("CITY",),
    "fte": ("FTE", "FTECOUNT"),
    "end_date": ("ENDDATEIFLEAVER", "ENDDATE", "LEAVERENDDATE"),
}

FTE_EXACT_ROSTER_TITLES = {
    "AGENT", "AGENTS", "AGENTLIST", "AGENTROSTER", "FTE", "FTECOUNT",
    "FTEAGENT", "FTEAGENTS", "HEADCOUNT", "ROSTER",
}


def _clean(value: Any) -> str | None:
    text = str(value).strip() if value is not None else ""
    return text or None


def _number(value: Any) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _filename_date_bounds(name: str) -> tuple[date | None, date | None]:
    """Return optional start/end hints without making filenames authoritative."""
    values = [datetime.strptime(value, "%Y-%m-%d").date() for value in FILENAME_DATE_RE.findall(name)]
    if not values:
        return None, None
    return min(values), max(values)


def discover_sources(config: Config) -> list[SourceCandidate]:
    candidates: list[SourceCandidate] = []
    single = config.source_path("fte_file")
    if single.is_file() and not single.name.startswith("~$"):
        candidates.append(SourceCandidate("fte", single))
    specs = (
        ("schedule", "schedule_folder", ("*.txt",)),
        ("lilo", "lilo_folder", ("*.csv",)),
        ("agent_status", "agent_status_folder", ("*.csv",)),
        ("forecast", "forecast_folder", ("*.txt",)),
        ("apbe", "apbe_folder", ("*.xlsx", "*.csv")),
        ("apfr", "apfr_folder", ("*.xlsx", "*.csv")),
        ("apde", "apde_folder", ("*.xlsx", "*.csv")),
        ("calls", "call_folder", ("*.csv",)),
    )
    for family, key, patterns in specs:
        if family in {"agent_status"} and not config.modules.get("agent_status", True):
            continue
        if family in {"forecast"} and not config.modules.get("forecast", True):
            continue
        if family in {"apbe", "apfr", "apde"} and not config.modules.get("intraday", True):
            continue
        if family == "calls" and not config.modules.get("pcs", True):
            continue
        folder = config.source_path(key)
        if not folder.is_dir():
            continue
        for pattern in patterns:
            for path in sorted(folder.glob(pattern)):
                if path.is_file() and not path.name.startswith("~$"):
                    candidates.append(SourceCandidate(family, path.resolve()))
    return candidates


def _fte_column_indexes(
    headers: tuple[Any, ...] | list[Any],
) -> tuple[dict[str, int], dict[str, list[int]]]:
    normalized: dict[str, list[int]] = {}
    for index, value in enumerate(headers):
        if str(value or "").strip():
            normalized.setdefault(normalize_header(value), []).append(index)
    indexes: dict[str, int] = {}
    duplicates: dict[str, list[int]] = {}
    for field, aliases in FTE_HEADER_ALIASES.items():
        matches = sorted({index for alias in aliases for index in normalized.get(alias, [])})
        if len(matches) == 1:
            indexes[field] = matches[0]
        elif len(matches) > 1:
            duplicates[field] = matches
    return indexes, duplicates


def _find_fte_table(workbook, path: Path) -> tuple[str, int, Any, dict[str, int]]:
    def confidence_tier(sheet_title: str, recognized: set[str]) -> int:
        normalized_title = normalize_header(sheet_title)
        if normalized_title in FTE_EXACT_ROSTER_TITLES:
            return 3
        title_words = set(re.findall(r"[A-Z0-9]+", sheet_title.upper()))
        if title_words & {"ROSTER", "HEADCOUNT"}:
            return 2
        org_fields = {"lob", "market", "language", "location", "city"}
        has_leadership = bool(recognized & {"team_leader", "ops_manager"})
        if "employment_status" in recognized and has_leadership and len(recognized & org_fields) >= 2:
            return 1
        return 0

    candidates: list[tuple[int, Any, int, dict[str, int]]] = []
    ambiguous: list[str] = []
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        for header_row, values in enumerate(rows, 1):
            if header_row > 100:
                break
            indexes, duplicates = _fte_column_indexes(values)
            identity_present = "agent_id" in indexes or "agent_id" in duplicates
            name_present = "agent_name" in indexes or "agent_name" in duplicates
            recognized = set(indexes) | set(duplicates)
            tier = confidence_tier(sheet.title, recognized)
            if identity_present and name_present and duplicates and tier:
                columns = ", ".join(sorted(duplicates))
                ambiguous.append(f"{sheet.title!r} row {header_row} has multiple aliases for: {columns}")
                break
            if {"agent_id", "agent_name"} <= indexes.keys() and tier:
                has_data = any(
                    (
                        indexes["agent_id"] < len(row) and row[indexes["agent_id"]] not in (None, "")
                    )
                    and (
                        indexes["agent_name"] < len(row) and str(row[indexes["agent_name"]] or "").strip()
                    )
                    for row in rows
                )
                if has_data:
                    candidates.append((tier, sheet, header_row, indexes))
                break
    if ambiguous:
        raise SourceSchemaError(
            f"FTE workbook has ambiguous roster headers. {'; '.join(ambiguous)}. "
            "Keep only one ID and one name column in the authoritative roster table."
        )
    if candidates:
        highest_tier = max(candidate[0] for candidate in candidates)
        best = [candidate for candidate in candidates if candidate[0] == highest_tier]
        if len(best) > 1:
            locations = ", ".join(f"{item[1].title!r} row {item[2]}" for item in best)
            raise SourceSchemaError(
                f"FTE workbook has multiple equally likely agent tables: {locations}. "
                "Keep one authoritative roster table or rename its sheet to Agent/Roster/FTE."
            )
        _, sheet, header_row, indexes = best[0]
        rows = sheet.iter_rows(min_row=header_row + 1, values_only=True)
        return sheet.title, header_row, rows, indexes
    searched = ", ".join(workbook.sheetnames) or "(no worksheets)"
    raise SourceSchemaError(
        f"FTE agent table was not found in {path.name}. Looked in sheets: {searched}. "
        "Expected an Agent/FTE/Roster/Headcount sheet with Client ID/Agent ID and Name/Agent Name."
    )


def _fte_register_sheet(workbook, names: set[str], label: str):
    matches = [sheet for sheet in workbook.worksheets if normalize_header(sheet.title) in names]
    if len(matches) > 1:
        raise SourceSchemaError(
            f"FTE workbook contains multiple {label} sheets: "
            + ", ".join(repr(sheet.title) for sheet in matches)
        )
    return matches[0] if matches else None


def _register_header(sheet, required: set[str], label: str) -> tuple[int, dict[str, int], Any]:
    rows = sheet.iter_rows(values_only=True)
    for header_row, values in enumerate(rows, 1):
        if header_row > 25:
            break
        normalized = {
            normalize_header(value): index
            for index, value in enumerate(values)
            if str(value or "").strip()
        }
        if required <= normalized.keys():
            return header_row, normalized, rows
    raise SourceSchemaError(
        f"FTE {label} sheet {sheet.title!r} is missing required template headers. "
        "Copy the current standard FTE template and keep its headers unchanged."
    )


def _parse_fte_registers(workbook, file_id: str) -> tuple[list[dict[str, Any]], list[str]]:
    output: list[dict[str, Any]] = []
    rejected: list[str] = []

    pto = _fte_register_sheet(workbook, {"PTO", "PTOS"}, "PTO")
    if pto is not None:
        required = {
            "CLIENTID", "STARTDATE", "ENDDATE", "DAYCOVERAGE", "STARTTIME",
            "ENDTIME", "PTOTYPE", "APPROVALSTATUS",
        }
        header_row, indexes, rows = _register_header(pto, required, "PTO")
        for source_row, values in enumerate(rows, header_row + 1):
            get = lambda header: values[indexes[header]] if indexes[header] < len(values) else None
            if not any(value not in (None, "") for value in values):
                continue
            agent_id = normalize_id(get("CLIENTID"))
            start_date = parse_date(get("STARTDATE"))
            end_date = parse_date(get("ENDDATE"))
            coverage_text = " ".join(str(get("DAYCOVERAGE") or "").strip().upper().replace("_", " ").split())
            coverage = {"FULL DAY": "FULL_DAY", "PARTIAL DAY": "PARTIAL_DAY"}.get(coverage_text)
            start_time = parse_time(get("STARTTIME"))
            end_time = parse_time(get("ENDTIME"))
            absence_type = _clean(get("PTOTYPE"))
            status = str(get("APPROVALSTATUS") or "").strip().upper()
            problems = []
            if not agent_id:
                problems.append("Client ID is required")
            if start_date is None or end_date is None:
                problems.append("valid Start date and End date are required")
            elif end_date < start_date:
                problems.append("End date is before Start date")
            if coverage is None:
                problems.append("Day coverage must be Full day or Partial day")
            if coverage == "PARTIAL_DAY":
                if start_date is not None and end_date is not None and start_date != end_date:
                    problems.append("Partial day PTO must use the same Start date and End date")
                if start_time is None or end_time is None or end_time <= start_time:
                    problems.append("Partial day PTO requires Start time before End time")
            if not absence_type:
                problems.append("PTO type is required")
            if status not in {"APPROVED", "PENDING", "CANCELLED"}:
                problems.append("Approval status must be Approved, Pending or Cancelled")
            if problems:
                rejected.append(f"{pto.title} row {source_row}: {'; '.join(problems)}")
                continue
            output.append({
                "source_file_id": file_id, "source_sheet": pto.title,
                "source_row": source_row, "source_kind": "PTO",
                "agent_id": agent_id,
                "agent_name": _clean(values[indexes["NAME"]]) if "NAME" in indexes and indexes["NAME"] < len(values) else None,
                "start_date": start_date, "end_date": end_date,
                "day_coverage": coverage,
                "start_time": start_time if coverage == "PARTIAL_DAY" else None,
                "end_time": end_time if coverage == "PARTIAL_DAY" else None,
                "absence_type": absence_type, "record_status": status,
                "comment": _clean(values[indexes["COMMENT"]]) if "COMMENT" in indexes and indexes["COMMENT"] < len(values) else None,
            })

    away = _fte_register_sheet(
        workbook, {"AWAY", "AWAYPEOPLE", "LONGABSENCE", "LONGABSENCES"}, "Away",
    )
    if away is not None:
        required = {"CLIENTID", "STARTDATE", "ENDDATE", "AWAYTYPE", "CASESTATUS"}
        header_row, indexes, rows = _register_header(away, required, "Away")
        for source_row, values in enumerate(rows, header_row + 1):
            get = lambda header: values[indexes[header]] if indexes[header] < len(values) else None
            if not any(value not in (None, "") for value in values):
                continue
            agent_id = normalize_id(get("CLIENTID"))
            start_date = parse_date(get("STARTDATE"))
            end_date = parse_date(get("ENDDATE"))
            absence_type = _clean(get("AWAYTYPE"))
            status = str(get("CASESTATUS") or "").strip().upper()
            problems = []
            if not agent_id:
                problems.append("Client ID is required")
            if start_date is None:
                problems.append("a valid Start date is required")
            if start_date is not None and end_date is not None and end_date < start_date:
                problems.append("End date is before Start date")
            if not absence_type:
                problems.append("Away type is required")
            if status not in {"ACTIVE", "PLANNED", "CLOSED", "CANCELLED"}:
                problems.append("Case status must be Active, Planned, Closed or Cancelled")
            if end_date is None and status != "ACTIVE":
                problems.append("only an Active case may have a blank End date")
            if problems:
                rejected.append(f"{away.title} row {source_row}: {'; '.join(problems)}")
                continue
            output.append({
                "source_file_id": file_id, "source_sheet": away.title,
                "source_row": source_row, "source_kind": "AWAY",
                "agent_id": agent_id,
                "agent_name": _clean(values[indexes["NAME"]]) if "NAME" in indexes and indexes["NAME"] < len(values) else None,
                "start_date": start_date, "end_date": end_date,
                "day_coverage": "FULL_DAY", "start_time": None, "end_time": None,
                "absence_type": absence_type, "record_status": status,
                "comment": _clean(values[indexes["COMMENT"]]) if "COMMENT" in indexes and indexes["COMMENT"] < len(values) else None,
            })
    return output, rejected


def parse_fte(path: Path, file_id: str) -> ParseResult:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        sheet_name, header_row, rows, index = _find_fte_table(workbook, path)
        output: list[dict[str, Any]] = []
        for source_row, values in enumerate(rows, header_row + 1):
            get = lambda field: values[index[field]] if field in index and index[field] < len(values) else None
            if get("agent_id") is None and not _clean(get("agent_name")):
                continue
            output.append({
                "source_file_id": file_id,
                "source_row": source_row,
                "agent_id": normalize_id(get("agent_id")),
                "employment_status": _clean(get("employment_status")),
                "agent_name": _clean(get("agent_name")),
                "team_leader": _clean(get("team_leader")),
                "ops_manager": _clean(get("ops_manager")),
                "lob": _clean(get("lob")),
                "market": _clean(get("market")),
                "language": _clean(get("language")),
                "location": _clean(get("location")),
                "city": _clean(get("city")),
                "fte": _number(get("fte")),
                "end_date": parse_date(get("end_date")),
            })
        if not output:
            raise SourceSchemaError(
                f"FTE agent table on sheet {sheet_name!r} has headers but no populated agent rows: {path.name}"
            )
        time_off, rejected = _parse_fte_registers(workbook, file_id)
        return ParseResult({"raw.fte_agent": output, "raw.fte_time_off": time_off}, rejected)
    finally:
        workbook.close()


def _parse_start_end_schedule(path: Path, file_id: str, scope: AgentScope | None = None) -> ParseResult:
    """Normalize Verint's wide one-column-per-day StartEndTimes extract."""
    shifts: list[dict[str, Any]] = []
    rejected: list[str] = []
    scoped_out = 0
    with path.open("r", encoding="cp1252", newline="") as handle:
        reader = csv.reader(handle, delimiter="\t")
        try:
            headers = next(reader)
        except StopIteration as exc:
            raise SourceSchemaError("StartEndTimes extract is empty") from exc
        normalized = [str(value or "").strip() for value in headers]
        if len(normalized) < 3 or normalized[:2] != ["Name", "Data Source IDs"]:
            raise SourceSchemaError("StartEndTimes must begin with Name and Data Source IDs")
        date_columns = [(index, parse_date(header)) for index, header in enumerate(normalized[2:], 2)]
        date_columns = [(index, value) for index, value in date_columns if value is not None]
        if not date_columns:
            raise SourceSchemaError("StartEndTimes contains no date columns")
        for physical_row, values in enumerate(reader, 2):
            name = _clean(values[0] if values else None)
            raw_id = normalize_id(values[1] if len(values) > 1 else None, reject_placeholders=False)
            source_agent_id = normalize_id(values[1] if len(values) > 1 else None)
            if not name and raw_id is None:
                continue
            for ordinal, (column, business_date) in enumerate(date_columns, 1):
                raw_assignment = _clean(values[column] if column < len(values) else None)
                if not raw_assignment:
                    continue
                agent_id = source_agent_id
                if scope is not None:
                    resolved_id = scope.resolve(agent_id, name, business_date)
                    if resolved_id is None:
                        scoped_out += 1
                        continue
                    agent_id = resolved_id
                assignment, start, end = parse_verint_interval(raw_assignment)
                is_off = raw_assignment.strip().upper() == "OFF"
                parse_ok = is_off or bool(start and end and end > start)
                shifts.append({
                    "source_file_id": file_id,
                    "source_row": physical_row * 1000 + ordinal,
                    "schedule_date": start.date() if start else business_date,
                    "agent_id_raw": raw_id,
                    "agent_id": agent_id,
                    "agent_name": name,
                    "scheduling_period": f"{date_columns[0][1]:%m/%d/%Y} to {date_columns[-1][1]:%m/%d/%Y}",
                    "shift_assignment": raw_assignment,
                    "assignment": assignment or raw_assignment,
                    "assignment_type": classify_assignment(raw_assignment, assignment),
                    "scheduled_start": start,
                    "scheduled_end": end,
                    "shift_events": None,
                    "parse_ok": parse_ok,
                })
                if not parse_ok:
                    rejected.append(f"row {physical_row}, {business_date:%Y-%m-%d}: shift interval could not be parsed")
    return ParseResult({"raw.schedule_shift": shifts, "raw.schedule_event": []}, rejected, scoped_out)


def _schedule_variant(path: Path) -> str:
    with path.open("r", encoding="cp1252", newline="") as probe:
        headers = next(csv.reader(probe, delimiter="\t"), [])
    stripped = [str(value or "").strip() for value in headers]
    activities = {"Name", "Data Source IDs", "Scheduling Period", "Shift Assignment", "Shift Events"}
    if activities <= set(stripped):
        return "ACTIVITIES"
    date_headers = [value for value in stripped[2:] if value]
    if (
        stripped[:2] == ["Name", "Data Source IDs"]
        and date_headers
        and len(date_headers) == len(set(date_headers))
        and all(parse_date(value) is not None for value in date_headers)
    ):
        return "START_END"
    raise SourceSchemaError(
        "Schedule header is neither a StartEndTimes export nor an Activities export"
    )


def parse_schedule(path: Path, file_id: str, scope: AgentScope | None = None) -> ParseResult:
    variant = _schedule_variant(path)
    if variant == "START_END":
        result = _parse_start_end_schedule(path, file_id, scope)
        result.source_variant = variant
        return result
    shifts: list[dict[str, Any]] = []
    events: list[dict[str, Any]] = []
    rejected: list[str] = []
    scoped_out = 0
    marker = None
    with path.open("r", encoding="cp1252", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        required = {"Name", "Data Source IDs", "Scheduling Period", "Shift Assignment", "Shift Events"}
        missing = sorted(required - set(reader.fieldnames or []))
        if missing:
            raise SourceSchemaError(f"Schedule missing columns: {', '.join(missing)}")
        for source_row, row in enumerate(reader, 2):
            name = _clean(row.get("Name"))
            raw_id = normalize_id(row.get("Data Source IDs"), reject_placeholders=False)
            agent_id = normalize_id(row.get("Data Source IDs"))
            if name and raw_id is None:
                possible = parse_date(name)
                if possible:
                    marker = possible
                    continue
            if not name and raw_id is None:
                continue
            if marker is None:
                raise SourceSchemaError(f"Agent row before date marker at line {source_row}")
            raw_assignment = _clean(row.get("Shift Assignment"))
            assignment, start, end = parse_verint_interval(raw_assignment)
            is_off = (raw_assignment or "").strip().upper() == "OFF"
            schedule_date = start.date() if start else marker
            if scope is not None:
                resolved_id = scope.resolve(agent_id, name, schedule_date)
                if resolved_id is None:
                    scoped_out += 1
                    continue
                agent_id = resolved_id
            parse_ok = is_off or bool(start and end and end > start)
            shifts.append({
                "source_file_id": file_id, "source_row": source_row,
                "schedule_date": schedule_date, "agent_id_raw": raw_id,
                "agent_id": agent_id, "agent_name": name,
                "scheduling_period": _clean(row.get("Scheduling Period")),
                "shift_assignment": raw_assignment, "assignment": assignment or raw_assignment,
                "assignment_type": classify_assignment(raw_assignment, assignment),
                "scheduled_start": start, "scheduled_end": end,
                "shift_events": _clean(row.get("Shift Events")), "parse_ok": parse_ok,
            })
            if not parse_ok:
                rejected.append(f"line {source_row}: shift interval could not be parsed")
            for event_index, part in enumerate(str(row.get("Shift Events") or "").split(";"), 1):
                if not part.strip():
                    continue
                activity, event_start, event_end = parse_verint_interval(part)
                event_ok = bool(event_start and event_end and event_end > event_start)
                events.append({
                    "source_file_id": file_id, "source_row": source_row, "event_index": event_index,
                    "schedule_date": schedule_date, "agent_id": agent_id, "agent_name": name,
                    "activity": activity, "activity_type": classify_event(activity),
                    "event_start": event_start, "event_end": event_end, "parse_ok": event_ok,
                })
                if not event_ok:
                    rejected.append(f"line {source_row} event {event_index}: interval could not be parsed")
    return ParseResult(
        {"raw.schedule_shift": shifts, "raw.schedule_event": events},
        rejected, scoped_out, variant,
    )


def parse_lilo(path: Path, file_id: str, scope: AgentScope | None = None) -> ParseResult:
    date_field, fallback_date = _validate_lilo_header(path)
    output: list[dict[str, Any]] = []
    rejected: list[str] = []
    scoped_out = 0
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {"[Agent]", "[Agent ID]", "[First Log-on Time]", "[Last Log-off Time]"}
        missing = sorted(required - set(reader.fieldnames or []))
        if missing:
            raise SourceSchemaError(f"LILO missing columns: {', '.join(missing)}")
        for source_row, row in enumerate(reader, 2):
            agent_id = normalize_id(row.get("[Agent ID]"))
            first = parse_datetime(row.get("[First Log-on Time]"))
            raw_last = parse_datetime(row.get("[Last Log-off Time]"))
            extract_date = (
                parse_date(row.get(date_field)) if date_field else None
            ) or (first.date() if first else None) or (raw_last.date() if raw_last else None) or fallback_date
            if extract_date is None:
                rejected.append(
                    f"line {source_row}: no row date; a multi-day LILO row with blank login/logout needs a Date column"
                )
                continue
            if scope is None and not agent_id:
                continue
            if scope is not None:
                resolved_id = scope.resolve(agent_id, row.get("[Agent]"), extract_date)
                if resolved_id is None:
                    scoped_out += 1
                    continue
                agent_id = resolved_id
            last = raw_last
            adjusted = False
            if first and last and last < first:
                last += timedelta(days=1)
                adjusted = True
            output.append({
                "source_file_id": file_id, "source_row": source_row,
                "extract_date": extract_date, "agent_id": agent_id,
                "agent_name": _clean(row.get("[Agent]")), "first_login": first,
                "raw_last_logout": raw_last, "last_logout": last,
                "overnight_adjusted": adjusted,
            })
    return ParseResult({"raw.lilo": output}, rejected=rejected, scoped_out=scoped_out)


def _validate_lilo_header(path: Path) -> tuple[str | None, datetime.date | None]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        header_row = next(reader, [])
        headers = set(header_row)
    required = {"[Agent]", "[Agent ID]", "[First Log-on Time]", "[Last Log-off Time]"}
    missing = sorted(required - headers)
    if missing:
        raise SourceSchemaError(f"LILO missing columns: {', '.join(missing)}")
    normalized = {normalize_header(header): header for header in header_row}
    date_field = next(
        (normalized[key] for key in ("BUSINESSDATE", "EXTRACTDATE", "REPORTDATE", "DATE") if key in normalized),
        None,
    )
    filename_start, filename_end = _filename_date_bounds(path.name)
    fallback_date = filename_start if filename_start is not None and filename_start == filename_end else None
    return date_field, fallback_date


def _insert_lilo_direct(
    conn: DatabaseConnection,
    path: Path,
    file_id: str,
    date_field: str | None,
    fallback_date,
    scope: AgentScope,
    progress: ProgressCallback | None = None,
) -> tuple[int, int, int]:
    """Stream large LILO CSVs into bounded SQLite inserts."""
    count = 0
    scoped_out = 0
    rejected = 0
    processed = 0
    batch: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for source_row, row in enumerate(reader, 2):
            processed += 1
            if progress is not None and processed % 5000 == 0:
                progress(processed, 0, f"LILO: {processed:,} rows scanned")
            agent_id = normalize_id(row.get("[Agent ID]"))
            first = parse_datetime(row.get("[First Log-on Time]"))
            raw_last = parse_datetime(row.get("[Last Log-off Time]"))
            extract_date = (
                parse_date(row.get(date_field)) if date_field else None
            ) or (first.date() if first else None) or (raw_last.date() if raw_last else None) or fallback_date
            if extract_date is None:
                rejected += 1
                continue
            resolved_id = scope.resolve(agent_id, row.get("[Agent]"), extract_date)
            if resolved_id is None:
                scoped_out += 1
                continue
            agent_id = resolved_id
            last = raw_last
            adjusted = False
            if first and last and last < first:
                last += timedelta(days=1)
                adjusted = True
            batch.append({
                "source_file_id": file_id,
                "source_row": source_row,
                "extract_date": extract_date,
                "agent_id": agent_id,
                "agent_name": _clean(row.get("[Agent]")),
                "first_login": first,
                "raw_last_logout": raw_last,
                "last_logout": last,
                "overnight_adjusted": adjusted,
            })
            if len(batch) >= 5000:
                _insert_rows(conn, "raw.lilo", batch)
                count += len(batch)
                batch.clear()
    if batch:
        _insert_rows(conn, "raw.lilo", batch)
        count += len(batch)
    if progress is not None and processed:
        progress(processed, 0, f"LILO: {processed:,} rows scanned")
    return count, scoped_out, rejected


STATUS_REQUIRED_HEADERS = {
    "[Serial Number]", "[Status]", "[Status Start Date and Time]", "[Agent]",
    "[Agent ID]", "[Status Duration]", "[Queue]",
}


def _status_reader(path: Path):
    handle = path.open("r", encoding="utf-8-sig", newline="")
    reader = csv.DictReader(handle)
    missing = sorted(STATUS_REQUIRED_HEADERS - set(reader.fieldnames or []))
    if missing:
        handle.close()
        raise SourceSchemaError(f"Agent Status missing columns: {', '.join(missing)}")
    return handle, reader


def _status_record(
    row: dict[str, Any], file_id: str, source_row: int, scope: AgentScope | None,
) -> tuple[dict[str, Any] | None, str | None]:
    agent_id = normalize_id(row.get("[Agent ID]"))
    start = parse_datetime(row.get("[Status Start Date and Time]"))
    seconds = duration_seconds(row.get("[Status Duration]"))
    end = start + timedelta(seconds=seconds) if start is not None and seconds is not None else None
    if not start or not end or end <= start:
        return None, "invalid status interval"
    if scope is not None:
        resolved_id = scope.resolve(agent_id, row.get("[Agent]"), start.date())
        if resolved_id is None:
            return None, "outside roster"
        agent_id = resolved_id
    serial = _clean(row.get("[Serial Number]"))
    if not serial:
        serial = hashlib.sha256(repr(sorted(row.items())).encode("utf-8")).hexdigest()
    status = _clean(row.get("[Status]"))
    return {
        "source_file_id": file_id, "source_row": source_row,
        # A range filename is only a label; every row timestamp is authoritative.
        "serial_number": serial, "extract_date": start.date(), "agent_id": agent_id,
        "agent_name": _clean(row.get("[Agent]")), "status": status,
        "actual_category": classify_status(status), "status_start": start,
        "status_end": end, "duration_seconds": seconds, "queue": _clean(row.get("[Queue]")),
    }, None


def parse_agent_status(path: Path, file_id: str, scope: AgentScope | None = None) -> ParseResult:
    output: list[dict[str, Any]] = []
    rejected: list[str] = []
    scoped_out = 0
    handle, reader = _status_reader(path)
    try:
        for source_row, row in enumerate(reader, 2):
            record, reason = _status_record(row, file_id, source_row, scope)
            if reason == "outside roster":
                scoped_out += 1
            elif reason:
                rejected.append(f"line {source_row}: {reason}")
            elif record is not None:
                output.append(record)
    finally:
        handle.close()
    return ParseResult({"raw.agent_status": output}, rejected, scoped_out)


def _insert_status_direct(
    conn: DatabaseConnection,
    path: Path,
    file_id: str,
    scope: AgentScope,
    progress: ProgressCallback | None = None,
) -> tuple[int, int, int]:
    count = scoped_out = rejected = processed = 0
    batch: list[dict[str, Any]] = []
    handle, reader = _status_reader(path)
    try:
        for source_row, row in enumerate(reader, 2):
            processed += 1
            if progress is not None and processed % 5000 == 0:
                progress(processed, 0, f"Agent Status: {processed:,} rows scanned")
            record, reason = _status_record(row, file_id, source_row, scope)
            if reason == "outside roster":
                scoped_out += 1
                continue
            if reason:
                rejected += 1
                continue
            if record is not None:
                batch.append(record)
            if len(batch) >= 5000:
                _insert_rows(conn, "raw.agent_status", batch)
                count += len(batch)
                batch.clear()
    finally:
        handle.close()
    if batch:
        _insert_rows(conn, "raw.agent_status", batch)
        count += len(batch)
    if progress is not None and processed:
        progress(processed, 0, f"Agent Status: {processed:,} rows scanned")
    return count, scoped_out, rejected


CALL_REQUIRED_HEADERS = {
    "Call Date/Time", "Call ID", "Call Reference Number", "Agent ID", "Agent",
    "Talk Time", "Hold Time", "Total Wrap Time", "Call End Date/Time",
}


def _boolean(value: Any) -> bool | None:
    text = str(value or "").strip().upper()
    if text in {"1", "TRUE", "YES", "Y"}:
        return True
    if text in {"0", "FALSE", "NO", "N"}:
        return False
    return None


def _call_header_map(path: Path) -> dict[str, str]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        fieldnames = csv.DictReader(handle).fieldnames or []
    mapping = {str(header or "").strip().strip("[]"): header for header in fieldnames}
    missing = sorted(CALL_REQUIRED_HEADERS - set(mapping))
    if missing:
        raise SourceSchemaError(f"Call by Call missing columns: {', '.join(missing)}")
    return mapping


def _call_record(
    row: dict[str, Any],
    headers: dict[str, str],
    file_id: str,
    source_row: int,
    scope: AgentScope | None,
    queue_mapping: QueueMapping | None = None,
) -> tuple[dict[str, Any] | None, str | None]:
    get = lambda name: row.get(headers.get(name, ""))
    start = parse_datetime(get("Call Date/Time"))
    if start is None:
        return None, "invalid call date/time"
    source_agent_id = normalize_id(get("Agent ID"))
    agent_name = _clean(get("Agent"))
    agent_id = source_agent_id
    queue = _clean(get("Queue"))
    queue_is_mapped = (
        queue_mapping is not None
        and queue_mapping.map_actual("STORM", queue, None, None).status == "MAPPED"
    )
    if scope is not None:
        roster_agent_id = scope.resolve(source_agent_id, agent_name, start.date())
        if roster_agent_id is not None:
            agent_id = roster_agent_id
        elif not queue_is_mapped:
            return None, "outside roster"
        # Mapped queue interactions are required for service demand even when
        # they were abandoned (no Agent ID) or handled by another operation.
        # Agent-level marts still admit only rows joined to the governed FTE
        # dimension, so this does not expand PCS or attendance scope.
    elif agent_id is None and not queue_is_mapped:
        return None, "outside roster"
    end = parse_datetime(get("Call End Date/Time"))
    if end and end < start:
        end += timedelta(days=1)
    call_reference = _clean(get("Call Reference Number"))
    call_id = _clean(get("Call ID"))
    direction = _clean(get("Call Direction"))
    clearing_party = _clean(get("Clearing Party"))
    shared_reference = _clean(get("Shared Call Reference"))
    # End time and survey answers can mature in a later overlapping export, so
    # they are deliberately excluded from the stable leg identity.
    key_parts = [call_reference, call_id, direction, agent_id, start, clearing_party]
    if not any(key_parts[:2]):
        key_parts.append(json.dumps(row, sort_keys=True, ensure_ascii=False, default=str))
    call_key = hashlib.sha256("|".join(str(value or "") for value in key_parts).encode("utf-8")).hexdigest()
    interaction_key = shared_reference or call_reference or call_id or call_key
    questions = [_clean(get(f"Question {number}")) for number in range(1, 11)]
    record: dict[str, Any] = {
        "source_file_id": file_id,
        "source_row": source_row,
        "call_key": call_key,
        "interaction_key": interaction_key,
        "business_date": start.date(),
        "call_start": start,
        "call_end": end,
        "communication_type": _clean(get("Communication Type")),
        "call_direction": direction,
        "originating_address": _clean(get("Originating Address")),
        "business_partner_id": _clean(get("BusinessPartnerID")),
        "lob": _clean(get("LineOfBusiness")),
        "destination_address": _clean(get("Destination Address")),
        "service": _clean(get("Service")),
        "call_reference_number": call_reference,
        "call_id": call_id,
        "call_progress": _clean(get("CallProgress")),
        "queue_wait_seconds": duration_seconds(get("Total Queue Wait Time")),
        "queue_id": normalize_id(get("Queue ID"), reject_placeholders=False),
        "queue": queue,
        "call_treatment_id": normalize_id(get("Call Treatment ID"), reject_placeholders=False),
        "call_treatment": _clean(get("Call Treatment")),
        "agent_group_id": normalize_id(get("Agent Group ID"), reject_placeholders=False),
        "agent_group": _clean(get("Agent Group")),
        "called_user_group": _clean(get("Called User Group")),
        "agent_id": agent_id,
        "agent_name": agent_name,
        "clearing_party": clearing_party,
        "talk_seconds": duration_seconds(get("Talk Time")),
        "hold_seconds": duration_seconds(get("Hold Time")),
        "wrap_seconds": duration_seconds(get("Total Wrap Time")),
        "completion_code": _clean(get("Call Completion Code")),
        "transferred": _boolean(get("Transferred?")),
        "conference_at": parse_datetime(get("Date/Time Conferenced")),
        "conference_duration_seconds": duration_seconds(get("Call Duration Until Conferenced")),
        "shared_call_reference": shared_reference,
        "ringing_seconds": duration_seconds(get("Ringing Duration")),
        "internal": _boolean(get("Internal")),
        "direct": _boolean(get("Direct")),
        "menu_progress": _clean(get("MenuProgress")),
        "recording_mode": _clean(get("CallRecordingMode")),
        "recording_consent": _clean(get("CallRecordingConsent")),
        "language_ivr": _clean(get("LanguageIVR")),
        "language": _clean(get("Language")),
        "voicebot_id": normalize_id(get("VoicebotID"), reject_placeholders=False),
        "voicebot_destination": _clean(get("VoicebotDestination")),
        "voicebot_percentage_split": _number(get("VoicebotPercentageSplit")),
        "voicebot_name": _clean(get("VoicebotName")),
        "routing_intent": _clean(get("RoutingIntent")),
        "agent_intent": _clean(get("AgentIntent")),
        "licence_plate": _clean(get("LicencePlate")),
        "ani": _clean(get("ANI")),
        "post_call_survey_mode": _clean(get("PostCallSurveyMode")),
        "pcs_status": _clean(get("PCSStatus")),
        "session_identifier": _clean(get("SessionIdentifier")),
        "callback_id": _clean(get("CallbackID")),
        "callback_at": parse_datetime(get("CallbackDateTime")),
        "callback_status": _clean(get("CallbackStatus")),
        "callback_offered": _boolean(get("CallbackOffered")),
    }
    for number, value in enumerate(questions, 1):
        record[f"question_{number}"] = value
        record[f"question_{number}_score"] = _number(value)
    return record, None


def parse_calls(
    path: Path,
    file_id: str,
    scope: AgentScope | None = None,
    queue_mapping: QueueMapping | None = None,
) -> ParseResult:
    headers = _call_header_map(path)
    output: list[dict[str, Any]] = []
    rejected: list[str] = []
    scoped_out = 0
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for source_row, row in enumerate(csv.DictReader(handle), 2):
            record, reason = _call_record(
                row, headers, file_id, source_row, scope, queue_mapping,
            )
            if reason == "outside roster":
                scoped_out += 1
            elif reason:
                rejected.append(f"line {source_row}: {reason}")
            elif record is not None:
                output.append(record)
    return ParseResult({"raw.call_leg": output}, rejected, scoped_out)


def _insert_calls_direct(
    conn: DatabaseConnection,
    path: Path,
    file_id: str,
    scope: AgentScope,
    queue_mapping: QueueMapping,
    progress: ProgressCallback | None = None,
) -> tuple[int, int, int]:
    headers = _call_header_map(path)
    count = scoped_out = rejected = processed = 0
    batch: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for source_row, row in enumerate(csv.DictReader(handle), 2):
            processed += 1
            if progress is not None and processed % 2000 == 0:
                progress(processed, 0, f"Call by Call: {processed:,} rows scanned")
            record, reason = _call_record(
                row, headers, file_id, source_row, scope, queue_mapping,
            )
            if reason == "outside roster":
                scoped_out += 1
                continue
            if reason:
                rejected += 1
                continue
            if record is not None:
                batch.append(record)
            if len(batch) >= 2000:
                _insert_rows(conn, "raw.call_leg", batch)
                count += len(batch)
                batch.clear()
    if batch:
        _insert_rows(conn, "raw.call_leg", batch)
        count += len(batch)
    if progress is not None and processed:
        progress(processed, 0, f"Call by Call: {processed:,} rows scanned")
    return count, scoped_out, rejected


FORECAST_NAMES = {
    "volume_forecast": "Volume (Absolute For)",
    "abandons_forecast": "Abandons (Absolute For)",
    "sl_forecast": "Service Level (Absolute For)",
    "sl_required": "Service Level (Absolute Req)",
    "aht_forecast_seconds": "Activity Handling Time (Absolute For)",
    "headcount_forecast": "Headcount Staffing (Absolute For)",
    "net_staffing_forecast": "Net Staffing (Absolute For)",
    "fte_forecast": "Full Time Equivalents (Absolute For)",
    "fte_required": "Full Time Equivalents (Absolute Req)",
}


def parse_forecast(path: Path, file_id: str) -> ParseResult:
    with path.open("r", encoding="cp1252", newline="") as handle:
        lines = handle.readlines()
    header_index = next((i for i, line in enumerate(lines) if line.lstrip().startswith("Queue Name\tDate\tTime")), None)
    if header_index is None:
        raise SourceSchemaError("Forecast Queue Name / Date / Time header was not found")
    reader = csv.DictReader(lines[header_index:], delimiter="\t")
    # Some reviewed Verint exports intentionally contain volume only. The
    # remaining forecast measures are optional and stay NULL, never invented.
    required = {"Queue Name", "Date", "Time", "Time Interval", FORECAST_NAMES["volume_forecast"]}
    missing = sorted(required - set(reader.fieldnames or []))
    if missing:
        raise SourceSchemaError(f"Forecast missing columns: {', '.join(missing)}")
    output: list[dict[str, Any]] = []
    rejected: list[str] = []
    for source_row, row in enumerate(reader, header_index + 2):
        business_date = parse_date(row.get("Date"))
        interval_time = parse_time(row.get("Time"))
        interval_seconds = duration_seconds(row.get("Time Interval"))
        if business_date is None or interval_time is None:
            rejected.append(f"line {source_row}: invalid forecast date/time")
            continue
        result = {
            "source_file_id": file_id, "source_row": source_row,
            "queue_name": _clean(row.get("Queue Name")), "business_date": business_date,
            "interval_time": interval_time, "interval_minutes": int(interval_seconds / 60) if interval_seconds is not None else None,
            "interval_start": datetime.combine(business_date, interval_time),
        }
        for target, source in FORECAST_NAMES.items():
            value = _number(row.get(source))
            result[target] = value / 100 if target in {"sl_forecast", "sl_required"} and value is not None else value
        output.append(result)
    return ParseResult({"raw.forecast_interval": output}, rejected)


def _find_excel_header(workbook, path: Path) -> tuple[Any, list[str], Any]:
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        for source_row, values in enumerate(rows, 1):
            headers = [str(value or "").strip() for value in values]
            normalized = {normalize_header(value) for value in headers}
            if "DATE" in normalized and "15MINUTEPERIODSOFDAY" in normalized:
                return sheet, headers, (source_row, rows)
    raise SourceSchemaError(f"Could not find the queue-actual header row in {path.name}")


def _parse_queue_actual_rows(
    headers: list[str],
    rows,
    file_id: str,
    source_system: str,
) -> ParseResult:
        normalized = {normalize_header(name): index for index, name in enumerate(headers) if name}

        def require(*names: str) -> int:
            for name in names:
                key = normalize_header(name)
                if key in normalized:
                    return normalized[key]
            raise SourceSchemaError(f"{source_system} missing required column: {' / '.join(names)}")

        date_i = require("Date")
        time_i = require("15 Minute Periods of Day")
        partner_i = require("BusinessPartnerID")
        lob_i = require("LineOfBusiness")
        offered_i = require("Offered_calls (w/o short calls)", "APPELS ENTRANTS", "Offered Calls")
        answered_i = require("Answered_Calls", "APPELS RÉP", "APPELS REP", "Answered Calls")
        abandoned_i = require("Abandoned_Calls (w/o short calls)", "APPELS ABAN", "Abandoned Calls")
        ans15_i = require("Answered_Calls <= 15s", "APPELS RÉP <= 15s", "APPELS REP <= 15s", "Answered Calls <= 15s")
        ans20_i = require("Answered_Calls <= 20s", "APPELS RÉP <= 20s", "APPELS REP <= 20s", "Answered Calls <= 20s")
        ans30_i = require("Answered_Calls <= 30s", "APPELS RÉP <= 30s", "APPELS REP <= 30s", "Answered Calls <= 30s")
        abandoned20_i = next((normalized.get(normalize_header(name)) for name in (
            "Abandoned_Calls <= 20s", "Abandoned_Calls (w/o s.c.) <= 20s", "Abandoned Calls <= 20s",
            "APPELS ABAN <= 20s", "Abandoned <= 20s",
        ) if normalized.get(normalize_header(name)) is not None), None)
        short_i = require("Short_calls < 5s", "Short Calls < 5s")
        asa_i = require("Average_Speed_of_Answer", "Average Speed of Answer")
        talk_i = require("Average_Talk_Time", "Average Talk Time")
        hold_i = require("Average_Hold_Time", "Average Hold Time")
        wrap_i = require("Average Total Wrap Time", "Average_Total_Wrap_Time")
        handled_i = next((normalized.get(normalize_header(name)) for name in (
            "Average_Handled_Time", "Average Handled Time", "Average Handle Time",
        ) if normalized.get(normalize_header(name)) is not None), None)
        queue_i = normalized.get("QUEUE")
        queue_id_i = normalized.get("QUEUEID")
        language_i = normalized.get("LANGUAGE")
        output: list[dict[str, Any]] = []
        rejected: list[str] = []
        for source_row, values in rows:
            get = lambda index: values[index] if index is not None and index < len(values) else None
            business_date = parse_date(get(date_i))
            if business_date is None:
                if str(get(date_i) or "").strip().upper() not in {"", "TOTAL"}:
                    rejected.append(f"line {source_row}: invalid date")
                continue
            interval_time = parse_time(get(time_i))
            if interval_time is None:
                rejected.append(f"line {source_row}: invalid interval")
                continue
            talk = duration_seconds(get(talk_i))
            hold = duration_seconds(get(hold_i))
            wrap = duration_seconds(get(wrap_i))
            parts = [value for value in (talk, hold, wrap) if value is not None]
            partner = _clean(get(partner_i))
            queue = _clean(get(queue_i)) if queue_i is not None else partner
            interval_start = datetime.combine(business_date, interval_time)
            output.append({
                "source_file_id": file_id, "source_row": source_row, "source_system": source_system,
                "business_date": business_date, "interval_time": interval_time,
                "interval_start": interval_start, "hour_start": interval_start.replace(minute=0, second=0, microsecond=0),
                "language": _clean(get(language_i)) if language_i is not None else ({"APFR": "FR", "APDE": "DE"}.get(source_system)),
                "queue_id": normalize_id(get(queue_id_i), reject_placeholders=False) if queue_id_i is not None else None,
                "queue": queue, "business_partner": partner, "lob": _clean(get(lob_i)),
                "offered": _number(get(offered_i)), "answered": _number(get(answered_i)),
                "abandoned": _number(get(abandoned_i)), "short_calls": _number(get(short_i)),
                "answered_15s": _number(get(ans15_i)), "answered_20s": _number(get(ans20_i)),
                "answered_30s": _number(get(ans30_i)), "asa_seconds": duration_seconds(get(asa_i)),
                "aht_seconds": duration_seconds(get(handled_i)) if handled_i is not None else (sum(parts) if parts else None),
                "abandoned_20s": _number(get(abandoned20_i)) if abandoned20_i is not None else None,
            })
        return ParseResult({"raw.queue_actual": output}, rejected)


def parse_queue_actual(path: Path, file_id: str, source_system: str) -> ParseResult:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            try:
                headers = [str(value or "").strip() for value in next(reader)]
            except StopIteration as exc:
                raise SourceSchemaError(f"{source_system} file is empty") from exc
            return _parse_queue_actual_rows(headers, enumerate(reader, 2), file_id, source_system)
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        _, headers, state = _find_excel_header(workbook, path)
        header_row, rows = state
        return _parse_queue_actual_rows(headers, enumerate(rows, header_row + 1), file_id, source_system)
    finally:
        workbook.close()


PARSERS: dict[str, Callable[[Path, str, AgentScope | None], ParseResult]] = {
    "fte": lambda path, file_id, scope: parse_fte(path, file_id),
    "schedule": parse_schedule,
    "lilo": parse_lilo,
    "agent_status": parse_agent_status,
    "forecast": lambda path, file_id, scope: parse_forecast(path, file_id),
    "apbe": lambda path, file_id, scope: parse_queue_actual(path, file_id, "APBE"),
    "apfr": lambda path, file_id, scope: parse_queue_actual(path, file_id, "APFR"),
    "apde": lambda path, file_id, scope: parse_queue_actual(path, file_id, "APDE"),
    "calls": parse_calls,
}


def _load_agent_scope(conn: DatabaseConnection) -> AgentScope:
    rows = conn.execute(
        """SELECT r.source_file_id, r.source_row, r.agent_id, r.agent_name,
                  r.employment_status, r.end_date
           FROM raw.fte_agent r
           JOIN meta.source_file f ON f.file_id=r.source_file_id
           WHERE f.active=true AND f.status='SUCCESS'"""
    ).fetchall()
    ranked: dict[str, tuple[str, date | None, str | None]] = {}
    for source_file_id, source_row, agent_id, agent_name, status, end_date in rows:
        normalized_id = normalize_id(agent_id)
        name_key = _normalize_agent_name(agent_name)
        normalized_status = " ".join(str(status or "").strip().upper().replace("_", " ").split())
        parsed_end = parse_date(end_date)
        if (not normalized_id and not name_key) or normalized_status not in {"ACTIVE", "LEAVER"}:
            continue
        if normalized_status == "LEAVER" and parsed_end is None:
            continue
        roster_key = normalized_id or f"@NAME:{source_file_id}:{source_row}"
        candidate = (normalized_status, parsed_end, name_key)
        current = ranked.get(roster_key)
        if current is None or (
            (candidate[0] == "ACTIVE", candidate[1] or date.min)
            > (current[0] == "ACTIVE", current[1] or date.min)
        ):
            ranked[roster_key] = candidate
    ids = set(ranked)
    if not ids:
        raise SourceSchemaError(
            "Agent scope is empty. Load a valid FTE roster before agent-level extracts; "
            "no worldwide rows were admitted."
        )
    names: dict[str, set[str]] = {}
    for roster_key, (_, _, name_key) in ranked.items():
        if name_key:
            names.setdefault(name_key, set()).add(roster_key)
    unique_names = {name: next(iter(matches)) for name, matches in names.items() if len(matches) == 1}
    payload = "\n".join([
        f"policy:{AGENT_SCOPE_POLICY_VERSION}",
        *(
            f"id:{value}|status:{ranked[value][0]}|end:{ranked[value][1] or ''}"
            for value in sorted(ids)
        ),
        *(f"name:{key}={value}" for key, value in sorted(unique_names.items())),
    ])
    fingerprint = hashlib.sha256(payload.encode("utf-8")).hexdigest()
    eligibility = {
        agent_id: (values[0], values[1])
        for agent_id, values in ranked.items()
    }
    return AgentScope(frozenset(ids), unique_names, fingerprint, eligibility)


def _insert_rows(conn: DatabaseConnection, table: str, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    columns = TABLE_COLUMNS[table]
    row_placeholders = "(" + ", ".join("?" for _ in columns) + ")"
    # Bounded multi-row statements keep SQLite transaction overhead low while
    # staying below its host-parameter limit.
    batch_size = max(1, min(500, conn.max_variable_number // len(columns)))
    for offset in range(0, len(rows), batch_size):
        batch = rows[offset : offset + batch_size]
        sql = f"INSERT INTO {table} ({', '.join(columns)}) VALUES " + ", ".join(row_placeholders for _ in batch)
        params = [row.get(column) for row in batch for column in columns]
        conn.execute(sql, params)


def ingest_all(
    conn: DatabaseConnection,
    config: Config,
    families: set[str] | None = None,
    progress: ProgressCallback | None = None,
) -> IngestSummary:
    summary = IngestSummary()
    selected = set(families or ())
    if selected & AGENT_SCOPED_FAMILIES:
        selected.add("fte")
    candidates = [
        candidate for candidate in discover_sources(config)
        if not selected or candidate.family in selected
    ]
    total = len(candidates)
    queue_mapping = load_queue_mapping(config.queue_mapping)
    if progress is not None:
        progress(0, total, "Scanning source files")
    for index, candidate in enumerate(candidates, 1):
        path = candidate.path
        if progress is not None:
            progress(index - 1, total, f"Loading {candidate.family}: {path.name}")
        sha256 = file_sha256(path)
        path_text = str(path)
        try:
            scope = _load_agent_scope(conn) if candidate.family in AGENT_SCOPED_FAMILIES else None
        except Exception as exc:
            summary.failed += 1
            summary.errors.append(f"{candidate.family}: {path.name}: {exc}")
            if progress is not None:
                progress(index, total, f"Failed {candidate.family}: {path.name}")
            continue
        scope_fingerprint = scope.fingerprint if scope is not None else ""
        if candidate.family == "fte":
            # A parser upgrade must re-read an unchanged workbook so newly
            # governed worksheets do not remain invisible in an older DB.
            scope_fingerprint = FTE_PARSER_POLICY_VERSION
        if candidate.family == "schedule":
            scope_fingerprint = f"{scope_fingerprint}|{SCHEDULE_PARSER_POLICY_VERSION}"
        if candidate.family == "calls":
            # Re-read unchanged Call-by-Call files whenever the queue catalog
            # changes: mapped queue demand is part of the admitted scope.
            scope_fingerprint = (
                f"{scope_fingerprint}|{CALL_PARSER_POLICY_VERSION}|"
                f"queue-map:{queue_mapping.sha256}"
            )
        existing = conn.execute(
            """SELECT file_id, active, row_count, scoped_out_count FROM meta.source_file
               WHERE source_family=? AND source_path=? AND sha256=?
                 AND coalesce(scope_fingerprint, '')=? AND status='SUCCESS'""",
            [candidate.family, path_text, sha256, scope_fingerprint],
        ).fetchone()
        if existing:
            existing_id, active, existing_rows, existing_scoped_out = existing
            if active:
                summary.skipped += 1
                summary.scoped_out += existing_scoped_out or 0
                if progress is not None:
                    progress(index, total, f"Unchanged {candidate.family}: {path.name}")
                continue
            # The path changed A -> B -> A. Its original raw rows are still
            # immutable in the hub, so safely reactivate them without parsing
            # or duplicating the same content fingerprint.
            conn.execute("BEGIN TRANSACTION")
            try:
                conn.execute(
                    "UPDATE meta.source_file SET active=false WHERE source_family=? AND source_path=? AND active=true",
                    [candidate.family, path_text],
                )
                conn.execute(
                    "UPDATE meta.source_file SET active=true, loaded_at=? WHERE file_id=?",
                    [datetime.now(), existing_id],
                )
                conn.execute("COMMIT")
            except Exception:
                if conn.in_transaction:
                    conn.execute("ROLLBACK")
                raise
            summary.loaded += 1
            summary.rows += existing_rows or 0
            summary.scoped_out += existing_scoped_out or 0
            if progress is not None:
                progress(index, total, f"Reactivated {candidate.family}: {path.name}")
            continue
        stat = path.stat()
        file_id = hashlib.sha256(
            f"{candidate.family}|{path_text}|{sha256}|{scope_fingerprint}".encode("utf-8")
        ).hexdigest()
        discovered_at = datetime.now()
        modified_at = datetime.fromtimestamp(stat.st_mtime)
        try:
            result = None if candidate.family in {"lilo", "agent_status", "calls"} else PARSERS[candidate.family](path, file_id, scope)
            conn.execute("BEGIN TRANSACTION")
            try:
                conn.execute(
                    "UPDATE meta.source_file SET active=false WHERE source_family=? AND source_path=? AND active=true",
                    [candidate.family, path_text],
                )
                if candidate.family == "lilo":
                    date_field, fallback_date = _validate_lilo_header(path)
                    row_count, scoped_out_count, rejected_count = _insert_lilo_direct(
                        conn, path, file_id, date_field, fallback_date, scope, progress
                    )
                elif candidate.family == "agent_status":
                    row_count, scoped_out_count, rejected_count = _insert_status_direct(
                        conn, path, file_id, scope, progress
                    )
                elif candidate.family == "calls":
                    row_count, scoped_out_count, rejected_count = _insert_calls_direct(
                        conn, path, file_id, scope, queue_mapping, progress
                    )
                else:
                    row_count = result.row_count
                    rejected_count = len(result.rejected)
                    scoped_out_count = result.scoped_out
                    for table, rows in result.tables.items():
                        _insert_rows(conn, table, rows)
                source_variant = result.source_variant if result is not None else None
                load_note = (
                    "; ".join(result.rejected[:20])
                    if result is not None and result.rejected else None
                )
                conn.execute(
                    """INSERT INTO meta.source_file(
                           file_id, source_family, source_path, file_name, sha256, size_bytes,
                           modified_at, discovered_at, loaded_at, active, status, row_count,
                           rejected_count, error_message, scope_fingerprint, scoped_out_count,
                           source_variant
                       ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, true, 'SUCCESS', ?, ?, ?, ?, ?, ?)
                       ON CONFLICT(file_id) DO UPDATE SET
                           source_family=excluded.source_family, source_path=excluded.source_path,
                           file_name=excluded.file_name, sha256=excluded.sha256,
                           size_bytes=excluded.size_bytes, modified_at=excluded.modified_at,
                           discovered_at=excluded.discovered_at, loaded_at=excluded.loaded_at,
                           active=true, status='SUCCESS', row_count=excluded.row_count,
                           rejected_count=excluded.rejected_count,
                           error_message=excluded.error_message,
                           scope_fingerprint=excluded.scope_fingerprint,
                           scoped_out_count=excluded.scoped_out_count,
                           source_variant=excluded.source_variant""",
                    [file_id, candidate.family, path_text, path.name, sha256, stat.st_size, modified_at, discovered_at, datetime.now(), row_count, rejected_count, load_note, scope_fingerprint, scoped_out_count, source_variant],
                )
                conn.execute("COMMIT")
            except Exception:
                if conn.in_transaction:
                    conn.execute("ROLLBACK")
                raise
            summary.loaded += 1
            summary.rows += row_count
            summary.scoped_out += scoped_out_count
        except Exception as exc:
            conn.execute(
                """INSERT INTO meta.source_file(
                       file_id, source_family, source_path, file_name, sha256, size_bytes,
                       modified_at, discovered_at, loaded_at, active, status, row_count,
                       rejected_count, error_message, scope_fingerprint, scoped_out_count,
                       source_variant
                   ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, false, 'ERROR', 0, 0, ?, ?, 0, NULL)
                   ON CONFLICT(file_id) DO UPDATE SET
                       modified_at=excluded.modified_at, discovered_at=excluded.discovered_at,
                       loaded_at=excluded.loaded_at, active=false, status='ERROR',
                       row_count=0, rejected_count=0, error_message=excluded.error_message,
                       scope_fingerprint=excluded.scope_fingerprint, scoped_out_count=0""",
                [file_id, candidate.family, path_text, path.name, sha256, stat.st_size, modified_at, discovered_at, datetime.now(), str(exc)[:4000], scope_fingerprint],
            )
            summary.failed += 1
            summary.errors.append(f"{candidate.family}: {path.name}: {exc}")
        if progress is not None:
            progress(index, total, f"Processed {candidate.family}: {path.name}")
    return summary
