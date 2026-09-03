"""Import persistent PCS coaching decisions from a generated ACTIONS sheet."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from .config import Config
from .database import DatabaseConnection


REQUIRED_HEADERS = {
    "Coaching Key",
    "Coaching Status",
    "Coaching Date",
    "Coach",
    "Coaching Comment",
}

_STATUS_ALIASES = {
    "": "PENDING",
    "PENDING": "PENDING",
    "OPEN": "PENDING",
    "TO DO": "PENDING",
    "TODO": "PENDING",
    "COMPLETED": "COMPLETED",
    "COMPLETE": "COMPLETED",
    "DONE": "COMPLETED",
    "OK": "COMPLETED",
    "YES": "COMPLETED",
    "Y": "COMPLETED",
    "NOT REQUIRED": "NOT_REQUIRED",
    "NOT NEEDED": "NOT_REQUIRED",
    "N/A": "NOT_REQUIRED",
    "NA": "NOT_REQUIRED",
}


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    return date.fromisoformat(text) if text else None


def _status(value) -> str:
    text = " ".join(str(value or "").strip().upper().replace("_", " ").split())
    try:
        return _STATUS_ALIASES[text]
    except KeyError as exc:
        raise ValueError(
            f"Unknown coaching status {value!r}. Use Pending, Completed or Not required."
        ) from exc


def _valid_low_score_keys(
    conn: DatabaseConnection,
    config: Config,
    coaching_keys: set[str],
) -> set[str]:
    if not coaching_keys:
        return set()
    primary_score = f"question_{config.pcs.primary_score_question}_score"
    allowed = ", ".join(f"{value:g}" for value in config.pcs.allowed_scores)
    valid: set[str] = set()
    keys = sorted(coaching_keys)
    batch_size = max(1, min(900, conn.max_variable_number - 1))
    for offset in range(0, len(keys), batch_size):
        batch = keys[offset:offset + batch_size]
        placeholders = ",".join("?" for _ in batch)
        valid.update(
            str(row[0])
            for row in conn.execute(
                f"""SELECT call_key FROM core.clean_call_leg
                    WHERE call_key IN ({placeholders})
                      AND upper(coalesce(call_direction,''))='I'
                      AND {primary_score} IN ({allowed})
                      AND {primary_score} <= ?""",
                [*batch, config.pcs.negative_score_maximum],
            ).fetchall()
        )
    return valid


def import_pcs_coaching(conn: DatabaseConnection, config: Config, path: Path) -> int:
    """Import blue coaching fields without trusting workbook evidence columns."""

    path = path.resolve()
    if not path.is_file():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if "ACTIONS" not in workbook.sheetnames:
            raise ValueError("The workbook has no ACTIONS sheet")
        sheet = workbook["ACTIONS"]
        headers = [str(cell.value or "").strip() for cell in sheet[4]]
        missing = sorted(REQUIRED_HEADERS - set(headers))
        if missing:
            raise ValueError(f"ACTIONS is missing coaching columns: {', '.join(missing)}")
        index = {header: headers.index(header) for header in headers if header}
        decisions = []
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=5, values_only=True), start=5
        ):
            coaching_key = str(values[index["Coaching Key"]] or "").strip()
            if not coaching_key or coaching_key == "No rows for this period.":
                continue
            status = _status(values[index["Coaching Status"]])
            coaching_date = _as_date(values[index["Coaching Date"]])
            coach = str(values[index["Coach"]] or "").strip() or None
            comment = str(values[index["Coaching Comment"]] or "").strip() or None
            # Do not fill the database with untouched default Pending rows.
            if status == "PENDING" and coaching_date is None and coach is None and comment is None:
                continue
            decisions.append((
                row_number, coaching_key, status, coaching_date, coach, comment,
            ))

        valid_keys = _valid_low_score_keys(
            conn, config, {decision[1] for decision in decisions},
        )
        for row_number, coaching_key, *_ in decisions:
            if coaching_key not in valid_keys:
                raise ValueError(
                    f"ACTIONS row {row_number} has an unknown or no-longer-eligible Coaching Key"
                )

        imported = 0
        conn.execute("SAVEPOINT import_pcs_coaching")
        try:
            for _, coaching_key, status, coaching_date, coach, comment in decisions:
                conn.execute(
                    """
                    INSERT INTO core.pcs_coaching_action VALUES (?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(coaching_key) DO UPDATE SET
                        coaching_status=excluded.coaching_status,
                        coaching_date=excluded.coaching_date,
                        coach=excluded.coach,
                        coaching_comment=excluded.coaching_comment,
                        updated_at=excluded.updated_at,
                        imported_from=excluded.imported_from
                    """,
                    [
                        coaching_key, status, coaching_date, coach, comment,
                        datetime.now(), path.name,
                    ],
                )
                imported += 1
            conn.execute("RELEASE SAVEPOINT import_pcs_coaching")
            return imported
        except Exception:
            conn.execute("ROLLBACK TO SAVEPOINT import_pcs_coaching")
            conn.execute("RELEASE SAVEPOINT import_pcs_coaching")
            raise
    finally:
        workbook.close()
