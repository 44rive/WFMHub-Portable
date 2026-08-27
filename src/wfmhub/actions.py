"""Import persistent human decisions from an exported GAPS worksheet."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from .database import DatabaseConnection


REQUIRED_HEADERS = {
    "Correction ID",
    "Confirmed Activity",
    "Validation Status",
    "Owner",
    "Comment",
    "Injected Date",
}


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    return date.fromisoformat(text) if text else None


def import_actions(conn: DatabaseConnection, path: Path) -> int:
    path = path.resolve()
    if not path.is_file():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if "GAPS" not in workbook.sheetnames:
            raise ValueError("The workbook has no GAPS sheet")
        sheet = workbook["GAPS"]
        headers = [str(cell.value or "").strip() for cell in sheet[4]]
        missing = sorted(REQUIRED_HEADERS - set(headers))
        if missing:
            raise ValueError(f"GAPS is missing decision columns: {', '.join(missing)}")
        index = {header: headers.index(header) for header in headers if header}
        imported = 0
        conn.execute("SAVEPOINT import_actions")
        try:
            for values in sheet.iter_rows(min_row=5, values_only=True):
                correction_id = str(values[index["Correction ID"]] or "").strip()
                if not correction_id:
                    continue
                fields = {
                    "confirmed_activity": values[index["Confirmed Activity"]],
                    "validation_status": values[index["Validation Status"]],
                    "owner": values[index["Owner"]],
                    "comment": values[index["Comment"]],
                    "injected_date": _as_date(values[index["Injected Date"]]),
                }
                # An untouched exported row says Open. It is still safe to import,
                # but blank rows do not create noise in the persistent action table.
                if not any(value not in (None, "", "Open") for value in fields.values()):
                    continue
                conn.execute(
                    """
                    INSERT INTO core.correction_action VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(correction_id) DO UPDATE SET
                        confirmed_activity=excluded.confirmed_activity,
                        validation_status=excluded.validation_status,
                        owner=excluded.owner,
                        comment=excluded.comment,
                        injected_date=excluded.injected_date,
                        updated_at=excluded.updated_at,
                        imported_from=excluded.imported_from
                    """,
                    [correction_id, fields["confirmed_activity"], fields["validation_status"] or "Open", fields["owner"], fields["comment"], fields["injected_date"], datetime.now(), path.name],
                )
                conn.execute(
                    """
                    UPDATE mart.correction_candidate SET
                        confirmed_activity=?, validation_status=?, owner=?, comment=?, injected_date=?
                    WHERE correction_id=?
                    """,
                    [fields["confirmed_activity"], fields["validation_status"] or "Open", fields["owner"], fields["comment"], fields["injected_date"], correction_id],
                )
                imported += 1
            conn.execute("RELEASE SAVEPOINT import_actions")
            return imported
        except Exception:
            conn.execute("ROLLBACK TO SAVEPOINT import_actions")
            conn.execute("RELEASE SAVEPOINT import_actions")
            raise
    finally:
        workbook.close()
