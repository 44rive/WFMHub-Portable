"""Focused WFM/Operations report products using one workbook design contract."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import load_workbook

from .config import Config
from .database import DatabaseConnection
from .metrics import MetricCatalog, evaluate_metric, load_metric_catalog
from .report_packs import publish_report, report_current_path
from .reports import COLORS, _query
from .rules import load_rulebook
from .service_profiles import ServiceProfile, load_service_profiles
from .template_reports import DecisionWorkbook, KpiCard, ModelTable


@dataclass(frozen=True)
class NamedPeriod:
    label: str
    start: date
    end: date


def _month_start(value: date) -> date:
    return value.replace(day=1)


def _previous_month(value: date) -> tuple[date, date]:
    end = value.replace(day=1) - timedelta(days=1)
    return end.replace(day=1), end


def _comparison_periods(start: date, end: date) -> list[NamedPeriod]:
    previous_start, previous_end = _previous_month(end)
    prior_mtd_end = min(previous_end, previous_start + timedelta(days=end.day - 1))
    duration = (end - start).days + 1
    prior_equal_end = start - timedelta(days=1)
    prior_equal_start = prior_equal_end - timedelta(days=duration - 1)
    periods = [
        NamedPeriod("Latest day", end, end),
        NamedPeriod("Selected period", start, end),
        NamedPeriod("Current MTD", _month_start(end), end),
        NamedPeriod("Prior-month same days", previous_start, prior_mtd_end),
        NamedPeriod("Previous full month", previous_start, previous_end),
    ]
    if start != _month_start(end):
        periods.insert(2, NamedPeriod("Previous equal period", prior_equal_start, prior_equal_end))
    return periods


def _ratio(numerator: float | int | None, denominator: float | int | None) -> float | None:
    return float(numerator) / float(denominator) if denominator else None


def _delta(value: float | None, reference: float | None, percent: bool = False) -> str:
    if value is None or reference is None:
        return "Comparison not available"
    difference = value - reference
    return f"{difference:+.1%} vs prior" if percent else f"{difference:+.2f} vs prior"


def _output_path(config: Config, key: str, start: date, end: date, generated: datetime, output: Path | None) -> Path:
    del start, end, generated
    return (output or report_current_path(config, key)).resolve()


def _source_state(conn: DatabaseConnection, families: Sequence[str], through: date, final: bool = False) -> tuple[str, str]:
    if not families:
        return ("FINAL" if final else "LIVE"), "All required datasets are available"
    placeholders = ",".join("?" for _ in families)
    rows = conn.execute(
        f"""SELECT source_family, newest_business_date, status
            FROM mart.source_health WHERE source_family IN ({placeholders})""",
        list(families),
    ).fetchall()
    found = {str(row[0]).lower(): row for row in rows}
    problems = []
    for family in families:
        if family.casefold() in {"start_end", "activities"}:
            variant = "START_END" if family.casefold() == "start_end" else "ACTIVITIES"
            variant_row = conn.execute(
                """SELECT max(r.schedule_date), count(*)
                   FROM raw.schedule_shift r
                   JOIN meta.source_file f ON f.file_id=r.source_file_id
                   WHERE f.active=true AND f.status='SUCCESS' AND f.source_variant=?""",
                [variant],
            ).fetchone()
            variant_date, variant_count = variant_row
            if family.casefold() == "start_end" and not variant_count:
                variant = "ACTIVITIES"
                variant_row = conn.execute(
                    """SELECT max(r.schedule_date), count(*)
                       FROM raw.schedule_shift r
                       JOIN meta.source_file f ON f.file_id=r.source_file_id
                       WHERE f.active=true AND f.status='SUCCESS'
                         AND f.source_variant='ACTIVITIES' AND r.parse_ok=true
                         AND r.scheduled_start IS NOT NULL
                         AND r.scheduled_end IS NOT NULL"""
                ).fetchone()
                variant_date, variant_count = variant_row
            if not variant_count:
                problems.append(f"{family}: no successful {variant} extract")
            elif variant_date is None or str(variant_date)[:10] < through.isoformat():
                problems.append(f"{family}: latest {variant_date or 'unknown'}")
            continue
        row = found.get(family.lower())
        if row is None:
            problems.append(f"{family}: no health record")
        elif str(row[2]).upper() != "SUCCESS":
            problems.append(f"{family}: {row[2]}")
        # FTE is a point-in-time scope roster rather than a dated fact source.
        # Its freshness is represented by load status and source hash, so a
        # NULL business date is expected and must not make every report red.
        elif family.casefold() != "fte" and (row[1] is None or str(row[1])[:10] < through.isoformat()):
            problems.append(f"{family}: latest {row[1] or 'unknown'}")
    if problems:
        return "INCOMPLETE", "; ".join(problems)
    if final:
        return "FINAL", f"Required sources loaded through {through}"
    if through >= date.today():
        return "PROVISIONAL", "Current-day values can still change before shifts and queues close"
    return "LIVE", f"Required sources loaded through {through}"


def _audit_rows(
    conn: DatabaseConnection,
    config: Config,
    report_key: str,
    start: date,
    end: date,
    extra: Iterable[Sequence[Any]] = (),
) -> list[Sequence[Any]]:
    latest = conn.execute(
        "SELECT run_id, finished_at, details FROM meta.refresh_run WHERE status='SUCCESS' ORDER BY finished_at DESC LIMIT 1"
    ).fetchone()
    rows: list[Sequence[Any]] = [
        ("Report", report_key, "WFM report product"),
        ("Selected period", f"{start} to {end}", "Dates included"),
        ("Last refreshed", datetime.now(), "Local work-machine time"),
        ("Refresh run", latest[0] if latest else None, latest[2] if latest else "No successful refresh metadata"),
        ("Prepared by", "Anass ASSRI", "WFM"),
    ]
    rows.extend(extra)
    return rows


def _atomic_book(
    config: Config,
    key: str,
    title: str,
    start: date,
    end: date,
    output: Path | None,
) -> tuple[DecisionWorkbook, Path, Path]:
    generated = datetime.now()
    target = _output_path(config, key, start, end, generated, output)
    target.parent.mkdir(parents=True, exist_ok=True)
    partial = target.with_name(f"{target.stem}.partial{target.suffix}")
    return DecisionWorkbook(partial, config, key, title, start, end, generated), partial, target


def _finish(book: DecisionWorkbook, partial: Path, target: Path) -> Path:
    try:
        book.close()
        publish_report(book.config, book.report_key, partial, target, book.generated)
    except Exception:
        partial.unlink(missing_ok=True)
        raise
    return target


def _pcs_aggregate(
    conn: DatabaseConnection,
    config: Config,
    period: NamedPeriod,
) -> tuple[Any, ...]:
    row = conn.execute(
        """SELECT coalesce(sum(pcs_score_sum),0), coalesce(sum(survey_responses),0),
                  coalesce(sum(pcs_participation_responses),0), coalesce(sum(pcs_status_calls),0),
                  coalesce(sum(low_score_responses),0), coalesce(sum(top_box_responses),0),
                  coalesce(sum(inbound_calls),0)
           FROM mart.agent_pcs_day WHERE business_date BETWEEN ? AND ?""",
        [period.start, period.end],
    ).fetchone()
    score_sum, valid, participants, eligible, low, high, inbound = row
    return (
        period.label, period.start, period.end, _ratio(score_sum, valid),
        _ratio(participants, eligible), valid, eligible, low, high, inbound,
    )


def _pcs_coaching_rows(
    conn: DatabaseConnection,
    config: Config,
    start: date,
    end: date,
) -> tuple[list[str], list[tuple[Any, ...]]]:
    """Return a self-contained coaching queue; manual fields stay in Excel."""

    primary = config.pcs.primary_score_question
    primary_score = f"question_{primary}_score"
    allowed_scores = ", ".join(f"{value:g}" for value in config.pcs.allowed_scores)
    return _query(
        conn,
        f"""SELECT coalesce(d.lob,c.lob) AS lob,
                   d.team_leader,
                   coalesce(d.canonical_name,c.agent_name) || ' [' || c.agent_id || ']' AS agent_selector,
                   coalesce(d.canonical_name,c.agent_name) AS agent_name,
                   c.agent_id,
                   CASE WHEN c.{primary_score}<=2 THEN 'HIGH' ELSE 'NORMAL' END AS priority,
                   c.business_date, c.call_start,
                   c.{primary_score} AS q1_score,
                   c.question_3 AS customer_comment,
                   c.call_reference_number,
                   coalesce(d.language,c.language) AS language,
                   c.call_key AS coaching_key,
                   'Pending' AS coaching_status,
                   NULL AS coach, NULL AS coaching_date, NULL AS due_date,
                   NULL AS coaching_comment,
                   d.ops_manager
            FROM core.clean_call_leg c
            LEFT JOIN core.dim_agent d ON d.agent_id=c.agent_id
            WHERE c.business_date BETWEEN ? AND ?
              AND upper(coalesce(c.call_direction,''))='I'
              AND c.{primary_score} IN ({allowed_scores})
              AND c.{primary_score} <= ?
            ORDER BY priority, c.business_date DESC, d.team_leader,
                     coalesce(d.canonical_name,c.agent_name), c.call_start""",
        [start, end, config.pcs.negative_score_maximum],
    )


def _pcs_sum_formula(column: str, from_name: str = "PCS_From", to_name: str = "PCS_To") -> str:
    scope = (
        f'(tblPcsData[Date]>={from_name})*(tblPcsData[Date]<={to_name})*'
        'IF(DASHBOARD!$K$6="All",1,--(tblPcsData[LOB]=DASHBOARD!$K$6))*'
        'IF(DASHBOARD!$N$6="All",1,--(tblPcsData[Team Leader]=DASHBOARD!$N$6))*'
        'IF(DASHBOARD!$Q$6="All",1,--(tblPcsData[Agent Selector]=DASHBOARD!$Q$6))'
    )
    return f"SUMPRODUCT({scope}*N(tblPcsData[{column}]))"


def _pcs_completed_formula() -> str:
    return (
        'SUMPRODUCT((tblCoaching[Date]>=PCS_From)*(tblCoaching[Date]<=PCS_To)*'
        'IF(DASHBOARD!$K$6="All",1,--(tblCoaching[LOB]=DASHBOARD!$K$6))*'
        'IF(DASHBOARD!$N$6="All",1,--(tblCoaching[Team Leader]=DASHBOARD!$N$6))*'
        'IF(DASHBOARD!$Q$6="All",1,--(tblCoaching[Agent Selector]=DASHBOARD!$Q$6))*'
        '--(tblCoaching[Coaching Status]="Completed"))'
    )


def _pcs_agent_sum_formula(
    excel_row: int,
    column: str,
    from_name: str | None = "PCS_From",
    to_name: str | None = "PCS_To",
) -> str:
    """Exact Agent Selector total used by the refreshable agent results."""

    date_scope = (
        f"(tblPcsData[Date]>={from_name})*(tblPcsData[Date]<={to_name})"
        if from_name and to_name
        else "(tblPcsData[Date]=PCS_Latest)"
    )
    return (
        f"SUMPRODUCT((tblPcsData[Agent Selector]=$C{excel_row})*"
        f"{date_scope}*N(tblPcsData[{column}]))"
    )


def _add_pcs_dashboard(
    book: DecisionWorkbook,
    status: str,
    status_text: str,
    start: date,
    end: date,
    lobs: Sequence[str],
    team_leaders: Sequence[str],
    agents: Sequence[str],
    data_start: date,
    latest: date,
    trend_count: int,
    minimum_sample: int,
    default_values: dict[str, float | int | None],
) -> None:
    """Create an Excel-native selector cockpit without Power Query or macros."""

    wb = book.report.workbook
    wb.set_calc_mode("auto")
    ws = wb.add_worksheet("DASHBOARD")
    ws.hide_gridlines(2)
    ws.set_tab_color(COLORS["gold"])
    ws.set_zoom(85)
    ws.set_landscape()
    ws.fit_to_pages(1, 1)
    ws.freeze_panes(4, 0)
    ws.merge_range("A1:R1", "PCS  /  PERFORMANCE & COACHING CONTROL", book.report.title)
    ws.merge_range(
        "A2:R2",
        f"Data updated {book.generated:%Y-%m-%d %H:%M}  |  latest PCS data {latest:%Y-%m-%d}  |  prepared by Anass ASSRI",
        book.report.subtitle,
    )
    badge = status if status in book.badge_formats else "INCOMPLETE"
    ws.merge_range("A4:R4", f"{badge}  /  {status_text}", book.badge_formats[badge])

    selector_label = wb.add_format({
        "font_name": "Aptos", "font_size": 8, "bold": True,
        "font_color": COLORS["muted"], "bg_color": COLORS["canvas"],
        "align": "left", "valign": "vcenter", "indent": 1,
    })
    selector = wb.add_format({
        "font_name": "Aptos Display", "font_size": 11, "bold": True,
        "font_color": COLORS["dark"], "bg_color": COLORS["white"],
        "border": 1, "border_color": COLORS["teal"], "align": "left",
        "valign": "vcenter", "indent": 1, "num_format": "yyyy-mm-dd",
    })
    for label, label_range, value_range, value in (
        ("PERIOD VIEW", "A5:C5", "A6:C7", "Current MTD"),
        ("CUSTOM FROM", "E5:F5", "E6:F7", start),
        ("CUSTOM TO", "H5:I5", "H6:I7", end),
        ("LOB", "K5:L5", "K6:L7", "All"),
        ("TEAM LEADER", "N5:O5", "N6:O7", "All"),
        ("AGENT", "Q5:R5", "Q6:R7", "All"),
    ):
        ws.merge_range(label_range, label, selector_label)
        if isinstance(value, date):
            ws.merge_range(value_range, value, selector)
        else:
            ws.merge_range(value_range, value, selector)
    ws.set_row(5, 22)
    ws.set_row(6, 22)
    period_choices = [
        "Latest day", "Current week", "Previous week", "Current MTD",
        "Previous-month same days", "Previous full month", "Custom period",
    ]
    ws.data_validation("A6", {"validate": "list", "source": period_choices})
    ws.data_validation("E6", {"validate": "date", "criteria": "between", "minimum": data_start, "maximum": latest})
    ws.data_validation("H6", {"validate": "date", "criteria": "between", "minimum": data_start, "maximum": latest})

    del lobs, team_leaders, agents
    ws.data_validation("K6", {"validate": "list", "source": "=PCS_LOB_LIST"})
    ws.data_validation("N6", {"validate": "list", "source": "=PCS_TL_LIST"})
    ws.data_validation("Q6", {"validate": "list", "source": "=PCS_AGENT_LIST"})

    # Period selectors follow the latest date inside the refreshable PCS table.
    # Replacing PCS_DATA therefore advances Current week/MTD without rebuilding.
    wb.define_name("PCS_Latest", "=MAX(tblPcsData[Date])")
    wb.define_name(
        "PCS_From",
        '=IF(DASHBOARD!$A$6="Latest day",PCS_Latest,'
        'IF(DASHBOARD!$A$6="Current week",PCS_Latest-WEEKDAY(PCS_Latest,2)+1,'
        'IF(DASHBOARD!$A$6="Previous week",PCS_Latest-WEEKDAY(PCS_Latest,2)-6,'
        'IF(DASHBOARD!$A$6="Current MTD",EOMONTH(PCS_Latest,-1)+1,'
        'IF(DASHBOARD!$A$6="Previous-month same days",EOMONTH(PCS_Latest,-2)+1,'
        'IF(DASHBOARD!$A$6="Previous full month",EOMONTH(PCS_Latest,-2)+1,DASHBOARD!$E$6))))))',
    )
    wb.define_name(
        "PCS_To",
        '=IF(DASHBOARD!$A$6="Latest day",PCS_Latest,'
        'IF(DASHBOARD!$A$6="Current week",PCS_Latest,'
        'IF(DASHBOARD!$A$6="Previous week",PCS_Latest-WEEKDAY(PCS_Latest,2),'
        'IF(DASHBOARD!$A$6="Current MTD",PCS_Latest,'
        'IF(DASHBOARD!$A$6="Previous-month same days",EDATE(PCS_Latest,-1),'
        'IF(DASHBOARD!$A$6="Previous full month",EOMONTH(PCS_Latest,-1),DASHBOARD!$H$6))))))',
    )
    wb.define_name("PCS_Prior_From", "=EOMONTH(PCS_Latest,-2)+1")
    wb.define_name("PCS_Prior_To", "=EDATE(PCS_Latest,-1)")

    scope_count = _pcs_sum_formula("Valid Q1")
    ws.merge_range("A8:R8", "", book.report.note)
    ws.write_formula(
        "A8", f'=IF({scope_count}=0,"NO MATCHING PCS DATA - CHECK THE SELECTORS",'
        f'"Showing "&TEXT(PCS_From,"yyyy-mm-dd")&" to "&TEXT(PCS_To,"yyyy-mm-dd"))',
        book.report.note, "Showing current MTD",
    )

    score_sum = _pcs_sum_formula("Q1 Score Sum")
    valid = _pcs_sum_formula("Valid Q1")
    participating = _pcs_sum_formula("Q1 Nonblank")
    eligible = _pcs_sum_formula("PCS Status 1")
    low = _pcs_sum_formula("Score <= 3")
    positive = _pcs_sum_formula("Score > 3")
    inbound = _pcs_sum_formula("Inbound Call Legs")
    completed = _pcs_completed_formula()
    cards = [
        ("PCS AVERAGE", f'=IFERROR({score_sum}/{valid},"")', book.card_decimal, "Weighted score / valid responses", default_values.get("pcs_average")),
        ("PARTICIPATION", f'=IFERROR({participating}/{eligible},"")', book.card_percent, "Q1 nonblank / PCS Status 1", default_values.get("participation")),
        ("VALID RESPONSES", f"={valid}", book.card_integer, f"Low sample below {minimum_sample}", default_values.get("valid")),
        ("INBOUND CALL LEGS", f"={inbound}", book.card_integer, "Inbound legs in selected scope", default_values.get("inbound")),
        ("SCORE <= 3", f"={low}", book.card_integer, "Coaching opportunities", default_values.get("low")),
        ("POSITIVE > 3", f"={positive}", book.card_integer, "Positive valid responses", default_values.get("positive")),
        ("COACHING COMPLETED", f"={completed}", book.card_integer, "Updates as the team fills COACHING", 0),
        ("ACTIONS RATE", f'=IFERROR({completed}/{low},"")', book.card_percent, "Completed / score <= 3", 0),
    ]
    for index, (label, formula, fmt, note, cached) in enumerate(cards):
        row = 9 if index < 4 else 14
        column = (index % 4) * 4
        ws.merge_range(row, column, row, column + 2, label, book.report.kpi_label)
        ws.merge_range(row + 1, column, row + 2, column + 2, "", fmt)
        ws.write_formula(row + 1, column, formula, fmt, cached if cached is not None else "")
        ws.merge_range(row + 3, column, row + 3, column + 2, note, book.card_compare)
        ws.set_row(row + 1, 26)
        ws.set_row(row + 2, 26)

    table_row = 20
    ws.merge_range(table_row, 0, table_row, 9, "PERIOD BENCHMARK", book.report.section)
    compare_headers = ["Period", "Start", "End", "PCS Average", "Participation %", "Valid Responses", "PCS Status 1", "Score <= 3", "Score > 3", "Inbound Legs"]
    for column, header in enumerate(compare_headers):
        ws.write(table_row + 2, column, header, book.report.header)
    current_values = [
        "Selected scope", "=PCS_From", "=PCS_To",
        f'=IFERROR({_pcs_sum_formula("Q1 Score Sum")}/{_pcs_sum_formula("Valid Q1")},"")',
        f'=IFERROR({_pcs_sum_formula("Q1 Nonblank")}/{_pcs_sum_formula("PCS Status 1")},"")',
        f'={_pcs_sum_formula("Valid Q1")}', f'={_pcs_sum_formula("PCS Status 1")}',
        f'={_pcs_sum_formula("Score <= 3")}', f'={_pcs_sum_formula("Score > 3")}',
        f'={_pcs_sum_formula("Inbound Call Legs")}',
    ]
    prior_values = [
        "Previous-month same days", "=PCS_Prior_From", "=PCS_Prior_To",
        f'=IFERROR({_pcs_sum_formula("Q1 Score Sum", "PCS_Prior_From", "PCS_Prior_To")}/{_pcs_sum_formula("Valid Q1", "PCS_Prior_From", "PCS_Prior_To")},"")',
        f'=IFERROR({_pcs_sum_formula("Q1 Nonblank", "PCS_Prior_From", "PCS_Prior_To")}/{_pcs_sum_formula("PCS Status 1", "PCS_Prior_From", "PCS_Prior_To")},"")',
        f'={_pcs_sum_formula("Valid Q1", "PCS_Prior_From", "PCS_Prior_To")}',
        f'={_pcs_sum_formula("PCS Status 1", "PCS_Prior_From", "PCS_Prior_To")}',
        f'={_pcs_sum_formula("Score <= 3", "PCS_Prior_From", "PCS_Prior_To")}',
        f'={_pcs_sum_formula("Score > 3", "PCS_Prior_From", "PCS_Prior_To")}',
        f'={_pcs_sum_formula("Inbound Call Legs", "PCS_Prior_From", "PCS_Prior_To")}',
    ]
    for offset, values in enumerate((current_values, prior_values)):
        for column, value in enumerate(values):
            fmt = book.report.date if column in {1, 2} else book.report.percent if column == 4 else book.report.decimal if column == 3 else book.report.integer if column >= 5 else book.report.body
            if isinstance(value, str) and value.startswith("="):
                ws.write_formula(table_row + 3 + offset, column, value, fmt)
            else:
                ws.write(table_row + 3 + offset, column, value, fmt)

    chart = wb.add_chart({"type": "line"})
    for name, column, color, secondary in (
        ("PCS Average", 5, COLORS["teal"], False),
        ("Participation", 6, COLORS["gold"], True),
    ):
        chart.add_series({
            "name": name,
            "categories": ["_LOOKUPS", 1, 0, trend_count, 0],
            "values": ["_LOOKUPS", 1, column, trend_count, column],
            "line": {"color": color, "width": 2.25},
            "marker": {"type": "circle", "size": 4, "border": {"color": color}, "fill": {"color": COLORS["white"]}},
            "y2_axis": secondary,
        })
    chart.set_title({"name": "Daily PCS & participation"})
    chart.set_legend({"position": "bottom"})
    chart.set_chartarea({"border": {"none": True}, "fill": {"color": COLORS["white"]}})
    chart.set_plotarea({"border": {"none": True}, "fill": {"color": COLORS["white"]}})
    chart.set_y_axis({"min": 1, "max": 5, "major_unit": 1, "major_gridlines": {"visible": False}, "name": "PCS score"})
    chart.set_y2_axis({"min": 0, "max": 1, "major_unit": 0.2, "num_format": "0%", "name": "Participation"})
    ws.insert_chart("L21", chart, {"x_scale": 1.15, "y_scale": 1.05})

    note_row = table_row + 11
    ws.merge_range(note_row, 0, note_row, 17, "HOW TO USE THIS PAGE", book.report.section)
    ws.merge_range(note_row + 1, 0, note_row + 1, 17, "1. Choose a period and your Team Leader name. Every KPI, benchmark and chart on this page follows those selectors.", book.report.note)
    ws.merge_range(note_row + 2, 0, note_row + 2, 17, "2. Open AGENT_RESULTS for the realization list. Quality records completed follow-up in the blue COACHING columns.", book.report.note)
    ws.set_column("A:R", 11)
    ws.set_column("A:A", 13)
    ws.set_column("K:R", 13)


def _add_pcs_team_view(book: DecisionWorkbook, minimum_sample: int) -> None:
    """Create the novice-facing PCS view driven by the Dashboard selectors."""

    wb = book.report.workbook
    ws = wb.add_worksheet("TEAM_VIEW")
    ws.hide_gridlines(2)
    ws.set_tab_color(COLORS["gold"])
    ws.set_zoom(85)
    ws.freeze_panes(10, 0)
    ws.merge_range("A1:AA1", "PCS  /  TEAM REALISATIONS & COACHING", book.report.title)
    ws.merge_range(
        "A2:AA2",
        "Select period, LOB, Team Leader and Agent on DASHBOARD. This page follows the same selection automatically.",
        book.report.subtitle,
    )
    selector_label = wb.add_format({
        "font_name": "Aptos", "font_size": 8, "bold": True,
        "font_color": COLORS["muted"], "bg_color": COLORS["canvas"],
        "align": "left", "valign": "vcenter", "indent": 1,
    })
    selector_value = wb.add_format({
        "font_name": "Aptos Display", "font_size": 11, "bold": True,
        "font_color": COLORS["dark"], "bg_color": COLORS["white"],
        "border": 1, "border_color": COLORS["teal"], "align": "left",
        "valign": "vcenter", "indent": 1,
    })
    selectors = (
        ("PERIOD", "A4:C4", "A5:C6", "=DASHBOARD!$A$6"),
        ("LOB", "E4:G4", "E5:G6", "=DASHBOARD!$K$6"),
        ("TEAM LEADER", "I4:L4", "I5:L6", "=DASHBOARD!$N$6"),
        ("AGENT", "N4:R4", "N5:R6", "=DASHBOARD!$Q$6"),
    )
    for label, label_range, value_range, formula in selectors:
        ws.merge_range(label_range, label, selector_label)
        ws.merge_range(value_range, "", selector_value)
        first = value_range.split(":", 1)[0]
        ws.write_formula(first, formula, selector_value, "All")
    ws.write_url(
        "T5", "internal:'DASHBOARD'!A1", book.report.editable,
        string="CHANGE FILTERS ON DASHBOARD",
    )
    ws.merge_range(
        "A8:L8", "AGENT REALISATIONS", book.report.section,
    )
    agent_headers = [
        "LOB", "Team Leader", "Agent Selector", "Agent ID", "Language",
        "PCS Average", "Participation %", "Valid Q1", "PCS Status 1",
        "Score <= 3", "Prior MTD PCS", "Priority",
    ]
    for column, header in enumerate(agent_headers):
        ws.write(9, column, header, book.report.header)
    agent_formula = (
        '=LET(d,tblPcsData,'
        'm,(d[Date]>=PCS_From)*(d[Date]<=PCS_To)*'
        'IF(DASHBOARD!$K$6="All",1,--(d[LOB]=DASHBOARD!$K$6))*'
        'IF(DASHBOARD!$N$6="All",1,--(d[Team Leader]=DASHBOARD!$N$6))*'
        'IF(DASHBOARD!$Q$6="All",1,--(d[Agent Selector]=DASHBOARD!$Q$6)),'
        'pm,(d[Date]>=PCS_Prior_From)*(d[Date]<=PCS_Prior_To)*'
        'IF(DASHBOARD!$K$6="All",1,--(d[LOB]=DASHBOARD!$K$6))*'
        'IF(DASHBOARD!$N$6="All",1,--(d[Team Leader]=DASHBOARD!$N$6))*'
        'IF(DASHBOARD!$Q$6="All",1,--(d[Agent Selector]=DASHBOARD!$Q$6)),'
        'a,SORT(UNIQUE(FILTER(d[Agent Selector],m,""))),'
        'v,MAP(a,LAMBDA(x,SUMPRODUCT(m*(d[Agent Selector]=x)*N(d[Valid Q1])))),'
        's,MAP(a,LAMBDA(x,SUMPRODUCT(m*(d[Agent Selector]=x)*N(d[Q1 Score Sum])))),'
        'e,MAP(a,LAMBDA(x,SUMPRODUCT(m*(d[Agent Selector]=x)*N(d[PCS Status 1])))),'
        'p,MAP(a,LAMBDA(x,SUMPRODUCT(m*(d[Agent Selector]=x)*N(d[Q1 Nonblank])))),'
        'lo,MAP(a,LAMBDA(x,SUMPRODUCT(m*(d[Agent Selector]=x)*N(d[Score <= 3])))),'
        'pv,MAP(a,LAMBDA(x,SUMPRODUCT(pm*(d[Agent Selector]=x)*N(d[Valid Q1])))),'
        'ps,MAP(a,LAMBDA(x,SUMPRODUCT(pm*(d[Agent Selector]=x)*N(d[Q1 Score Sum])))),'
        'IFERROR(HSTACK('
        'XLOOKUP(a,d[Agent Selector],d[LOB],""),'
        'XLOOKUP(a,d[Agent Selector],d[Team Leader],""),a,'
        'XLOOKUP(a,d[Agent Selector],d[Agent ID],""),'
        'XLOOKUP(a,d[Agent Selector],d[Language],""),'
        'IFERROR(s/v,""),IFERROR(p/e,""),v,e,lo,IFERROR(ps/pv,""),'
        f'IF(v=0,"NO RESPONSE",IF(v<{minimum_sample},"LOW SAMPLE",IF(lo>0,"COACH","ON TRACK")))),'
        '"No matching agent data"))'
    )
    ws.write_dynamic_array_formula("A11", agent_formula, book.report.body, "Open in desktop Excel")
    ws.conditional_format(
        "L11:L1048576",
        {"type": "text", "criteria": "containing", "value": "COACH", "format": book.report.error},
    )
    ws.merge_range("N8:AA8", "COACHING OPPORTUNITIES", book.report.section)
    coaching_headers = [
        "LOB", "Team Leader", "Agent Selector", "Agent ID", "Priority", "Date",
        "Call Start", "Q1 Score", "Customer Comment", "Call Reference Number",
        "Coaching Key", "Action Status",
    ]
    for column, header in enumerate(coaching_headers, 13):
        ws.write(9, column, header, book.report.header)
    coaching_formula = (
        '=LET(q,tblCoachingQueue,'
        'm,(q[Date]>=PCS_From)*(q[Date]<=PCS_To)*'
        'IF(DASHBOARD!$K$6="All",1,--(q[LOB]=DASHBOARD!$K$6))*'
        'IF(DASHBOARD!$N$6="All",1,--(q[Team Leader]=DASHBOARD!$N$6))*'
        'IF(DASHBOARD!$Q$6="All",1,--(q[Agent Selector]=DASHBOARD!$Q$6)),'
        'IFERROR(FILTER(CHOOSECOLS(q,1,2,3,5,6,7,8,9,10,11,13,14),m),'
        '"No coaching opportunities in this selection"))'
    )
    ws.write_dynamic_array_formula("N11", coaching_formula, book.report.body, "Open in desktop Excel")
    ws.write_url(
        "T7", "internal:'COACHING'!A1", book.report.editable,
        string="OPEN PERMANENT COACHING LOG",
    )
    ws.set_column("A:A", 18)
    ws.set_column("B:B", 22)
    ws.set_column("C:C", 30)
    ws.set_column("D:E", 15)
    ws.set_column("F:K", 15)
    ws.set_column("L:L", 16)
    ws.set_column("M:M", 3)
    ws.set_column("N:AA", 18)
    ws.set_column("V:V", 34)
    ws.set_landscape()
    ws.fit_to_pages(1, 0)


def _previous_coaching_values(path: Path) -> dict[str, dict[str, Any]]:
    """Carry the team's editable cells forward without importing them to SQLite."""
    if not path.exists():
        return {}
    try:
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    except Exception:
        return {}
    try:
        if "COACHING" not in workbook.sheetnames:
            return {}
        sheet = workbook["COACHING"]
        headers = {
            str(cell.value).strip(): index
            for index, cell in enumerate(next(sheet.iter_rows(min_row=4, max_row=4)), 1)
            if cell.value is not None
        }
        key_column = headers.get("Coaching Key")
        editable = ("Coaching Status", "Coach", "Coaching Date", "Due Date", "Coaching Comment")
        if key_column is None:
            return {}
        output: dict[str, dict[str, Any]] = {}
        for values in sheet.iter_rows(min_row=5, values_only=True):
            key = values[key_column - 1] if key_column <= len(values) else None
            if not key:
                continue
            output[str(key)] = {
                field: values[column - 1] if column <= len(values) else None
                for field in editable
                if (column := headers.get(field)) is not None
            }
        return output
    finally:
        workbook.close()


def _previous_table_values(
    path: Path,
    sheet_name: str,
    key_header: str,
    editable_headers: Sequence[str],
) -> dict[str, dict[str, Any]]:
    """Read manual values from a prior report without importing them to the Hub."""

    if not path.exists():
        return {}
    try:
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    except Exception:
        return {}
    try:
        if sheet_name not in workbook.sheetnames:
            return {}
        sheet = workbook[sheet_name]
        headers = {
            str(cell.value).strip(): index
            for index, cell in enumerate(next(sheet.iter_rows(min_row=4, max_row=4)), 1)
            if cell.value is not None
        }
        key_column = headers.get(key_header)
        if key_column is None:
            return {}
        output: dict[str, dict[str, Any]] = {}
        for values in sheet.iter_rows(min_row=5, values_only=True):
            key = values[key_column - 1] if key_column <= len(values) else None
            if not key:
                continue
            output[str(key)] = {
                field: values[column - 1] if column <= len(values) else None
                for field in editable_headers
                if (column := headers.get(field)) is not None
            }
        return output
    finally:
        workbook.close()


def _carry_table_values(
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    key_header: str,
    editable_headers: Sequence[str],
    previous: dict[str, dict[str, Any]],
) -> list[tuple[Any, ...]]:
    display = [header.replace("_", " ").title().replace("Id", "ID") for header in headers]
    indexes = {header: index for index, header in enumerate(display)}
    key_index = indexes.get(key_header)
    if key_index is None:
        return [tuple(row) for row in rows]
    output: list[tuple[Any, ...]] = []
    for raw in rows:
        values = list(raw)
        saved = previous.get(str(values[key_index]), {})
        for field in editable_headers:
            if field in indexes and field in saved:
                values[indexes[field]] = saved[field]
        output.append(tuple(values))
    return output


def _carry_coaching_forward(
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    previous: dict[str, dict[str, Any]],
) -> list[tuple[Any, ...]]:
    display = [header.replace("_", " ").title().replace("Id", "ID") for header in headers]
    indexes = {header: index for index, header in enumerate(display)}
    key_index = indexes.get("Coaching Key")
    if key_index is None:
        return [tuple(row) for row in rows]
    output = []
    for raw in rows:
        values = list(raw)
        saved = previous.get(str(values[key_index]), {})
        for field in ("Coaching Status", "Coach", "Coaching Date", "Due Date", "Coaching Comment"):
            if field in indexes and saved.get(field) not in (None, ""):
                values[indexes[field]] = saved[field]
        output.append(tuple(values))
    return output


def _add_pcs_lookups(
    book: DecisionWorkbook,
    dates: Sequence[date],
    scope_rows: Sequence[Sequence[Any]],
) -> None:
    """Hidden chart calculations and cascading selector lists."""
    del scope_rows
    ws = book.report.workbook.add_worksheet("_LOOKUPS")
    headers = [
        "Date", "Q1 Score Sum", "Valid Q1", "Q1 Nonblank", "PCS Status 1",
        "PCS Average", "Participation",
    ]
    for column, header in enumerate(headers):
        ws.write(0, column, header)
    for row_index, business_date in enumerate(dates, 1):
        excel_row = row_index + 1
        ws.write_datetime(row_index, 0, datetime.combine(business_date, datetime.min.time()))
        criteria = (
            f'(tblPcsData[Date]=$A${excel_row})*'
            'IF(DASHBOARD!$K$6="All",1,--(tblPcsData[LOB]=DASHBOARD!$K$6))*'
            'IF(DASHBOARD!$N$6="All",1,--(tblPcsData[Team Leader]=DASHBOARD!$N$6))*'
            'IF(DASHBOARD!$Q$6="All",1,--(tblPcsData[Agent Selector]=DASHBOARD!$Q$6))'
        )
        for column, source in enumerate(
            ("Q1 Score Sum", "Valid Q1", "Q1 Nonblank", "PCS Status 1"), 1,
        ):
            ws.write_formula(
                row_index, column,
                f'=IF(OR($A${excel_row}<PCS_From,$A${excel_row}>PCS_To),NA(),SUMPRODUCT({criteria}*N(tblPcsData[{source}])))',
            )
        ws.write_formula(
            row_index, 5,
            f'=IFERROR($B${excel_row}/$C${excel_row},NA())',
        )
        ws.write_formula(
            row_index, 6,
            f'=IFERROR($D${excel_row}/$E${excel_row},NA())',
        )
    ws.write("J2", "All")
    ws.write_dynamic_array_formula(
        "J3", '=SORT(UNIQUE(FILTER(tblPcsData[LOB],tblPcsData[LOB]<>"","")))',
    )
    ws.write("K2", "All")
    ws.write_dynamic_array_formula(
        "K3",
        '=SORT(UNIQUE(FILTER(tblPcsData[Team Leader],'
        '(tblPcsData[Team Leader]<>"")*IF(DASHBOARD!$K$6="All",1,'
        'tblPcsData[LOB]=DASHBOARD!$K$6),"")))',
    )
    ws.write("L2", "All")
    ws.write_dynamic_array_formula(
        "L3",
        '=SORT(UNIQUE(FILTER(tblPcsData[Agent Selector],'
        '(tblPcsData[Agent Selector]<>"")*IF(DASHBOARD!$K$6="All",1,'
        'tblPcsData[LOB]=DASHBOARD!$K$6)*IF(DASHBOARD!$N$6="All",1,'
        'tblPcsData[Team Leader]=DASHBOARD!$N$6),"")))',
    )
    wb = book.report.workbook
    wb.define_name("PCS_LOB_LIST", "=_LOOKUPS!$J$2:INDEX(_LOOKUPS!$J:$J,COUNTA(_LOOKUPS!$J:$J))")
    wb.define_name("PCS_TL_LIST", "=_LOOKUPS!$K$2:INDEX(_LOOKUPS!$K:$K,COUNTA(_LOOKUPS!$K:$K))")
    wb.define_name("PCS_AGENT_LIST", "=_LOOKUPS!$L$2:INDEX(_LOOKUPS!$L:$L,COUNTA(_LOOKUPS!$L:$L))")
    ws.hide()


def build_pcs_performance_workbook(
    conn: DatabaseConnection,
    config: Config,
    start: date,
    end: date,
    output: Path | None = None,
) -> Path:
    """Build one polished, self-contained PCS workbook with Excel selectors."""

    from .shared_feeds import publish_pcs_feeds

    publish_pcs_feeds(conn, config, start, end)
    book, partial, target = _atomic_book(config, "pcs", "PCS PERFORMANCE", start, end, output)
    latest_value = conn.execute(
        "SELECT max(business_date) FROM mart.agent_pcs_day"
    ).fetchone()[0]
    latest = latest_value or end
    if isinstance(latest, str):
        latest = date.fromisoformat(latest[:10])
    metric_catalog = load_metric_catalog(config.home, config.metric_catalog)
    pcs_method = metric_catalog.method_for("pcs_average", latest, {})
    minimum_sample = int(pcs_method.minimum_sample) if pcs_method is not None else 1
    status, status_text = _source_state(conn, ("fte", "calls"), latest)
    previous_start, previous_end = _previous_month(latest)
    data_start = min(start, previous_start)
    selector_rows = conn.execute(
        """SELECT DISTINCT coalesce(lob,''), coalesce(team_leader,''),
                  coalesce(agent_name,'Agent') || ' [' || agent_id || ']', agent_id
           FROM mart.agent_pcs_day WHERE business_date BETWEEN ? AND ?""",
        [data_start, latest],
    ).fetchall()
    lobs = sorted({str(row[0]) for row in selector_rows if row[0]})
    team_leaders = sorted({str(row[1]) for row in selector_rows if row[1]})
    agents = sorted({str(row[2]) for row in selector_rows if row[2]})
    month_start = latest.replace(day=1)
    default_period = NamedPeriod("Current MTD", month_start, latest)
    default_aggregate = _pcs_aggregate(conn, config, default_period)
    default_values = {
        "pcs_average": default_aggregate[3], "participation": default_aggregate[4],
        "valid": default_aggregate[5], "low": default_aggregate[7],
        "positive": default_aggregate[8], "inbound": default_aggregate[9],
    }
    trend_end = max(latest, date(latest.year, 12, 31))
    trend_dates = [
        data_start + timedelta(days=offset)
        for offset in range((trend_end - data_start).days + 1)
    ] or [latest]
    _add_pcs_dashboard(
        book, status, status_text, start, end, lobs, team_leaders, agents,
        data_start, latest, len(trend_dates), minimum_sample, default_values,
    )
    _add_pcs_team_view(book, minimum_sample)

    agent_headers, agent_raw = _query(
        conn,
        """SELECT agent_id, max(agent_name) AS agent_name,
                  max(team_leader) AS team_leader, max(lob) AS lob,
                  max(language) AS language
           FROM mart.agent_pcs_day WHERE business_date BETWEEN ? AND ?
           GROUP BY agent_id ORDER BY max(lob), max(team_leader), max(agent_name)""",
        [data_start, latest],
    )
    del agent_headers
    agent_rows = []
    for offset, values in enumerate(agent_raw, 5):
        agent_id, agent_name, tl, lob, language = values
        day_score = _pcs_agent_sum_formula(offset, "Q1 Score Sum", None, None)
        day_valid = _pcs_agent_sum_formula(offset, "Valid Q1", None, None)
        selected_score = _pcs_agent_sum_formula(offset, "Q1 Score Sum")
        selected_valid = _pcs_agent_sum_formula(offset, "Valid Q1")
        prior_score = _pcs_agent_sum_formula(offset, "Q1 Score Sum", "PCS_Prior_From", "PCS_Prior_To")
        prior_valid = _pcs_agent_sum_formula(offset, "Valid Q1", "PCS_Prior_From", "PCS_Prior_To")
        participating = _pcs_agent_sum_formula(offset, "Q1 Nonblank")
        eligible = _pcs_agent_sum_formula(offset, "PCS Status 1")
        low = _pcs_agent_sum_formula(offset, "Score <= 3")
        agent_rows.append((
            lob, tl, f"{agent_name or 'Agent'} [{agent_id}]", agent_id,
            agent_name, language,
            f'=IF($C{offset}="","",IF(M{offset}=0,"NO RESPONSE",IF(M{offset}<{minimum_sample},"LOW SAMPLE",IF(O{offset}>0,"COACH","ON TRACK"))))',
            f'=IFERROR({day_score}/{day_valid},"")',
            f'=IFERROR({selected_score}/{selected_valid},"")',
            f'=IFERROR({prior_score}/{prior_valid},"")',
            f'=IF(OR(I{offset}="",J{offset}=""),"",I{offset}-J{offset})',
            f'=IFERROR({participating}/{eligible},"")',
            f'={selected_valid}', f'={eligible}', f'={low}',
            f'=IF($C{offset}="","",IF(O{offset}>0,"Open COACHING",IF(M{offset}=0,"Check sample","Monitor")))',
        ))
    agent_result_headers = [
        "lob", "team_leader", "agent_selector", "agent_id", "agent_name",
        "language", "priority",
        "latest_day_average", "selected_period_average", "prior_mtd_average", "movement",
        "participation_rate", "valid_q1", "pcs_status_1", "score_le_3", "next_action",
    ]
    ws = book.table(
        "AGENT_RESULTS", "PCS agent realizations",
        "Use a personal Sheet View, then filter LOB, Team Leader or Agent Selector. Results follow the period selected on DASHBOARD and update from PCS_DATA.",
        agent_result_headers, agent_rows or [tuple(None for _ in agent_result_headers)],
    )
    if agent_rows:
        priority_column = agent_result_headers.index("priority")
        ws.conditional_format(4, priority_column, 3 + len(agent_rows), priority_column, {
            "type": "text", "criteria": "containing", "value": "COACH", "format": book.report.error,
        })

    action_headers, actions = _pcs_coaching_rows(conn, config, data_start, end)
    queue_source = list(actions)
    actions = _carry_coaching_forward(
        action_headers, actions, _previous_coaching_values(target),
    )
    action_rows = actions or [tuple(None for _ in action_headers)]
    action_sheet = book.table(
        "COACHING",
        "PCS coaching action plan",
        "Use a personal Sheet View, filter LOB, Team Leader or Agent Selector, and fill only the five blue action columns.",
        action_headers,
        action_rows,
        editable_headers={"Coaching Status", "Coaching Date", "Coach", "Due Date", "Coaching Comment"},
    )
    if action_rows:
        status_col = action_headers.index("coaching_status")
        action_sheet.data_validation(
            4, status_col, 3 + len(action_rows), status_col,
            {
                "validate": "list",
                "source": ["Pending", "Planned", "Completed", "Not required"],
                "input_title": "Coaching status",
                "input_message": "Choose one of the four action statuses.",
                "error_title": "Invalid status",
                "error_message": "Use Pending, Planned, Completed or Not required.",
            },
        )
        action_sheet.conditional_format(
            4, status_col, 3 + len(action_rows), status_col,
            {"type": "text", "criteria": "containing", "value": "Pending", "format": book.report.error},
        )
        for hidden_header in ("ops_manager",):
            if hidden_header in action_headers:
                column = action_headers.index(hidden_header)
                action_sheet.set_column(column, column, None, None, {"hidden": True})

    queue_headers = [
        "lob", "team_leader", "agent_selector", "agent_name", "agent_id",
        "priority", "business_date", "call_start", "q1_score",
        "customer_comment", "call_reference_number", "language",
        "coaching_key", "action_status",
    ]
    queue_indexes = {header: index for index, header in enumerate(action_headers)}
    queue_rows = []
    for excel_row, values in enumerate(queue_source, 5):
        queue_rows.append(tuple(
            values[queue_indexes[header]] if header != "action_status" else
            '=IFERROR(XLOOKUP([@[Coaching Key]],tblCoaching[Coaching Key],tblCoaching[Coaching Status]),"Not started")'
            for header in queue_headers
        ))
    queue_sheet = book.table(
        "COACHING_QUEUE", "PCS coaching opportunity queue",
        "Filter LOB, Team Leader or Agent Selector. Action Status reads the shared COACHING action plan by Coaching Key.",
        queue_headers, queue_rows or [tuple(None for _ in queue_headers)],
    )
    if queue_rows:
        queue_sheet.conditional_format(
            4, queue_headers.index("action_status"),
            3 + len(queue_rows), queue_headers.index("action_status"),
            {"type": "text", "criteria": "containing", "value": "Not started", "format": book.report.error},
        )
    data_headers, data_rows = _query(
        conn,
        """SELECT lob, team_leader,
                  coalesce(agent_name,'Agent') || ' [' || agent_id || ']' AS agent_selector,
                  agent_id, agent_name, business_date, ops_manager,
                  language, inbound_calls AS inbound_call_legs,
                  pcs_status_calls AS pcs_status_1,
                  pcs_participation_responses AS q1_nonblank,
                  survey_responses AS valid_q1,
                  pcs_score_sum AS q1_score_sum,
                  pcs_average, pcs_participation_rate AS participation_rate,
                  low_score_responses AS score_le_3,
                  top_box_responses AS score_gt_3,
                  pcs_invalid_responses AS invalid_q1,
                  CASE WHEN survey_responses<? THEN 'LOW_SAMPLE' ELSE 'OK' END AS sample_state
           FROM mart.agent_pcs_day WHERE business_date BETWEEN ? AND ?
           ORDER BY business_date, lob, team_leader, agent_name""",
        [minimum_sample, data_start, latest],
    )
    book.table("PCS_DATA", "PCS clean calculation table", "One agent/day. Use a personal Sheet View before filtering LOB, Team Leader or Agent Selector.", data_headers, data_rows or [tuple(None for _ in data_headers)])
    book.table(
        "HELP", "PCS shared workbook guide",
        "Keep this workbook as the team file. Refresh its data; do not replace the workbook while coaching notes are in progress.",
        ["Step", "What to do", "What changes", "Important"],
        [
            (1, "Run WFMHub > Refresh source data once > Agent PCS", "The database and fixed PCS CSV feeds are updated", "Original extracts are never changed"),
            (2, f"Use {config.feed / 'PCS' / 'PCS_AGENT_DAY_CURRENT.csv'} as the PCS_DATA source", "Dashboard, period cards and agent formulas recalculate", "Keep the PCS_DATA column names unchanged"),
            (3, "Use Data > Refresh All after the one-time Power Query link; until then replace only the rows inside PCS_DATA", "Latest day, Current week and Current MTD follow the latest Date", "Do not paste over the header row"),
            (4, "Refresh COACHING_QUEUE from its fixed CSV; copy columns A:M for each new case into the next COACHING row", "The same Coaching Key cannot be confused with another call", "Do not copy the Action Status formula"),
            (5, "Quality works only in the five blue COACHING columns", "Status, coach, dates and comments remain in the shared workbook", "Use a personal Sheet View before filtering"),
            (6, "Agent ID is the matching key; Agent Selector is Name [ID] for people", "Duplicate names remain separate", "Never match agents by display name alone"),
        ],
    )
    book.definitions([
        ("PCS Average", "Sum of valid inbound Q1 scores / valid inbound Q1 responses", "Customer experience result", "Only configured discrete Q1 scores are valid"),
        ("PCS Participation", "Inbound raw Q1 nonblank / inbound PCSStatus=1", "Survey participation opportunity", "Invalid nonblank Q1 remains in the numerator"),
        ("Score <= 3", "Count of valid Q1 responses at or below 3", "Follow-up volume", "A count, not a percentage"),
        ("Actions Rate", "Completed rows in COACHING / valid Q1 responses at or below 3", "Coaching completion", "Saved notes are matched by Coaching Key"),
        ("Low sample", f"Fewer than {minimum_sample} valid responses in the selected period", "Interpretation warning", "Use a larger sample before drawing conclusions"),
        ("Selectors", "LOB narrows Team Leader; LOB and Team Leader narrow Agent", "Management view", "Change an upstream selector if a previous choice is no longer valid"),
        ("Agent realizations", "Latest day, selected period and previous-month same-days at agent grain", "TL action list", "Filter Team Leader directly on the AGENT_RESULTS table"),
    ])
    _add_pcs_lookups(book, trend_dates, selector_rows)
    book.audit(_audit_rows(conn, config, "pcs", start, end))
    return _finish(book, partial, target)


def _service_rows(
    conn: DatabaseConnection,
    profile: ServiceProfile,
    start: date,
    end: date,
) -> list[dict[str, Any]]:
    scope_marks = ",".join("?" for _ in profile.service_scopes)
    system_marks = ",".join("?" for _ in profile.source_systems)
    cursor = conn.execute(
        f"""SELECT business_date, interval_start, hour_start, source_system, queue,
                   service_scope, comparison_scope, designation, mapping_status,
                   offered, answered, abandoned, short_abandoned,
                   answered_within_target, handled_seconds, source_file
            FROM mart.service_interval
            WHERE business_date BETWEEN ? AND ?
              AND service_scope IN ({scope_marks})
              AND source_system IN ({system_marks})
            ORDER BY business_date, interval_start, source_system, queue""",
        [start, end, *profile.service_scopes, *profile.source_systems],
    )
    headers = [item[0] for item in cursor.description]
    return [dict(zip(headers, row)) for row in cursor.fetchall()]


def _profile_metric(catalog: MetricCatalog, profile: ServiceProfile, metric_id: str, on_date: date):
    methods = {}
    for service_scope in profile.service_scopes:
        for source_system in profile.source_systems:
            method = catalog.method_for(metric_id, on_date, {
                "lob": service_scope,
                "source_system": source_system,
            })
            if method is None:
                raise ValueError(
                    f"Service profile {profile.profile_id!r} selects metric {metric_id!r}, "
                    f"but no method applies to {service_scope}/{source_system} on {on_date}"
                )
            methods[(method.method_id, method.effective_from, method.priority)] = method
    if len(methods) != 1:
        names = ", ".join(key[0] for key in methods)
        raise ValueError(
            f"Service profile {profile.profile_id!r} crosses incompatible {metric_id} methods: {names}. "
            "Split it into separate service profiles."
        )
    return next(iter(methods.values()))


def _service_aggregate(
    rows: Iterable[dict[str, Any]],
    profile: ServiceProfile,
    catalog: MetricCatalog,
    on_date: date,
) -> dict[str, float | str | None]:
    values = list(rows)
    offered = sum(float(row.get("offered") or 0) for row in values)
    answered = sum(float(row.get("answered") or 0) for row in values)
    abandoned = sum(float(row.get("abandoned") or 0) for row in values)
    short = sum(float(row.get("short_abandoned") or 0) for row in values)
    within = sum(float(row.get("answered_within_target") or 0) for row in values)
    handled = sum(float(row.get("handled_seconds") or 0) for row in values)
    components = {
        "offered": offered,
        "answered": answered,
        "abandoned": abandoned,
        "short_abandoned": short,
        "answered_within_target": within,
        "handled_seconds": handled,
    }
    service_level = evaluate_metric(
        _profile_metric(catalog, profile, profile.service_level_metric, on_date), components,
    )
    availability = evaluate_metric(
        _profile_metric(catalog, profile, profile.availability_metric, on_date), components,
    )
    aht = evaluate_metric(
        _profile_metric(catalog, profile, profile.aht_metric, on_date), components,
    )
    return {
        "offered": offered,
        "answered": answered,
        "short_abandoned": short,
        "within_target": within,
        "service_level": service_level.value,
        "service_target": service_level.method.target,
        "service_state": service_level.state,
        "service_method": service_level.method.method_id,
        "availability": availability.value,
        "aht_seconds": aht.value,
    }


def build_service_performance_workbook(
    conn: DatabaseConnection,
    config: Config,
    start: date,
    end: date,
    output: Path | None = None,
    profile_id: str | None = None,
) -> Path:
    """Build LOB service performance, beginning with the Ford OEM profile."""

    catalog = load_service_profiles(config.home, config.service_profiles)
    profile = catalog.select(profile_id, end)
    metric_catalog = load_metric_catalog(config.home, config.metric_catalog)
    rulebook = load_rulebook(config.home, config.business_rules)
    book, partial, target = _atomic_book(config, "service", f"OEM FLASH  /  {profile.label.upper()}", start, end, output)
    periods = _comparison_periods(start, end)
    all_start, all_end = min(p.start for p in periods), max(p.end for p in periods)
    raw_rows = _service_rows(conn, profile, all_start, all_end)
    period_rows = []
    for period in periods:
        scoped = [row for row in raw_rows if period.start.isoformat() <= str(row["business_date"])[:10] <= period.end.isoformat()]
        metrics = _service_aggregate(scoped, profile, metric_catalog, period.end)
        period_rows.append((period.label, period.start, period.end, metrics["offered"], metrics["answered"], metrics["service_level"], metrics["availability"], metrics["aht_seconds"]))
    by_label = {row[0]: row for row in period_rows}
    latest, mtd, prior = by_label["Latest day"], by_label["Current MTD"], by_label["Prior-month same days"]
    latest_method = _profile_metric(metric_catalog, profile, profile.service_level_metric, end)
    forecast_marks = ",".join("?" for _ in profile.service_scopes)
    forecast_total, forecast_rows_count = conn.execute(
        f"""SELECT sum(volume_forecast), count(*) FROM mart.forecast_hour
            WHERE business_date=? AND service_scope IN ({forecast_marks})""",
        [end, *profile.service_scopes],
    ).fetchone()
    if not forecast_rows_count:
        forecast_total = None
    forecast_method = metric_catalog.method_for(
        "forecast_attainment", end, {"lob": profile.service_scopes[0]},
    )
    forecast_evaluation = evaluate_metric(
        forecast_method, {"actual_volume": latest[3], "forecast_volume": forecast_total},
    ) if forecast_method and forecast_total is not None else None
    families = tuple(system.lower() for system in profile.source_systems) + ("forecast",)
    status, status_text = _source_state(conn, families, end)
    latest_raw = [
        row for row in raw_rows
        if str(row["business_date"])[:10] == end.isoformat()
    ]
    group_metrics: dict[str, dict[str, float | str | None]] = {}
    for group in profile.groups:
        group_metrics[group.label] = _service_aggregate(
            [row for row in latest_raw if profile.group_for(row.get("queue")) == group.label],
            profile,
            metric_catalog,
            end,
        )

    def group_value(label: str, metric: str) -> float | str | None:
        return group_metrics.get(label, {}).get(metric)

    variance_calls = (
        float(latest[3] or 0) - float(forecast_total)
        if forecast_total is not None else None
    )
    book.dashboard(
        [
            KpiCard("OEM availability", latest[6], "percent", "Answered / offered"),
            KpiCard("OEM service level", latest[5], "percent", f"Target {latest_method.target:.0%}" if latest_method.target is not None else "No target"),
            KpiCard("Ford availability", group_value("Ford", "availability"), "percent", "Mapped Ford queues"),
            KpiCard("Ford service level", group_value("Ford", "service_level"), "percent", "Mapped Ford queues"),
            KpiCard("Toyota availability", group_value("Toyota", "availability"), "percent", "Mapped Toyota/Lexus queues"),
            KpiCard("Toyota service level", group_value("Toyota", "service_level"), "percent", "Mapped Toyota/Lexus queues"),
            KpiCard("Forecast attainment", forecast_evaluation.value if forecast_evaluation else None, "percent", "Actual offered / forecast"),
            KpiCard("Variance calls", variance_calls, "integer", "Actual offered - forecast"),
        ],
        status,
        status_text,
        ["Period", "Start", "End", "Offered", "Answered", "Service Level %", "Availability %", "AHT Seconds"],
        period_rows,
        [
            "This report includes only queues mapped into the selected service profile; queue membership is controlled in queue_mapping.csv.",
            f"{profile.label}: the extract's within-{rulebook.target_seconds}s counter is evaluated by {profile.service_level_metric}.{latest_method.method_id} from metric_catalog.toml.",
            "Service availability is answered / offered. It is never agent availability or adherence.",
            "Forecast contributes forecast only. APBE/APFR/APDE contribute actual performance only.",
            "Back-office workload stays NOT CONFIGURED until its source and calculation are agreed.",
        ],
        (("Service Level", 5), ("Availability", 6)),
        sheet_name="FLASH",
    )

    hourly: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in raw_rows:
        if start.isoformat() <= str(row["business_date"])[:10] <= end.isoformat():
            group = profile.group_for(row.get("queue"))
            hourly[(str(row["business_date"])[:10], str(row["hour_start"]), group)].append(row)
    forecast_rows = conn.execute(
        f"""SELECT hour_start, sum(volume_forecast)
            FROM mart.forecast_hour
            WHERE business_date BETWEEN ? AND ?
              AND service_scope IN ({forecast_marks})
            GROUP BY hour_start""",
        [start, end, *profile.service_scopes],
    ).fetchall()
    forecast_by_hour = {str(row[0])[:13]: float(row[1] or 0) for row in forecast_rows}
    staffing_marks = ",".join("?" for _ in profile.service_scopes)
    staffing_rows = conn.execute(
        f"""SELECT business_date, hour_key,
                   avg(scheduled_fte), avg(observed_fte), avg(productive_fte)
            FROM (
                SELECT business_date, substr(cast(interval_start AS TEXT),1,13) AS hour_key,
                       interval_start, sum(scheduled_fte) AS scheduled_fte,
                       sum(observed_fte) AS observed_fte,
                       sum(productive_fte) AS productive_fte
                FROM mart.staffing_interval
                WHERE business_date BETWEEN ? AND ?
                  AND lob IN ({staffing_marks})
                GROUP BY business_date, interval_start
            ) staffing
            GROUP BY business_date, hour_key""",
        [start, end, *profile.service_scopes],
    ).fetchall()
    staffing_by_hour = {
        f"{str(row[0])[:10]}T{str(row[1])[-2:]}": (
            float(row[2] or 0), float(row[3] or 0), float(row[4] or 0),
        )
        for row in staffing_rows
    }
    grouped_by_hour: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in raw_rows:
        if start.isoformat() <= str(row["business_date"])[:10] <= end.isoformat():
            grouped_by_hour[(str(row["business_date"])[:10], str(row["hour_start"]))].append(row)

    hourly_rows = []
    for (business_date, hour_start), rows in sorted(grouped_by_hour.items()):
        row_date = date.fromisoformat(business_date)
        total = _service_aggregate(rows, profile, metric_catalog, row_date)
        hour_key = str(hour_start)[:13]
        forecast = forecast_by_hour.get(hour_key)
        attainment = _ratio(total["offered"], forecast)
        scheduled, observed, productive = staffing_by_hour.get(
            f"{business_date}T{hour_key[-2:]}", (None, None, None),
        )
        hourly_rows.append((
            business_date, hour_start, "OEM TOTAL", "OEM Total",
            total["offered"], forecast,
            float(total["offered"] or 0) - forecast if forecast is not None else None,
            attainment, total["answered"], total["within_target"],
            total["short_abandoned"], total["service_level"], total["availability"],
            total["aht_seconds"], scheduled, observed, productive,
            total["service_method"], total["service_state"],
        ))
    for (business_date, hour_start, group), rows in sorted(hourly.items()):
        row_date = date.fromisoformat(business_date)
        metrics = _service_aggregate(rows, profile, metric_catalog, row_date)
        hourly_rows.append((
            business_date, hour_start, "QUEUE GROUP", group,
            metrics["offered"], None, None, None, metrics["answered"],
            metrics["within_target"], metrics["short_abandoned"],
            metrics["service_level"], metrics["availability"], metrics["aht_seconds"],
            None, None, None, metrics["service_method"], metrics["service_state"],
        ))
    hourly_rows.sort(key=lambda row: (str(row[0]), str(row[1]), 0 if row[2] == "OEM TOTAL" else 1, str(row[3])))
    headers = [
        "business_date", "hour_start", "row_type", "service_group", "offered",
        "forecast", "variance_calls", "forecast_attainment", "answered",
        "answered_within_target", "short_abandoned", "service_level",
        "service_availability", "aht_seconds", "scheduled_fte", "observed_fte",
        "productive_fte", "service_method", "service_state",
    ]
    ws = book.table("HOURLY", f"{profile.label} hourly control", "OEM total rows reconcile forecast and staffing; Ford / Toyota / Chery rows show mapped queue-group performance.", headers, hourly_rows)
    if hourly_rows:
        state_col = headers.index("service_state")
        ws.conditional_format(4, state_col, 3 + len(hourly_rows), state_col, {"type": "text", "criteria": "containing", "value": "BELOW_TARGET", "format": book.report.error})
    detail_headers = ["business_date", "interval_start", "source_system", "queue", "service_scope", "designation", "mapping_status", "offered", "answered", "abandoned", "short_abandoned", "answered_within_target", "handled_seconds", "source_file"]
    detail_rows = [tuple(row.get(header) for header in detail_headers) for row in raw_rows if start.isoformat() <= str(row["business_date"])[:10] <= end.isoformat()]
    book.table("QUEUES", "Mapped queue evidence", "Compact queue intervals for reconciliation; no original extract rows are modified.", detail_headers, detail_rows)
    action_rows = [row for row in hourly_rows if row[-1] == "BELOW_TARGET"]
    book.table("EXCEPTIONS", "Service intervals below target", "Prioritize intervals with high offered volume and validate staffing/forecast before escalation.", headers, action_rows)
    source_headers, source_rows = _query(
        conn,
        """SELECT source_family, newest_business_date, row_count, scoped_out_count,
                  status, newest_file, details
           FROM mart.source_health
           WHERE lower(source_family) IN ('apbe','apfr','apde','forecast','fte','schedule','lilo','agent_status')
           ORDER BY source_family""",
    )
    source_rows.append((
        "back_office", None, None, None, "NOT_CONFIGURED", None,
        "The reference workbook contains manually sourced backlog counters; their source is not configured yet.",
    ))
    book.table("DATA_STATUS", "Flash source status", "The Flash stays explicit about missing or stale inputs.", source_headers, source_rows)
    book.definitions([
        ("Service Level", f"({latest_method.numerator}) / ({latest_method.denominator})", "Contract performance", f"Method {latest_method.method_id}; ratio of summed counters"),
        ("Service Availability", "Answered / offered", "Ability of the service to answer demand", "Not an agent metric"),
        ("Forecast Attainment", "Actual offered / forecast", "Demand realization", "Forecast and actual joined only by mapped comparison scope"),
        ("AHT", "Total handled seconds / answered contacts", "Workload driver", "Weighted, never average of interval AHT"),
    ])
    book.audit(_audit_rows(conn, config, "service", start, end, [
        ("Service profile", profile.profile_id, f"{catalog.version} / {catalog.sha256}"),
        ("Included scopes", ", ".join(profile.service_scopes), ", ".join(profile.source_systems)),
        ("Service metric", profile.service_level_metric, f"{metric_catalog.version} / {metric_catalog.sha256}"),
    ]))
    return _finish(book, partial, target)


def build_realisations_workbook(
    conn: DatabaseConnection,
    config: Config,
    start: date,
    end: date,
    output: Path | None = None,
    profile_id: str | None = None,
) -> Path:
    """Build one management view across every configured service LOB.

    Supplying ``profile_id`` keeps the focused single-LOB route available for
    command-line use.  The normal menu intentionally includes all active
    profiles so Operations does not need to build four separate workbooks.
    """

    profiles = load_service_profiles(config.home, config.service_profiles)
    selected_profiles = (
        (profiles.select(profile_id, end),)
        if profile_id
        else tuple(profile for profile in profiles.profiles if profile.active_on(end))
    )
    metrics_catalog = load_metric_catalog(config.home, config.metric_catalog)
    title = (
        f"REALISATIONS  /  {selected_profiles[0].label.upper()}"
        if len(selected_profiles) == 1
        else "REALISATIONS  /  ALL MAPPED LOBS"
    )
    book, partial, target = _atomic_book(
        config, "realisations", title, start, end, output,
    )
    daily_rows: list[tuple[Any, ...]] = []
    all_source_rows: list[dict[str, Any]] = []
    profile_by_label = {profile.label: profile for profile in selected_profiles}

    for profile in selected_profiles:
        source_rows = _service_rows(conn, profile, start, end)
        for row in source_rows:
            all_source_rows.append({**row, "reporting_lob": profile.label})
        forecast_marks = ",".join("?" for _ in profile.service_scopes)
        forecasts = conn.execute(
            f"""SELECT business_date, sum(volume_forecast), sum(fte_forecast),
                       sum(fte_required), avg(sl_forecast), avg(sl_required),
                       CASE WHEN sum(volume_forecast)>0
                            THEN sum(aht_forecast_seconds*volume_forecast)/sum(volume_forecast) END
                FROM mart.forecast_hour
                WHERE business_date BETWEEN ? AND ?
                  AND service_scope IN ({forecast_marks})
                GROUP BY business_date ORDER BY business_date""",
            [start, end, *profile.service_scopes],
        ).fetchall()
        forecast_by_day = {str(row[0])[:10]: row[1:] for row in forecasts}
        staffing_marks = ",".join("?" for _ in profile.staffing_lobs)
        staffing_rows = conn.execute(
            f"""SELECT business_date, avg(scheduled_fte), avg(observed_fte),
                       avg(productive_fte), max(staffing_gap_fte)
                FROM (
                    SELECT business_date, interval_start,
                           sum(scheduled_fte) AS scheduled_fte,
                           sum(observed_fte) AS observed_fte,
                           sum(productive_fte) AS productive_fte,
                           sum(staffing_gap_fte) AS staffing_gap_fte
                    FROM mart.staffing_interval
                    WHERE business_date BETWEEN ? AND ? AND lob IN ({staffing_marks})
                    GROUP BY business_date, interval_start
                ) x GROUP BY business_date""",
            [start, end, *profile.staffing_lobs],
        ).fetchall()
        staffing_by_day = {str(row[0])[:10]: row[1:] for row in staffing_rows}
        absence_rows = conn.execute(
            f"""SELECT business_date, sum(planned_net_minutes)/60.0,
                       sum(final_absence_minutes)/60.0,
                       sum(final_vacation_minutes)/60.0,
                       sum(final_shrinkage_minutes)/60.0,
                       sum(CASE WHEN final_ledger_status NOT IN ('CLEAR','ABSENCE_RECORDED')
                                THEN 1 ELSE 0 END)
                FROM mart.verint_final_absence_agent_day
                WHERE business_date BETWEEN ? AND ? AND lob IN ({staffing_marks})
                GROUP BY business_date""",
            [start, end, *profile.staffing_lobs],
        ).fetchall()
        absence_by_day = {str(row[0])[:10]: row[1:] for row in absence_rows}
        rows_by_day: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in source_rows:
            rows_by_day[str(row["business_date"])[:10]].append(row)

        cursor = start
        while cursor <= end:
            key = cursor.isoformat()
            actual = _service_aggregate(
                rows_by_day.get(key, []), profile, metrics_catalog, cursor,
            )
            forecast = forecast_by_day.get(key, (None, None, None, None, None, None))
            staffing = staffing_by_day.get(key, (None, None, None, None))
            absence_values = absence_by_day.get(key, (None, None, None, None, 0))
            forecast_volume = forecast[0]
            planned_hours, absence_hours, vacation_hours, shrinkage_hours, review_cases = absence_values
            has_actual = bool(rows_by_day.get(key))
            has_forecast = forecast_volume is not None
            data_state = (
                "Provisional" if cursor >= date.today()
                else "Review" if review_cases
                else "No actual" if not has_actual
                else "No forecast" if not has_forecast
                else "Final"
            )
            daily_rows.append((
                cursor, cursor.strftime("%Y-%m"), cursor.isocalendar().week,
                f"Q{((cursor.month - 1) // 3) + 1}", profile.label,
                ", ".join(profile.service_scopes), actual["offered"],
                forecast_volume,
                float(actual["offered"] or 0) - float(forecast_volume)
                if forecast_volume is not None else None,
                _ratio(actual["offered"], forecast_volume), actual["answered"],
                actual["within_target"], actual["short_abandoned"],
                actual["service_level"], actual["availability"],
                actual["aht_seconds"],
                _ratio(
                    float(actual["aht_seconds"] or 0) * float(actual["answered"] or 0),
                    3600,
                ),
                staffing[0], staffing[1], staffing[2], staffing[3],
                planned_hours, absence_hours, _ratio(absence_hours, planned_hours),
                vacation_hours, shrinkage_hours, _ratio(shrinkage_hours, planned_hours),
                review_cases, data_state, profile.profile_id,
                ", ".join(profile.staffing_lobs),
            ))
            cursor += timedelta(days=1)
    daily_headers = [
        "business_date", "month", "iso_week", "quarter", "lob",
        "service_scopes", "actual_volume", "forecast_volume",
        "variance_calls", "forecast_attainment", "answered",
        "answered_within_target", "short_abandoned", "service_level",
        "service_availability", "aht_seconds", "processing_hours",
        "scheduled_fte", "observed_fte", "productive_fte", "peak_gap_fte",
        "planned_hours", "absence_hours", "absence_rate", "vacation_hours",
        "shrinkage_hours", "shrinkage_rate", "review_cases", "data_state",
        "profile_id", "staffing_lobs",
    ]

    total_actual = sum(float(row[6] or 0) for row in daily_rows)
    total_forecast = sum(float(row[7] or 0) for row in daily_rows if row[7] is not None)
    forecast_present = any(row[7] is not None for row in daily_rows)
    total_answered = sum(float(row[10] or 0) for row in daily_rows)
    total_within = sum(float(row[11] or 0) for row in daily_rows)
    total_short = sum(float(row[12] or 0) for row in daily_rows)
    total_handled_seconds = sum(
        float(row[15] or 0) * float(row[10] or 0) for row in daily_rows
    )
    total_planned = sum(float(row[21] or 0) for row in daily_rows)
    total_absence = sum(float(row[22] or 0) for row in daily_rows)
    total_shrinkage = sum(float(row[25] or 0) for row in daily_rows)
    availability_value = _ratio(total_answered, total_actual)
    aht_value = _ratio(total_handled_seconds, total_answered)
    profile_summary: list[tuple[Any, ...]] = []
    for profile in selected_profiles:
        rows = [row for row in daily_rows if row[4] == profile.label]
        offered = sum(float(row[6] or 0) for row in rows)
        forecast = sum(float(row[7] or 0) for row in rows if row[7] is not None)
        answered = sum(float(row[10] or 0) for row in rows)
        within = sum(float(row[11] or 0) for row in rows)
        short = sum(float(row[12] or 0) for row in rows)
        handled = sum(float(row[15] or 0) * float(row[10] or 0) for row in rows)
        planned_hours = sum(float(row[21] or 0) for row in rows)
        absence_hours = sum(float(row[22] or 0) for row in rows)
        shrinkage_hours = sum(float(row[25] or 0) for row in rows)
        profile_components = {
            "offered": offered, "answered": answered,
            "short_abandoned": short, "answered_within_target": within,
            "handled_seconds": handled,
        }
        sl_result = evaluate_metric(
            _profile_metric(metrics_catalog, profile, profile.service_level_metric, end),
            profile_components,
        )
        availability_result = evaluate_metric(
            _profile_metric(metrics_catalog, profile, profile.availability_metric, end),
            profile_components,
        )
        aht_result = evaluate_metric(
            _profile_metric(metrics_catalog, profile, profile.aht_metric, end),
            profile_components,
        )
        state = (
            "NO ACTUAL" if not any(float(row[6] or 0) for row in rows)
            else "NO FORECAST" if not any(row[7] is not None for row in rows)
            else "REVIEW" if any(row[-4] for row in rows)
            else "READY"
        )
        profile_summary.append((
            profile.label, offered,
            forecast if any(row[7] is not None for row in rows) else None,
            _ratio(offered, forecast), sl_result.value, sl_result.method.target,
            availability_result.value, aht_result.value,
            _ratio(absence_hours, planned_hours),
            _ratio(shrinkage_hours, planned_hours), state,
        ))
    status, status_text = _source_state(
        conn,
        tuple(sorted({
            system.lower()
            for profile in selected_profiles
            for system in profile.source_systems
        })) + ("forecast",),
        end,
    )
    review_days = sum(1 for row in daily_rows if row[-4])
    if review_days:
        status, status_text = "INCOMPLETE", f"{review_days:,} day(s) include absence review cases"
    book.dashboard(
        [
            KpiCard("Actual volume", total_actual, "integer"),
            KpiCard("Forecast volume", total_forecast if forecast_present else None, "integer"),
            KpiCard("Forecast attainment", _ratio(total_actual, total_forecast) if forecast_present else None, "percent"),
            KpiCard("Mapped LOBs", len(selected_profiles), "integer"),
            KpiCard("Service availability", availability_value, "percent"),
            KpiCard("Weighted AHT", aht_value, "decimal"),
            KpiCard("Absence rate", _ratio(total_absence, total_planned), "percent"),
            KpiCard("Shrinkage rate", _ratio(total_shrinkage, total_planned), "percent"),
        ],
        status,
        status_text,
        [
            "LOB", "Actual", "Forecast", "Attainment %", "Service Level %",
            "SL Target %", "Availability %", "AHT Seconds", "Absence Rate %",
            "Shrinkage Rate %", "State",
        ],
        profile_summary,
        [
            "Queue membership is maintained in Queue Mapping; service and roster LOB links are maintained in Service Profiles.",
            "Forecast comes from Verint. Actual volume, service level and AHT come from the mapped Storm performance sources.",
            "Service availability means answered contacts divided by offered contacts. It is not agent availability.",
            "Absence uses final Verint results for completed days; provisional or review cases remain visible.",
            "Adherence is intentionally excluded.",
        ],
        (("Service Level", 4), ("Availability", 6)),
    )
    book.table(
        "LOB_RESULTS", "Daily LOB results",
        "One normalized row per mapped management LOB and day. Use this sheet for pivots, charts and management checks.",
        daily_headers, daily_rows,
    )

    trend_rows: list[tuple[Any, ...]] = []
    for grain, index in (("Month", 1), ("ISO Week", 2), ("Quarter", 3)):
        grouped: dict[tuple[str, Any], list[tuple[Any, ...]]] = defaultdict(list)
        for row in daily_rows:
            grouped[(str(row[4]), row[index])].append(row)
        for (lob_label, label), group in sorted(grouped.items(), key=lambda item: (item[0][0], str(item[0][1]))):
            offered = sum(float(row[6] or 0) for row in group)
            forecast = sum(float(row[7] or 0) for row in group if row[7] is not None)
            answered = sum(float(row[10] or 0) for row in group)
            within = sum(float(row[11] or 0) for row in group)
            short = sum(float(row[12] or 0) for row in group)
            handled = sum(float(row[15] or 0) * float(row[10] or 0) for row in group)
            planned_hours = sum(float(row[21] or 0) for row in group)
            absence_hours = sum(float(row[22] or 0) for row in group)
            shrinkage_hours = sum(float(row[25] or 0) for row in group)
            trend_components = {
                "offered": offered, "answered": answered,
                "short_abandoned": short, "answered_within_target": within,
                "handled_seconds": handled,
            }
            trend_profile = profile_by_label[lob_label]
            trend_sl = evaluate_metric(
                _profile_metric(metrics_catalog, trend_profile, trend_profile.service_level_metric, end),
                trend_components,
            ).value
            trend_rows.append((
                lob_label, grain, label, min(row[0] for row in group), max(row[0] for row in group),
                offered, forecast if any(row[7] is not None for row in group) else None,
                _ratio(offered, forecast) if forecast else None,
                trend_sl, _ratio(answered, offered),
                _ratio(handled, answered), planned_hours, absence_hours,
                _ratio(absence_hours, planned_hours), shrinkage_hours,
                _ratio(shrinkage_hours, planned_hours),
            ))
    book.table(
        "TREND", "Period trend",
        "Month, ISO week and quarter summaries calculated from additive daily counters.",
        [
            "lob", "grain", "period", "start", "end", "actual_volume",
            "forecast_volume", "forecast_attainment", "service_level",
            "service_availability", "aht_seconds", "planned_hours",
            "absence_hours", "absence_rate", "shrinkage_hours", "shrinkage_rate",
        ],
        trend_rows,
    )
    detail_headers = [
        "reporting_lob", "business_date", "interval_start", "source_system", "queue",
        "service_scope", "designation", "mapping_status", "offered", "answered",
        "abandoned", "short_abandoned", "answered_within_target",
        "handled_seconds", "source_file",
    ]
    detail_rows = [tuple(row.get(header) for header in detail_headers) for row in all_source_rows]
    book.table(
        "DATA", "Mapped service data",
        "Filterable queue and interval evidence behind the LOB results.",
        detail_headers, detail_rows,
    )
    book.definitions([
        ("Actual / forecast", "Actual offered contacts / forecast contacts", "Demand realisation", "Use summed volumes"),
        ("Service level", "Configured numerator / denominator for each service profile", "Service performance", "Calculated from summed counters; never average LOB percentages"),
        ("Service availability", "Answered / offered", "Ability of the service to answer demand", "Not agent availability"),
        ("Weighted AHT", "Handled seconds / answered contacts", "Workload", "Never average daily AHT values"),
        ("Absence rate", "Final absence hours / planned hours", "Capacity impact", "Completed-day Verint result"),
    ])
    book.audit(_audit_rows(conn, config, "realisations", start, end, [
        ("Service profiles", ", ".join(profile.profile_id for profile in selected_profiles), profiles.version),
        ("Included service scopes", "; ".join(
            f"{profile.label}: {', '.join(profile.service_scopes)}"
            for profile in selected_profiles
        ), "Queue Mapping"),
        ("Included staffing LOBs", "; ".join(
            f"{profile.label}: {', '.join(profile.staffing_lobs)}"
            for profile in selected_profiles
        ), "Service Profiles"),
    ]))
    return _finish(book, partial, target)


def build_attendance_today_workbook(
    conn: DatabaseConnection,
    config: Config,
    start: date,
    end: date,
    output: Path | None = None,
) -> Path:
    """Build the attendance callout product for the full selected period."""

    book, partial, target = _atomic_book(
        config, "attendance", "ATTENDANCE CALLOUTS", start, end, output,
    )
    totals = conn.execute(
        """SELECT count(*),
                  coalesce(sum(CASE WHEN requires_call THEN 1 ELSE 0 END),0),
                  coalesce(sum(CASE WHEN call_action='CALL_NO_SHOW' THEN 1 ELSE 0 END),0),
                  coalesce(sum(CASE WHEN call_action='CALL_LATE' THEN 1 ELSE 0 END),0),
                  coalesce(sum(CASE WHEN call_action='CALL_NOT_SEEN_NOW' THEN 1 ELSE 0 END),0),
                  coalesce(sum(CASE WHEN attendance_result IN
                    ('Schedule parse error','Data not loaded','Missing actual evidence',
                     'Incomplete actual evidence','No schedule overlap')
                    THEN 1 ELSE 0 END),0)
           FROM mart.attendance_agent_day
           WHERE business_date BETWEEN ? AND ?
             AND assignment_type NOT IN ('Off','Planned absence')""",
        [start, end],
    ).fetchone()
    scheduled, call_now, no_show, late, not_seen, missing = totals
    status, status_text = _source_state(conn, ("fte", "start_end", "lilo", "agent_status"), end)
    if missing:
        status, status_text = "INCOMPLETE", f"{missing:,} scheduled row(s) do not have complete attendance evidence"
    book.dashboard(
        [
            KpiCard("Scheduled working", scheduled, "integer", "Selected agent-day rows"),
            KpiCard("Call/action cases", call_now, "integer", "Selected-period queue"),
            KpiCard("Confirmed no-show", no_show, "integer", "Only after completed shift"),
            KpiCard("Late", late, "integer", "Beyond configured tolerance"),
            KpiCard("Not seen yet", not_seen, "integer", "Provisional current-day state"),
            KpiCard("Missing evidence", missing, "integer", "Unknown, never no-show"),
        ],
        status,
        status_text,
        ["Result", "Agents"],
        [("Scheduled working", scheduled), ("Call/action cases", call_now), ("Confirmed no-show", no_show), ("Late", late), ("Not seen yet", not_seen), ("Missing evidence", missing)],
        [
            "This workbook is the selected-period callout register; it is not an adherence report.",
            "Choose Today for a live queue or Current Week for every callout case and daily trend in that week.",
            "An unfinished shift can be late or not seen, but can never be marked as early leave.",
            "No-show requires completed-shift evidence: either a loaded blank LILO row or sufficiently covered Agent Status that stays Logged Off.",
            "Use Attendance Review only after the operating day is complete.",
        ],
        (("Agents", 1),),
        "column",
    )
    headers, rows = _query(
        conn,
        """SELECT business_date, agent_id, agent_name, team_leader, ops_manager,
                  lob, language, scheduled_start, scheduled_end, shift_state,
                  call_action, attendance_result, actual_first_seen, actual_last_seen,
                  uncoded_late_minutes, no_show_minutes, actual_evidence,
                  source_loaded, is_provisional, evaluation_as_of
           FROM mart.attendance_agent_day
           WHERE business_date BETWEEN ? AND ? AND requires_call=true
           ORDER BY CASE call_action WHEN 'CALL_NO_SHOW' THEN 1 WHEN 'CALL_LATE' THEN 2 ELSE 3 END,
                    business_date, scheduled_start, lob, team_leader, agent_name""",
        [start, end],
    )
    ws = book.table(
        "ACTIONS", "Attendance callout cases",
        "Every actionable case in the selected period, ordered by severity, date and scheduled start.",
        headers, rows,
    )
    if rows:
        col = headers.index("call_action")
        ws.conditional_format(4, col, 3 + len(rows), col, {"type": "text", "criteria": "containing", "value": "CALL_NO_SHOW", "format": book.report.error})
    trend_headers, trend_rows = _query(
        conn,
        """SELECT business_date, count(*) AS scheduled_working,
                  sum(CASE WHEN requires_call THEN 1 ELSE 0 END) AS call_actions,
                  sum(CASE WHEN call_action='CALL_NO_SHOW' THEN 1 ELSE 0 END) AS no_shows,
                  sum(CASE WHEN call_action='CALL_LATE' THEN 1 ELSE 0 END) AS late,
                  sum(CASE WHEN source_loaded=false THEN 1 ELSE 0 END) AS missing_evidence
           FROM mart.attendance_agent_day
           WHERE business_date BETWEEN ? AND ? AND assignment_type NOT IN ('Off','Planned absence')
           GROUP BY business_date ORDER BY business_date""",
        [start, end],
    )
    book.table("TREND", "Attendance action trend", "Daily counts for the requested range; today remains provisional until shifts close.", trend_headers, trend_rows)
    book.definitions([
        ("Call no-show", "Completed shift + blank LILO row, or sufficient all-Logged-Off Agent Status coverage", "Call and validate absence", "Missing source is never a no-show"),
        ("Call late", "First observed evidence after scheduled start plus tolerance", "Contact / record explanation", "Current shift may still be running"),
        ("Not seen now", "Shift started, no observed evidence yet", "Immediate operational check", "Always provisional"),
        ("Early leave", "Last observed evidence before completed scheduled end", "Historical correction only", "Never evaluated before shift end"),
    ])
    book.audit(_audit_rows(conn, config, "attendance", start, end))
    return _finish(book, partial, target)


def build_staffing_coverage_workbook(
    conn: DatabaseConnection,
    config: Config,
    start: date,
    end: date,
    output: Path | None = None,
) -> Path:
    """Build actual staffing control and future required-versus-scheduled plan."""

    book, partial, target = _atomic_book(
        config, "staffing", "STAFFING & CAPACITY PLAN", start, end, output,
    )
    profiles = load_service_profiles(config.home, config.service_profiles)
    active_profiles = tuple(profile for profile in profiles.profiles if profile.active_on(end))
    staffing_profile: dict[str, ServiceProfile] = {}
    for profile in active_profiles:
        for lob in profile.staffing_lobs:
            staffing_profile.setdefault(lob.casefold(), profile)

    forecast_by_lob_hour: dict[tuple[str, str, datetime], tuple[Any, ...]] = {}
    for profile in active_profiles:
        for service_scope, staffing_lob in profile.staffing_pairs():
            forecast_rows = conn.execute(
                """SELECT business_date, hour_start, sum(volume_forecast),
                          sum(fte_forecast), sum(fte_required), avg(sl_required)
                   FROM mart.forecast_hour
                   WHERE business_date BETWEEN ? AND ? AND service_scope=?
                   GROUP BY business_date, hour_start""",
                [start, end, service_scope],
            ).fetchall()
            for business_date, hour_start, volume, fte_forecast, fte_required, sl_required in forecast_rows:
                stamp = hour_start
                if isinstance(stamp, str):
                    stamp = datetime.fromisoformat(stamp)
                key = (
                    staffing_lob.casefold(), str(business_date)[:10],
                    stamp.replace(minute=0, second=0, microsecond=0),
                )
                previous = forecast_by_lob_hour.get(key)
                if previous:
                    volume = float(previous[1] or 0) + float(volume or 0)
                    fte_forecast = float(previous[2] or 0) + float(fte_forecast or 0)
                    fte_required = float(previous[3] or 0) + float(fte_required or 0)
                forecast_by_lob_hour[key] = (
                    profile, volume, fte_forecast, fte_required, sl_required,
                )

    raw_headers, raw_rows = _query(
        conn,
        """SELECT business_date, interval_start, interval_end, lob, language,
                  scheduled_agents, observed_agents, productive_agents,
                  gross_scheduled_fte, planned_time_off_fte, scheduled_fte,
                  elapsed_scheduled_fte, observed_fte, productive_fte,
                  staffing_variance_fte, staffing_gap_fte, staffing_state,
                  evidence_basis, evaluation_as_of
           FROM mart.staffing_interval
           WHERE business_date BETWEEN ? AND ?
           ORDER BY business_date, interval_start, lob, language""",
        [start, end],
    )
    source_indexes = {header: index for index, header in enumerate(raw_headers)}
    known_languages: dict[str, set[str]] = defaultdict(set)
    evaluation_values: list[datetime] = []
    for raw in raw_rows:
        lob = str(raw[source_indexes["lob"]] or "")
        language = str(raw[source_indexes["language"]] or "Unspecified")
        known_languages[lob.casefold()].add(language)
        evaluation = raw[source_indexes["evaluation_as_of"]]
        if isinstance(evaluation, str):
            evaluation = datetime.fromisoformat(evaluation)
        if evaluation is not None:
            evaluation_values.append(evaluation)
    evaluation_default = max(evaluation_values, default=datetime.now())
    plan_headers = [
        "business_date", "iso_week", "interval_start", "interval_end",
        "reporting_lob", "roster_lob", "language", "mode",
        "forecast_volume_hour", "fte_forecast", "fte_required",
        "gross_scheduled_fte", "planned_time_off_fte", "net_scheduled_fte",
        "capacity_variance_fte", "capacity_gap_fte", "observed_fte",
        "productive_fte", "actual_gap_fte", "decision_state",
        "evidence_basis", "evaluation_as_of",
    ]
    plan_rows: list[tuple[Any, ...]] = []
    covered_intervals: set[tuple[str, str, datetime]] = set()
    for raw in raw_rows:
        values = {header: raw[index] for header, index in source_indexes.items()}
        business_date = values["business_date"]
        if isinstance(business_date, str):
            business_date = date.fromisoformat(business_date[:10])
        interval_start = values["interval_start"]
        if isinstance(interval_start, str):
            interval_start = datetime.fromisoformat(interval_start)
        interval_end = values["interval_end"]
        if isinstance(interval_end, str):
            interval_end = datetime.fromisoformat(interval_end)
        evaluation_as_of = values["evaluation_as_of"]
        if isinstance(evaluation_as_of, str):
            evaluation_as_of = datetime.fromisoformat(evaluation_as_of)
        roster_lob = str(values["lob"] or "")
        profile = staffing_profile.get(roster_lob.casefold())
        forecast = forecast_by_lob_hour.get((
            roster_lob.casefold(), business_date.isoformat(),
            interval_start.replace(minute=0, second=0, microsecond=0),
        ))
        if forecast:
            profile = forecast[0]
        _forecast_profile, volume, fte_forecast, fte_required, _sl_required = forecast or (None, None, None, None, None)
        net_scheduled = float(values["scheduled_fte"] or 0)
        capacity_variance = net_scheduled - float(fte_required) if fte_required is not None else None
        capacity_gap = max(0.0, -capacity_variance) if capacity_variance is not None else None
        future = interval_start > evaluation_as_of
        mode = "FUTURE PLAN" if future else "ACTUAL CONTROL"
        actual_gap = values["staffing_gap_fte"]
        if profile is None:
            state = "UNMAPPED LOB"
        elif fte_required is None:
            state = "NO FORECAST"
        elif future:
            state = "FUTURE GAP" if capacity_gap and capacity_gap > 0.001 else "FUTURE OK"
        else:
            state = str(values["staffing_state"] or "DATA MISSING").replace("_", " ")
        plan_rows.append((
            business_date, f"{business_date.isocalendar().year}-W{business_date.isocalendar().week:02d}",
            interval_start, interval_end, profile.label if profile else "Unmapped",
            roster_lob, values["language"], mode, volume, fte_forecast,
            fte_required, values["gross_scheduled_fte"],
            values["planned_time_off_fte"], net_scheduled, capacity_variance,
            capacity_gap, values["observed_fte"], values["productive_fte"],
            actual_gap, state, values["evidence_basis"], evaluation_as_of,
        ))
        covered_intervals.add((roster_lob.casefold(), business_date.isoformat(), interval_start))

    # A demand interval with zero scheduled agents does not exist in the
    # staffing mart. Add it here so an empty roster can never hide a shortage.
    for (lob_key, business_date_text, hour_start), forecast in forecast_by_lob_hour.items():
        profile, volume, fte_forecast, fte_required, _sl_required = forecast
        roster_lob = next(
            (lob for lob in profile.staffing_lobs if lob.casefold() == lob_key), lob_key,
        )
        business_date = date.fromisoformat(business_date_text)
        language = " / ".join(sorted(known_languages.get(lob_key, {"Unspecified"})))
        for offset in range(4):
            interval_start = hour_start + timedelta(minutes=15 * offset)
            key = (lob_key, business_date_text, interval_start)
            if key in covered_intervals:
                continue
            interval_end = interval_start + timedelta(minutes=15)
            required = float(fte_required or 0)
            future = interval_start > evaluation_default
            mode = "FUTURE PLAN" if future else "ACTUAL CONTROL"
            state = "FUTURE GAP" if future and required > 0 else "NO SCHEDULE"
            plan_rows.append((
                business_date,
                f"{business_date.isocalendar().year}-W{business_date.isocalendar().week:02d}",
                interval_start, interval_end, profile.label, roster_lob, language,
                mode, volume, fte_forecast, fte_required, 0.0, 0.0, 0.0,
                -required, required, None, None, required if not future else None,
                state, "Forecast demand with no scheduled roster interval",
                evaluation_default,
            ))

    plan_rows.sort(key=lambda row: (row[0], row[2], str(row[5]), str(row[6])))

    decision_state = plan_headers.index("decision_state")
    capacity_gap_column = plan_headers.index("capacity_gap_fte")
    actual_gap_column = plan_headers.index("actual_gap_fte")
    action_states = {"FUTURE GAP", "NO FORECAST", "UNMAPPED LOB", "NO SCHEDULE", "GAP", "PARTIAL GAP", "DATA MISSING"}
    actions = [row for row in plan_rows if str(row[decision_state]).upper() in action_states]
    future_rows = [row for row in plan_rows if row[plan_headers.index("mode")] == "FUTURE PLAN"]
    required_hours = sum(float(row[plan_headers.index("fte_required")] or 0) * 0.25 for row in future_rows)
    net_hours = sum(float(row[plan_headers.index("net_scheduled_fte")] or 0) * 0.25 for row in future_rows)
    pto_hours = sum(float(row[plan_headers.index("planned_time_off_fte")] or 0) * 0.25 for row in future_rows)
    gap_hours = sum(float(row[capacity_gap_column] or 0) * 0.25 for row in future_rows)
    future_gap_intervals = sum(1 for row in future_rows if row[decision_state] == "FUTURE GAP")
    no_forecast_intervals = sum(1 for row in future_rows if row[decision_state] == "NO FORECAST")
    actual_action_rows = [row for row in plan_rows if row[plan_headers.index("mode")] == "ACTUAL CONTROL"]
    actual_gap_intervals = sum(
        1 for row in actual_action_rows
        if str(row[decision_state]).upper() in {"GAP", "PARTIAL GAP"}
    )
    peak_gap = max(
        [float(row[capacity_gap_column] or 0) for row in future_rows]
        + [float(row[actual_gap_column] or 0) for row in actual_action_rows]
        + [0.0]
    )
    status, status_text = _source_state(conn, ("fte", "start_end", "forecast"), end)
    if no_forecast_intervals:
        status, status_text = "INCOMPLETE", f"{no_forecast_intervals:,} future interval(s) have no mapped forecast"
    book.dashboard(
        [
            KpiCard("Peak gap FTE", peak_gap, "decimal", "Largest 15-minute deficit"),
            KpiCard("Future gap intervals", future_gap_intervals, "integer"),
            KpiCard("Future required hours", required_hours, "decimal", "FTE-hours from Verint"),
            KpiCard("Future net scheduled", net_hours, "decimal", "After PTO and Away"),
            KpiCard("Future gap hours", gap_hours, "decimal"),
            KpiCard("PTO / Away impact", pto_hours, "decimal", "Capacity hours removed"),
            KpiCard("Actual gap intervals", actual_gap_intervals, "integer"),
            KpiCard("Forecast coverage", _ratio(required_hours - gap_hours, required_hours), "percent"),
        ],
        status,
        status_text,
        ["Measure", "Value"],
        [
            ("Required future FTE-hours", required_hours),
            ("Net scheduled future FTE-hours", net_hours),
            ("Future shortage FTE-hours", gap_hours),
            ("PTO / Away FTE-hours", pto_hours),
            ("Future gap intervals", future_gap_intervals),
            ("Actual gap intervals", actual_gap_intervals),
        ],
        [
            "Future plan compares Verint required FTE with net scheduled FTE after approved PTO and effective Away.",
            "Service-to-roster LOB links are editable in Service Profiles; no text guess is made in the report.",
            "Observed FTE is calculated from agent-seconds inside each 15-minute interval.",
            "Missing forecast remains NO FORECAST. It is never converted to zero demand.",
        ],
        (("Value", 1),),
        "column",
    )

    weekly: dict[tuple[str, str, str, str], list[tuple[Any, ...]]] = defaultdict(list)
    for row in plan_rows:
        weekly[(str(row[1]), str(row[4]), str(row[5]), str(row[6]))].append(row)
    weekly_rows = []
    for (iso_week, reporting_lob, roster_lob, language), rows in sorted(weekly.items()):
        required = sum(float(row[10] or 0) * 0.25 for row in rows)
        gross = sum(float(row[11] or 0) * 0.25 for row in rows)
        time_off = sum(float(row[12] or 0) * 0.25 for row in rows)
        net = sum(float(row[13] or 0) * 0.25 for row in rows)
        gap = sum(float(row[15] or 0) * 0.25 for row in rows)
        weekly_rows.append((
            iso_week, min(row[0] for row in rows), max(row[0] for row in rows),
            reporting_lob, roster_lob, language, required, gross, time_off, net,
            net - required if required else None, gap,
            _ratio(net, required),
            sum(1 for row in rows if row[19] == "FUTURE GAP"),
            sum(1 for row in rows if row[19] == "NO FORECAST"),
        ))
    book.table(
        "WEEKLY_PLAN", "Weekly capacity plan",
        "Required, gross scheduled, PTO/Away and net scheduled FTE-hours by ISO week, LOB and language.",
        [
            "iso_week", "start_date", "end_date", "reporting_lob", "roster_lob",
            "language", "required_fte_hours", "gross_scheduled_fte_hours",
            "planned_time_off_fte_hours", "net_scheduled_fte_hours",
            "capacity_variance_fte_hours", "capacity_gap_fte_hours",
            "forecast_coverage", "gap_intervals", "no_forecast_intervals",
        ],
        weekly_rows,
    )
    intraday = book.table(
        "INTRADAY", "15-minute staffing control and plan",
        "All selected dates. FUTURE PLAN uses forecast demand; ACTUAL CONTROL uses observed attendance evidence.",
        plan_headers, plan_rows,
    )
    if plan_rows:
        intraday.conditional_format(
            4, decision_state, 3 + len(plan_rows), decision_state,
            {"type": "text", "criteria": "containing", "value": "GAP", "format": book.report.error},
        )
    book.table(
        "ACTIONS", "Staffing exceptions",
        "Future shortages, missing forecasts, unmapped LOBs and actual gaps requiring action.",
        plan_headers, actions,
    )
    book.definitions([
        ("Required FTE", "Verint required FTE repeated across its four 15-minute intervals", "Demand requirement", "Forecast only; missing stays blank"),
        ("Gross scheduled FTE", "Scheduled agent-seconds / 900 before time off", "Roster capacity", "FTE roster LOB/language"),
        ("Net scheduled FTE", "Gross scheduled FTE - approved PTO/effective Away FTE", "Usable planned capacity", "Planned Away affects future only"),
        ("Future capacity gap", "MAX(0, required FTE - net scheduled FTE)", "Hiring, OT or redeployment action", "15-minute interval"),
        ("Observed FTE", "Observed agent-seconds / 900", "Actual presence", "LILO + Agent Status evidence"),
        ("Productive FTE", "Productive-status seconds / 900", "Available handling capacity", "Not adherence"),
        ("Actual staffing gap", "MAX(0, elapsed net scheduled FTE - observed FTE)", "Same-day staffing deficit", "Blank for future/missing evidence"),
    ])
    book.audit(_audit_rows(conn, config, "staffing", start, end, [
        ("Service profile mapping", profiles.version, profiles.sha256),
        ("Planning grain", "15 minutes", "Weekly summary uses FTE-hours"),
    ]))
    return _finish(book, partial, target)


def _exclusive_final_components(
    conn: DatabaseConnection,
    rulebook,
    start: date,
    end: date,
    flag: str,
) -> tuple[list[str], list[tuple[Any, ...]]]:
    """Allocate overlapping final Activities once for a component breakdown."""

    columns = {
        "absence": "e.counts_as_absence",
        "shrinkage": "e.counts_as_shrinkage",
    }
    if flag not in columns:
        raise ValueError(f"Unsupported final component flag {flag!r}")
    headers, raw_rows = _query(
        conn,
        f"""SELECT e.agent_day_key, e.business_date, e.agent_id, e.agent_name,
                   e.team_leader, e.lob, e.activity, e.category,
                   e.event_start, e.event_end, d.planned_net_minutes
            FROM mart.verint_final_absence_event e
            JOIN mart.verint_final_absence_agent_day d
              ON d.agent_day_key=e.agent_day_key
            WHERE e.business_date BETWEEN ? AND ? AND {columns[flag]}=true
            ORDER BY e.business_date, e.agent_id, e.event_start, e.event_end""",
        [start, end],
    )
    events = [dict(zip(headers, row)) for row in raw_rows]
    precedence = {rule.name: index for index, rule in enumerate(rulebook.activity_rules)}
    by_day: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for event in events:
        by_day[str(event["agent_day_key"])].append(event)
    totals: dict[tuple[Any, ...], int] = defaultdict(int)
    days: dict[tuple[Any, ...], set[Any]] = defaultdict(set)
    for day_events in by_day.values():
        remaining = int(day_events[0].get("planned_net_minutes") or 0)
        boundaries = sorted({
            stamp for event in day_events
            for stamp in (event["event_start"], event["event_end"])
            if stamp is not None
        })
        for left, right in zip(boundaries, boundaries[1:]):
            if right <= left:
                continue
            active = [
                event for event in day_events
                if event["event_start"] <= left and event["event_end"] >= right
            ]
            if not active:
                continue

            def rank(event: dict[str, Any]) -> tuple[int, str, str]:
                rule = rulebook.classify_activity(event.get("activity"))
                return (
                    precedence.get(rule.name if rule else "", len(precedence) + 1),
                    str(event.get("category") or "Other"),
                    str(event.get("activity") or ""),
                )

            chosen = min(active, key=rank)
            key = (
                chosen.get("category") or "OTHER",
                chosen.get("lob"), chosen.get("team_leader"),
                chosen.get("agent_id"), chosen.get("agent_name"),
            )
            allocated = min(remaining, int((right - left).total_seconds() // 60))
            if allocated <= 0:
                break
            totals[key] += allocated
            remaining -= allocated
            days[key].add(chosen.get("business_date"))
    output = [
        (*key, len(days[key]), minutes, minutes / 60.0)
        for key, minutes in totals.items()
        if minutes > 0
    ]
    output.sort(key=lambda row: (str(row[0]), str(row[1]), str(row[2]), str(row[4])))
    return (
        [
            "component", "lob", "team_leader", "agent_id", "agent_name",
            "agent_days", "minutes", "hours",
        ],
        output,
    )


def _legacy_build_final_absence_product_workbook(
    conn: DatabaseConnection,
    config: Config,
    start: date,
    end: date,
    output: Path | None = None,
) -> Path:
    """Build the final corrected Verint absence and shrinkage product."""

    book, partial, target = _atomic_book(config, "absence", "FINAL ABSENCE & SHRINKAGE", start, end, output)
    totals = conn.execute(
        """SELECT coalesce(sum(planned_net_minutes),0), coalesce(sum(final_absence_minutes),0),
                  coalesce(sum(final_vacation_minutes),0), coalesce(sum(final_unpaid_minutes),0),
                  coalesce(sum(final_shrinkage_minutes),0), coalesce(sum(final_unmapped_minutes),0),
                  coalesce(sum(CASE WHEN final_absence_day THEN 1 ELSE 0 END),0)
           FROM mart.verint_final_absence_agent_day
           WHERE business_date BETWEEN ? AND ?
             AND final_ledger_status IN ('CLEAR','ABSENCE_RECORDED')""",
        [start, end],
    ).fetchone()
    planned, absence, vacation, unpaid, shrinkage, unmapped, absence_days = totals
    unmapped = conn.execute(
        """SELECT coalesce(sum(final_unmapped_minutes),0)
           FROM mart.verint_final_absence_agent_day
           WHERE business_date BETWEEN ? AND ?""",
        [start, end],
    ).fetchone()[0]
    ledger_exceptions = conn.execute(
        """SELECT count(*) FROM mart.verint_final_absence_agent_day
           WHERE business_date BETWEEN ? AND ?
             AND final_ledger_status NOT IN ('CLEAR','ABSENCE_RECORDED')""",
        [start, end],
    ).fetchone()[0]
    uncoded_empty = conn.execute(
        """SELECT count(*) FROM mart.verint_final_absence_agent_day
           WHERE business_date BETWEEN ? AND ? AND final_ledger_status='UNCODED_EMPTY_SHIFT'""",
        [start, end],
    ).fetchone()[0]
    status, status_text = _source_state(conn, ("fte", "start_end", "activities"), end, final=True)
    if unmapped:
        status, status_text = "INCOMPLETE", f"{unmapped / 60:,.2f} hour(s) of Verint Activities are unmapped"
    if uncoded_empty:
        status, status_text = "INCOMPLETE", f"{uncoded_empty:,} scheduled shift(s) have no final code and no reliable work evidence"
    elif ledger_exceptions:
        status, status_text = "INCOMPLETE", f"{ledger_exceptions:,} final-ledger row(s) still require review"
    period_rows = []
    for period in _comparison_periods(start, end):
        row = conn.execute(
            """SELECT coalesce(sum(planned_net_minutes),0), coalesce(sum(final_absence_minutes),0),
                      coalesce(sum(final_vacation_minutes),0), coalesce(sum(final_shrinkage_minutes),0)
               FROM mart.verint_final_absence_agent_day
               WHERE business_date BETWEEN ? AND ?
                 AND final_ledger_status IN ('CLEAR','ABSENCE_RECORDED')""",
            [period.start, period.end],
        ).fetchone()
        period_rows.append((period.label, period.start, period.end, row[0] / 60, row[1] / 60, _ratio(row[1], row[0]), row[2] / 60, _ratio(row[3], row[0])))
    book.dashboard(
        [
            KpiCard("Finalized planned hours", planned / 60, "decimal"),
            KpiCard("Final absence hours", absence / 60, "decimal"),
            KpiCard("Final absence rate", _ratio(absence, planned), "percent"),
            KpiCard("Absence agent-days", absence_days, "integer"),
            KpiCard("Vacation hours", vacation / 60, "decimal"),
            KpiCard("Unpaid hours", unpaid / 60, "decimal"),
            KpiCard("Shrinkage rate", _ratio(shrinkage, planned), "percent"),
            KpiCard("Ledger exceptions", ledger_exceptions, "integer", f"{uncoded_empty:,} uncoded empty shift(s)"),
        ],
        status,
        status_text,
        ["Period", "Start", "End", "Planned Hours", "Absence Hours", "Absence Rate %", "Vacation Hours", "Shrinkage Rate %"],
        period_rows,
        [
            "This is the final corrected ledger: Verint Activities only, clipped to StartEndTimes schedule boundaries.",
            "LILO and Agent Status detect operational gaps but do not create final payroll categories.",
            "A scheduled working shift with neither a final Verint code nor reliable operational evidence is UNCODED_EMPTY_SHIFT, never a silent zero.",
            "Observed Agent Status/LILO gaps without complete Verint coverage remain ledger exceptions until corrected.",
            "A Verint code with no matching observed gap remains an exception too; the final ledger never assumes it is valid.",
            "Headline rates include finalized CLEAR/ABSENCE_RECORDED rows only; exception and current provisional rows cannot dilute the result.",
            "Overlapping Activities are unioned before totals; never sum event evidence directly.",
        ],
        (("Absence Rate", 5), ("Shrinkage Rate", 7)),
    )
    trend_headers, trend_rows = _query(
        conn,
        """SELECT business_date, sum(planned_net_minutes)/60.0 AS planned_hours,
                  sum(final_absence_minutes)/60.0 AS absence_hours,
                  CASE WHEN sum(planned_net_minutes)>0 THEN sum(final_absence_minutes)*1.0/sum(planned_net_minutes) END AS absence_rate,
                  sum(final_vacation_minutes)/60.0 AS vacation_hours,
                  sum(final_unpaid_minutes)/60.0 AS unpaid_hours,
                  sum(final_shrinkage_minutes)/60.0 AS shrinkage_hours,
                  CASE WHEN sum(planned_net_minutes)>0 THEN sum(final_shrinkage_minutes)*1.0/sum(planned_net_minutes) END AS shrinkage_rate
           FROM mart.verint_final_absence_agent_day WHERE business_date BETWEEN ? AND ?
             AND final_ledger_status IN ('CLEAR','ABSENCE_RECORDED')
           GROUP BY business_date ORDER BY business_date""",
        [start, end],
    )
    book.table("TREND", "Final absence daily trend", "Daily overlap-safe final counters.", trend_headers, trend_rows)
    detail_headers, detail_rows = _query(
        conn,
        """SELECT business_date, agent_id, agent_name, team_leader, ops_manager,
                  lob, language, planned_net_minutes/60.0 AS planned_net_hours,
                  final_absence_minutes/60.0 AS final_absence_hours,
                  final_vacation_minutes/60.0 AS final_vacation_hours,
                  final_unpaid_minutes/60.0 AS final_unpaid_hours,
                  final_shrinkage_minutes/60.0 AS final_shrinkage_hours,
                  final_unmapped_minutes/60.0 AS final_unmapped_hours,
                  final_absence_rate, final_absence_day, final_ledger_status
           FROM mart.verint_final_absence_agent_day WHERE business_date BETWEEN ? AND ?
           ORDER BY business_date, lob, team_leader, agent_name""",
        [start, end],
    )
    book.table("AGENT_DETAIL", "Final absence agent ledger", "Authoritative agent/day grain for payroll and Operations.", detail_headers, detail_rows)
    exceptions = [
        row for row in detail_rows
        if row[detail_headers.index("final_unmapped_hours")]
        or row[detail_headers.index("final_ledger_status")] not in {"CLEAR", "ABSENCE_RECORDED"}
    ]
    book.table("EXCEPTIONS", "Final-ledger exceptions", "Unmapped, unsupported, empty, uncorrected and partially corrected rows must be resolved before payroll use.", detail_headers, exceptions)
    book.definitions([
        ("Final absence rate", "Unioned classified absence minutes / planned net minutes", "Payroll absence", "Activities-only final ledger"),
        ("Final shrinkage rate", "Unioned configured shrinkage minutes / planned net minutes", "Capacity loss", "Includes configured nonproductive categories"),
        ("Planned net", "Scheduled span capped by configured standard day", "Common denominator", "StartEndTimes only"),
        ("Unmapped", "Verint Activity without an approved classification", "Rulebook action", "Blocks final status"),
        ("Uncoded empty shift", "Working shift with no final code and no reliable Agent Status/LILO evidence", "Correction completeness check", "Blocks final status; never treated as zero absence"),
        ("Uncorrected observed gap", "Agent Status/LILO proves a gap but Verint has no final code", "Correction completeness check", "Blocks final status"),
        ("Partial correction review", "Verint has a code but some observed gap minutes remain uncovered", "Correction completeness check", "Blocks final status"),
        ("Verint without observed gap", "A final Verint code has no matching Agent Status/LILO gap", "Correction validity check", "Blocks final status"),
    ])
    book.audit(_audit_rows(conn, config, "absence", start, end))
    return _finish(book, partial, target)


def _add_absence_team_view(
    book: DecisionWorkbook,
    start: date,
    end: date,
    data_start: date,
    latest: date,
) -> None:
    """Add one selector-driven view for Operations and Payroll reviewers."""

    wb = book.report.workbook
    wb.set_calc_mode("auto")
    ws = wb.add_worksheet("TEAM_VIEW")
    ws.hide_gridlines(2)
    ws.set_tab_color(COLORS["gold"])
    ws.set_zoom(85)
    ws.freeze_panes(16, 0)
    ws.merge_range("A1:AA1", "ABSENTEEISM & SHRINKAGE  /  TEAM REVIEW", book.report.title)
    ws.merge_range(
        "A2:AA2",
        f"Latest final-ledger date {latest:%Y-%m-%d}  |  select a period and team below  |  prepared by Anass ASSRI",
        book.report.subtitle,
    )
    selector_label = wb.add_format({
        "font_name": "Aptos", "font_size": 8, "bold": True,
        "font_color": COLORS["muted"], "bg_color": COLORS["canvas"],
        "align": "left", "valign": "vcenter", "indent": 1,
    })
    selector = wb.add_format({
        "font_name": "Aptos Display", "font_size": 11, "bold": True,
        "font_color": COLORS["dark"], "bg_color": COLORS["white"],
        "border": 1, "border_color": COLORS["teal"], "align": "left",
        "valign": "vcenter", "indent": 1, "num_format": "yyyy-mm-dd",
    })
    for label, label_range, value_range, value in (
        ("PERIOD VIEW", "A4:C4", "A5:C6", "Current MTD"),
        ("CUSTOM FROM", "E4:F4", "E5:F6", start),
        ("CUSTOM TO", "H4:I4", "H5:I6", end),
        ("LOB", "K4:L4", "K5:L6", "All"),
        ("TEAM LEADER", "N4:O4", "N5:O6", "All"),
        ("AGENT", "Q4:R4", "Q5:R6", "All"),
    ):
        ws.merge_range(label_range, label, selector_label)
        ws.merge_range(value_range, value, selector)
    ws.data_validation("A5", {
        "validate": "list",
        "source": [
            "Latest day", "Current week", "Previous week", "Current MTD",
            "Previous-month same days", "Previous full month", "Custom period",
        ],
    })
    ws.data_validation("E5", {
        "validate": "date", "criteria": "between", "minimum": data_start,
        "maximum": latest,
    })
    ws.data_validation("H5", {
        "validate": "date", "criteria": "between", "minimum": data_start,
        "maximum": latest,
    })
    ws.data_validation("K5", {"validate": "list", "source": "=ABS_LOB_LIST"})
    ws.data_validation("N5", {"validate": "list", "source": "=ABS_TL_LIST"})
    ws.data_validation("Q5", {"validate": "list", "source": "=ABS_AGENT_LIST"})
    wb.define_name("ABS_Latest", "=MAX(tblAbsenceData[Date])")
    wb.define_name(
        "ABS_From",
        '=IF(TEAM_VIEW!$A$5="Latest day",ABS_Latest,'
        'IF(TEAM_VIEW!$A$5="Current week",ABS_Latest-WEEKDAY(ABS_Latest,2)+1,'
        'IF(TEAM_VIEW!$A$5="Previous week",ABS_Latest-WEEKDAY(ABS_Latest,2)-6,'
        'IF(TEAM_VIEW!$A$5="Current MTD",EOMONTH(ABS_Latest,-1)+1,'
        'IF(TEAM_VIEW!$A$5="Previous-month same days",EOMONTH(ABS_Latest,-2)+1,'
        'IF(TEAM_VIEW!$A$5="Previous full month",EOMONTH(ABS_Latest,-2)+1,TEAM_VIEW!$E$5))))))',
    )
    wb.define_name(
        "ABS_To",
        '=IF(TEAM_VIEW!$A$5="Latest day",ABS_Latest,'
        'IF(TEAM_VIEW!$A$5="Current week",ABS_Latest,'
        'IF(TEAM_VIEW!$A$5="Previous week",ABS_Latest-WEEKDAY(ABS_Latest,2),'
        'IF(TEAM_VIEW!$A$5="Current MTD",ABS_Latest,'
        'IF(TEAM_VIEW!$A$5="Previous-month same days",EDATE(ABS_Latest,-1),'
        'IF(TEAM_VIEW!$A$5="Previous full month",EOMONTH(ABS_Latest,-1),TEAM_VIEW!$H$5))))))',
    )
    scope = (
        '(tblAbsenceData[Date]>=ABS_From)*(tblAbsenceData[Date]<=ABS_To)*'
        'IF(TEAM_VIEW!$K$5="All",1,--(tblAbsenceData[LOB]=TEAM_VIEW!$K$5))*'
        'IF(TEAM_VIEW!$N$5="All",1,--(tblAbsenceData[Team Leader]=TEAM_VIEW!$N$5))*'
        'IF(TEAM_VIEW!$Q$5="All",1,--(tblAbsenceData[Agent Selector]=TEAM_VIEW!$Q$5))'
    )
    final = '((tblAbsenceData[Final Ledger Status]="CLEAR")+(tblAbsenceData[Final Ledger Status]="ABSENCE_RECORDED"))'
    planned = f"SUMPRODUCT({scope}*{final}*N(tblAbsenceData[Planned Net Hours]))"
    absence = f"SUMPRODUCT({scope}*{final}*N(tblAbsenceData[Absence Hours]))"
    shrinkage = f"SUMPRODUCT({scope}*{final}*N(tblAbsenceData[Shrinkage Hours]))"
    review = (
        f'SUMPRODUCT({scope}*--(tblAbsenceData[Final Ledger Status]<>"CLEAR")*'
        '--(tblAbsenceData[Final Ledger Status]<>"ABSENCE_RECORDED"))'
    )
    ws.merge_range("A8:AA8", "SELECTED TEAM POSITION", book.report.section)
    cards = (
        ("ABSENCE RATE", f'=IFERROR({absence}/{planned},"")', book.card_percent),
        ("SHRINKAGE RATE", f'=IFERROR({shrinkage}/{planned},"")', book.card_percent),
        ("ABSENCE HOURS", f"={absence}", book.card_decimal),
        ("REVIEW CASES", f"={review}", book.card_integer),
    )
    for index, (label, formula, fmt) in enumerate(cards):
        column = index * 4
        ws.merge_range(9, column, 9, column + 2, label, book.report.kpi_label)
        ws.merge_range(10, column, 11, column + 2, "", fmt)
        ws.write_formula(10, column, formula, fmt, "")
    ws.merge_range("A14:L14", "AGENT RESULTS", book.report.section)
    agent_headers = [
        "LOB", "Team Leader", "Agent Selector", "Agent ID", "Planned Hours",
        "Absence Hours", "Absence Rate", "Shrinkage Hours", "Shrinkage Rate",
        "Vacation Hours", "Unpaid Hours", "Review Cases",
    ]
    for column, header in enumerate(agent_headers):
        ws.write(15, column, header, book.report.header)
    agent_formula = (
        '=LET(d,tblAbsenceData,'
        'm,(d[Date]>=ABS_From)*(d[Date]<=ABS_To)*'
        'IF(TEAM_VIEW!$K$5="All",1,--(d[LOB]=TEAM_VIEW!$K$5))*'
        'IF(TEAM_VIEW!$N$5="All",1,--(d[Team Leader]=TEAM_VIEW!$N$5))*'
        'IF(TEAM_VIEW!$Q$5="All",1,--(d[Agent Selector]=TEAM_VIEW!$Q$5)),'
        'f,((d[Final Ledger Status]="CLEAR")+(d[Final Ledger Status]="ABSENCE_RECORDED")),'
        'a,SORT(UNIQUE(FILTER(d[Agent Selector],m,""))),'
        'ph,MAP(a,LAMBDA(x,SUMPRODUCT(m*f*(d[Agent Selector]=x)*N(d[Planned Net Hours])))),'
        'ah,MAP(a,LAMBDA(x,SUMPRODUCT(m*f*(d[Agent Selector]=x)*N(d[Absence Hours])))),'
        'sh,MAP(a,LAMBDA(x,SUMPRODUCT(m*f*(d[Agent Selector]=x)*N(d[Shrinkage Hours])))),'
        'vh,MAP(a,LAMBDA(x,SUMPRODUCT(m*f*(d[Agent Selector]=x)*N(d[Vacation Hours])))),'
        'uh,MAP(a,LAMBDA(x,SUMPRODUCT(m*f*(d[Agent Selector]=x)*N(d[Unpaid Hours])))),'
        'rv,MAP(a,LAMBDA(x,SUMPRODUCT(m*(d[Agent Selector]=x)*--(d[Final Ledger Status]<>"CLEAR")*--(d[Final Ledger Status]<>"ABSENCE_RECORDED")))),'
        'IFERROR(HSTACK('
        'XLOOKUP(a,d[Agent Selector],d[LOB],""),'
        'XLOOKUP(a,d[Agent Selector],d[Team Leader],""),a,'
        'XLOOKUP(a,d[Agent Selector],d[Agent ID],""),ph,ah,IFERROR(ah/ph,""),'
        'sh,IFERROR(sh/ph,""),vh,uh,rv),"No matching agent data"))'
    )
    ws.write_dynamic_array_formula("A17", agent_formula, book.report.body, "Open in desktop Excel")
    ws.merge_range("N14:AA14", "CASES TO REVIEW", book.report.section)
    queue_headers = [
        "Case ID", "Date", "Agent ID", "Agent", "Team Leader", "LOB",
        "Result Status", "Absence Hours", "Shrinkage Hours", "Unmapped Hours",
        "Action Status",
    ]
    for column, header in enumerate(queue_headers, 13):
        ws.write(15, column, header, book.report.header)
    queue_formula = (
        '=LET(q,tblActionQueue,'
        'm,(q[Date]>=ABS_From)*(q[Date]<=ABS_To)*'
        'IF(TEAM_VIEW!$K$5="All",1,--(q[LOB]=TEAM_VIEW!$K$5))*'
        'IF(TEAM_VIEW!$N$5="All",1,--(q[Team Leader]=TEAM_VIEW!$N$5))*'
        'IF(TEAM_VIEW!$Q$5="All",1,--(q[Agent Selector]=TEAM_VIEW!$Q$5)),'
        'IFERROR(FILTER(CHOOSECOLS(q,1,2,3,4,6,8,10,12,13,14,15),m),'
        '"No cases in this selection"))'
    )
    ws.write_dynamic_array_formula("N17", queue_formula, book.report.body, "Open in desktop Excel")
    ws.write_url(
        "T10", "internal:'ACTIONS'!A1", book.report.editable,
        string="OPEN PERMANENT ACTION LOG",
    )
    ws.set_column("A:A", 18)
    ws.set_column("B:B", 22)
    ws.set_column("C:C", 30)
    ws.set_column("D:L", 15)
    ws.set_column("M:M", 3)
    ws.set_column("N:AA", 18)
    ws.set_landscape()
    ws.fit_to_pages(1, 0)


def _add_absence_component_view(book: DecisionWorkbook) -> None:
    """Show filtered absence and shrinkage components from exact intervals."""

    wb = book.report.workbook
    ws = wb.add_worksheet("COMPONENT_VIEW")
    ws.hide_gridlines(2)
    ws.set_tab_color(COLORS["gold"])
    ws.set_zoom(90)
    ws.merge_range("A1:N1", "ABSENCE & SHRINKAGE  /  COMPONENT VIEW", book.report.title)
    ws.merge_range(
        "A2:N2", "This view follows TEAM_VIEW period, LOB, Team Leader and Agent selectors.",
        book.report.subtitle,
    )
    ws.merge_range("A4:F4", "ABSENCE COMPONENTS", book.report.section)
    ws.merge_range("H4:N4", "SHRINKAGE COMPONENTS", book.report.section)
    headers = ("Component", "Hours", "Intervals")
    for column, header in enumerate(headers):
        ws.write(5, column, header, book.report.header)
        ws.write(5, column + 7, header, book.report.header)
    base_scope = (
        '(t[Date]>=ABS_From)*(t[Date]<=ABS_To)*'
        'IF(TEAM_VIEW!$K$5="All",1,--(t[LOB]=TEAM_VIEW!$K$5))*'
        'IF(TEAM_VIEW!$N$5="All",1,--(t[Team Leader]=TEAM_VIEW!$N$5))*'
        'IF(TEAM_VIEW!$Q$5="All",1,--(t[Agent Selector]=TEAM_VIEW!$Q$5))'
    )
    for cell, flag, empty_text in (
        ("A7", "Counts As Absence", "No absence components in this selection"),
        ("H7", "Counts As Shrinkage", "No shrinkage components in this selection"),
    ):
        formula = (
            '=LET(t,tblActivityDetail,'
            f'm,{base_scope}*--(t[{flag}]=TRUE),'
            'c,SORT(UNIQUE(FILTER(t[Category],m,""))),'
            'h,MAP(c,LAMBDA(x,SUMPRODUCT(m*(t[Category]=x)*N(t[Hours])))),'
            'n,MAP(c,LAMBDA(x,SUMPRODUCT(m*(t[Category]=x)))),'
            f'IFERROR(HSTACK(c,h,n),"{empty_text}"))'
        )
        ws.write_dynamic_array_formula(cell, formula, book.report.body, "Open in desktop Excel")
    ws.merge_range(
        "A22:N22", "Exact start/end evidence remains on ACTIVITY_DETAIL. Component hours are exclusive inside each KPI view; never add Absence Rate and Shrinkage Rate together.",
        book.report.note,
    )
    ws.write_url(
        "A24", "internal:'ACTIVITY_DETAIL'!A1", book.report.editable,
        string="OPEN EXACT VERINT ACTIVITY DETAIL",
    )
    ws.set_column("A:A", 28)
    ws.set_column("B:C", 15)
    ws.set_column("D:G", 4)
    ws.set_column("H:H", 28)
    ws.set_column("I:J", 15)


def _add_absence_lookups(book: DecisionWorkbook) -> None:
    """Create cascading LOB, Team Leader and Agent lists for Absenteeism."""

    wb = book.report.workbook
    ws = wb.add_worksheet("_LOOKUPS")
    ws.write("A1", "All")
    ws.write_dynamic_array_formula(
        "A2", '=SORT(UNIQUE(FILTER(tblAbsenceData[LOB],tblAbsenceData[LOB]<>"","")))',
    )
    ws.write("B1", "All")
    ws.write_dynamic_array_formula(
        "B2",
        '=SORT(UNIQUE(FILTER(tblAbsenceData[Team Leader],'
        '(tblAbsenceData[Team Leader]<>"")*IF(TEAM_VIEW!$K$5="All",1,'
        'tblAbsenceData[LOB]=TEAM_VIEW!$K$5),"")))',
    )
    ws.write("C1", "All")
    ws.write_dynamic_array_formula(
        "C2",
        '=SORT(UNIQUE(FILTER(tblAbsenceData[Agent Selector],'
        '(tblAbsenceData[Agent Selector]<>"")*IF(TEAM_VIEW!$K$5="All",1,'
        'tblAbsenceData[LOB]=TEAM_VIEW!$K$5)*IF(TEAM_VIEW!$N$5="All",1,'
        'tblAbsenceData[Team Leader]=TEAM_VIEW!$N$5),"")))',
    )
    wb.define_name("ABS_LOB_LIST", "=_LOOKUPS!$A$1:INDEX(_LOOKUPS!$A:$A,COUNTA(_LOOKUPS!$A:$A))")
    wb.define_name("ABS_TL_LIST", "=_LOOKUPS!$B$1:INDEX(_LOOKUPS!$B:$B,COUNTA(_LOOKUPS!$B:$B))")
    wb.define_name("ABS_AGENT_LIST", "=_LOOKUPS!$C$1:INDEX(_LOOKUPS!$C:$C,COUNTA(_LOOKUPS!$C:$C))")
    ws.hide()


def build_final_absence_product_workbook(
    conn: DatabaseConnection,
    config: Config,
    start: date,
    end: date,
    output: Path | None = None,
) -> Path:
    """Build the final Absenteeism and Shrinkage collaboration report."""

    from .shared_feeds import publish_absence_feeds

    publish_absence_feeds(conn, config, start, end)
    rulebook = load_rulebook(config.home, config.business_rules)
    book, partial, target = _atomic_book(
        config, "absence", "ABSENTEEISM & SHRINKAGE", start, end, output,
    )
    available_start, available_end = conn.execute(
        "SELECT min(business_date), max(business_date) FROM mart.verint_final_absence_agent_day"
    ).fetchone()
    data_start = (
        available_start if isinstance(available_start, date)
        else date.fromisoformat(str(available_start)[:10]) if available_start else start
    )
    latest = (
        available_end if isinstance(available_end, date)
        else date.fromisoformat(str(available_end)[:10]) if available_end else end
    )
    totals = conn.execute(
        """SELECT coalesce(sum(planned_net_minutes),0),
                  coalesce(sum(final_absence_minutes),0),
                  coalesce(sum(final_vacation_minutes),0),
                  coalesce(sum(final_unpaid_minutes),0),
                  coalesce(sum(final_shrinkage_minutes),0),
                  coalesce(sum(final_unmapped_minutes),0),
                  coalesce(sum(CASE WHEN final_absence_day THEN 1 ELSE 0 END),0),
                  count(*),
                  coalesce(sum(CASE WHEN final_ledger_status IN ('CLEAR','ABSENCE_RECORDED')
                                    THEN planned_net_minutes ELSE 0 END),0),
                  coalesce(sum(CASE WHEN final_ledger_status NOT IN ('CLEAR','ABSENCE_RECORDED')
                                    THEN 1 ELSE 0 END),0)
           FROM mart.verint_final_absence_agent_day
           WHERE business_date BETWEEN ? AND ?""",
        [start, end],
    ).fetchone()
    (
        all_planned, all_absence, all_vacation, all_unpaid, all_shrinkage,
        unmapped, absence_days, agent_days, finalized_planned, exceptions,
    ) = totals
    finalized = conn.execute(
        """SELECT coalesce(sum(planned_net_minutes),0),
                  coalesce(sum(final_absence_minutes),0),
                  coalesce(sum(final_vacation_minutes),0),
                  coalesce(sum(final_unpaid_minutes),0),
                  coalesce(sum(final_shrinkage_minutes),0)
           FROM mart.verint_final_absence_agent_day
           WHERE business_date BETWEEN ? AND ?
             AND final_ledger_status IN ('CLEAR','ABSENCE_RECORDED')""",
        [start, end],
    ).fetchone()
    planned, absence, vacation, unpaid, shrinkage = finalized
    uncoded_empty = conn.execute(
        """SELECT count(*) FROM mart.verint_final_absence_agent_day
           WHERE business_date BETWEEN ? AND ?
             AND final_ledger_status='UNCODED_EMPTY_SHIFT'""",
        [start, end],
    ).fetchone()[0]
    status, status_text = _source_state(
        conn, ("fte", "start_end", "activities"), end, final=True,
    )
    if unmapped:
        status, status_text = "INCOMPLETE", f"{unmapped / 60:,.2f} hour(s) still need an activity mapping"
    elif uncoded_empty:
        status, status_text = "INCOMPLETE", f"{uncoded_empty:,} scheduled shift(s) have no final code or reliable login evidence"
    elif exceptions:
        status, status_text = "INCOMPLETE", f"{exceptions:,} case(s) still require review"

    period_rows = []
    for period in _comparison_periods(start, end):
        values = conn.execute(
            """SELECT coalesce(sum(planned_net_minutes),0),
                      coalesce(sum(final_absence_minutes),0),
                      coalesce(sum(final_vacation_minutes),0),
                      coalesce(sum(final_shrinkage_minutes),0)
               FROM mart.verint_final_absence_agent_day
               WHERE business_date BETWEEN ? AND ?
                 AND final_ledger_status IN ('CLEAR','ABSENCE_RECORDED')""",
            [period.start, period.end],
        ).fetchone()
        period_rows.append((
            period.label, period.start, period.end, values[0] / 60,
            values[1] / 60, _ratio(values[1], values[0]), values[2] / 60,
            values[3] / 60, _ratio(values[3], values[0]),
        ))
    coverage = _ratio(finalized_planned, all_planned)
    book.dashboard(
        [
            KpiCard("Finalized planned hours", planned / 60, "decimal"),
            KpiCard("Absence hours", absence / 60, "decimal"),
            KpiCard("Absence rate", _ratio(absence, planned), "percent"),
            KpiCard("Absence agent-days", absence_days, "integer"),
            KpiCard("Shrinkage hours", shrinkage / 60, "decimal"),
            KpiCard("Shrinkage rate", _ratio(shrinkage, planned), "percent"),
            KpiCard("Vacation / unpaid", (vacation + unpaid) / 60, "decimal"),
            KpiCard("Finalized coverage", coverage, "percent", f"{exceptions:,} review case(s)"),
        ],
        status,
        status_text,
        [
            "Period", "Start", "End", "Planned Hours", "Absence Hours",
            "Absence Rate %", "Vacation Hours", "Shrinkage Hours",
            "Shrinkage Rate %",
        ],
        period_rows,
        [
            "Final results use Verint Activities inside the preferred StartEndTimes boundary or its reviewed Shift Assignment fallback.",
            "Agent Status and LILO are used to find missing or unsupported corrections; they do not assign payroll categories.",
            "Absence and shrinkage are parallel views. Do not add their percentages together.",
            "Open ACTIONS for unresolved cases and use ACTIVITY_DETAIL when an exact interval needs investigation.",
        ],
        (("Absence rate", 5), ("Shrinkage rate", 8)),
    )
    _add_absence_team_view(book, start, end, data_start, latest)

    team_headers, team_rows = _query(
        conn,
        """SELECT lob, team_leader, count(DISTINCT agent_id) AS agents,
                  count(*) AS agent_days,
                  sum(planned_net_minutes)/60.0 AS planned_hours,
                  sum(final_absence_minutes)/60.0 AS absence_hours,
                  CASE WHEN sum(planned_net_minutes)>0
                       THEN sum(final_absence_minutes)*1.0/sum(planned_net_minutes) END AS absence_rate,
                  sum(final_shrinkage_minutes)/60.0 AS shrinkage_hours,
                  CASE WHEN sum(planned_net_minutes)>0
                       THEN sum(final_shrinkage_minutes)*1.0/sum(planned_net_minutes) END AS shrinkage_rate,
                  sum(final_vacation_minutes)/60.0 AS vacation_hours,
                  sum(final_unpaid_minutes)/60.0 AS unpaid_hours,
                  sum(CASE WHEN final_ledger_status NOT IN ('CLEAR','ABSENCE_RECORDED')
                           THEN 1 ELSE 0 END) AS review_cases
           FROM mart.verint_final_absence_agent_day
           WHERE business_date BETWEEN ? AND ?
           GROUP BY lob, team_leader
           ORDER BY lob, team_leader""",
        [start, end],
    )
    team_sheet = book.table(
        "TEAM_SUMMARY", "Team absence and shrinkage",
        "Filter LOB or Team Leader. Rates use the summed hours shown in the same row.",
        team_headers, team_rows,
    )
    if team_rows:
        for name in ("absence_rate", "shrinkage_rate"):
            column = team_headers.index(name)
            team_sheet.conditional_format(
                4, column, 3 + len(team_rows), column,
                {"type": "3_color_scale", "min_color": COLORS["green_light"],
                 "mid_color": COLORS["amber_light"], "max_color": COLORS["red_light"]},
            )

    agent_headers, agent_rows = _query(
        conn,
        """SELECT lob, team_leader,
                  coalesce(agent_name,'Agent') || ' [' || agent_id || ']' AS agent_selector,
                  agent_id, agent_name, count(*) AS scheduled_days,
                  sum(planned_net_minutes)/60.0 AS planned_hours,
                  sum(final_absence_minutes)/60.0 AS absence_hours,
                  CASE WHEN sum(planned_net_minutes)>0
                       THEN sum(final_absence_minutes)*1.0/sum(planned_net_minutes) END AS absence_rate,
                  sum(final_shrinkage_minutes)/60.0 AS shrinkage_hours,
                  CASE WHEN sum(planned_net_minutes)>0
                       THEN sum(final_shrinkage_minutes)*1.0/sum(planned_net_minutes) END AS shrinkage_rate,
                  sum(final_vacation_minutes)/60.0 AS vacation_hours,
                  sum(final_unpaid_minutes)/60.0 AS unpaid_hours,
                  sum(CASE WHEN final_absence_day THEN 1 ELSE 0 END) AS absence_days,
                  sum(CASE WHEN final_ledger_status NOT IN ('CLEAR','ABSENCE_RECORDED')
                           THEN 1 ELSE 0 END) AS review_cases
           FROM mart.verint_final_absence_agent_day
           WHERE business_date BETWEEN ? AND ?
           GROUP BY lob, team_leader, agent_id, agent_name
           ORDER BY lob, team_leader, agent_name, agent_id""",
        [start, end],
    )
    book.table(
        "AGENT_RESULTS", "Agent absence and shrinkage",
        "Use a personal Sheet View before filtering LOB, Team Leader or Agent Selector.",
        agent_headers, agent_rows,
    )

    action_headers, action_rows = _query(
        conn,
        """SELECT lob, team_leader,
                  coalesce(agent_name,'Agent') || ' [' || agent_id || ']' AS agent_selector,
                  business_date, agent_id, agent_name, final_ledger_status,
                  planned_net_minutes/60.0 AS planned_net_hours,
                  final_absence_minutes/60.0 AS absence_hours,
                  final_shrinkage_minutes/60.0 AS shrinkage_hours,
                  final_unmapped_minutes/60.0 AS unmapped_hours,
                  CASE WHEN final_ledger_status IN ('CLEAR','ABSENCE_RECORDED')
                       THEN 'Pending' ELSE 'Needs review' END AS review_status,
                  NULL AS owner, NULL AS due_date, NULL AS action,
                  NULL AS comment, agent_day_key AS case_id
           FROM mart.verint_final_absence_agent_day
           WHERE business_date BETWEEN ? AND ?
             AND (final_absence_day=true
                  OR final_ledger_status NOT IN ('CLEAR','ABSENCE_RECORDED'))
           ORDER BY business_date DESC, lob, team_leader, agent_name""",
        [start, end],
    )
    editable = ("Review Status", "Owner", "Due Date", "Action", "Comment")
    action_rows = _carry_table_values(
        action_headers, action_rows, "Case ID", editable,
        _previous_table_values(target, "ACTIONS", "Case ID", editable),
    )
    action_sheet = book.table(
        "ACTIONS", "Absence review and follow-up",
        "Filter your team and complete only the blue columns. Case ID keeps saved work attached to the correct agent-day.",
        action_headers, action_rows,
        editable_headers=set(editable),
    )
    if action_rows:
        review_column = action_headers.index("review_status")
        action_sheet.data_validation(
            4, review_column, 3 + len(action_rows), review_column,
            {"validate": "list", "source": ["Pending", "Needs review", "In progress", "Resolved", "No action"]},
        )

    queue_headers, queue_source = _query(
        conn,
        """SELECT agent_day_key AS case_id, business_date,
                  agent_id, agent_name,
                  coalesce(agent_name,'Agent') || ' [' || agent_id || ']' AS agent_selector,
                  team_leader, ops_manager, lob, language,
                  final_ledger_status AS result_status,
                  planned_net_minutes/60.0 AS planned_net_hours,
                  final_absence_minutes/60.0 AS absence_hours,
                  final_shrinkage_minutes/60.0 AS shrinkage_hours,
                  final_unmapped_minutes/60.0 AS unmapped_hours
           FROM mart.verint_final_absence_agent_day
           WHERE business_date BETWEEN ? AND ?
             AND (final_absence_day=true
                  OR final_ledger_status NOT IN ('CLEAR','ABSENCE_RECORDED'))
           ORDER BY business_date DESC, lob, team_leader, agent_name""",
        [data_start, latest],
    )
    queue_headers.append("action_status")
    queue_rows = [
        (*row, '=IFERROR(XLOOKUP([@[Case ID]],tblActions[Case ID],tblActions[Review Status]),"Not started")')
        for row in queue_source
    ]
    queue_sheet = book.table(
        "ACTION_QUEUE", "Absence action queue",
        "Refreshable cases and recorded absence days. Action Status reads the permanent ACTIONS log by Case ID.",
        queue_headers, queue_rows or [tuple(None for _ in queue_headers)],
    )
    if queue_rows:
        action_status_column = queue_headers.index("action_status")
        queue_sheet.conditional_format(
            4, action_status_column, 3 + len(queue_rows), action_status_column,
            {"type": "text", "criteria": "containing", "value": "Not started", "format": book.report.error},
        )

    absence_headers, absence_components = _exclusive_final_components(
        conn, rulebook, start, end, "absence",
    )
    shrinkage_headers, shrinkage_components = _exclusive_final_components(
        conn, rulebook, start, end, "shrinkage",
    )
    book.table(
        "ABSENCE_COMPONENTS", "Absence components",
        "Overlapping Activities are counted once so component hours reconcile to the selected final absence scope.",
        absence_headers, absence_components,
    )
    book.table(
        "SHRINKAGE_COMPONENTS", "Shrinkage components",
        "Overlapping Activities are counted once inside the shrinkage view. Do not add this table to Absence Components.",
        shrinkage_headers, shrinkage_components,
    )
    _add_absence_component_view(book)

    activity_headers, activity_rows = _query(
        conn,
        """SELECT business_date, lob, team_leader,
                  coalesce(agent_name,'Agent') || ' [' || agent_id || ']' AS agent_selector,
                  agent_id, agent_name, activity, category, event_start, event_end,
                  minutes, hours, counts_as_absence, counts_as_vacation,
                  counts_as_unpaid, counts_as_shrinkage, mapped,
                  evidence_type, event_key
           FROM mart.verint_final_absence_event
           WHERE business_date BETWEEN ? AND ?
           ORDER BY business_date, lob, team_leader, agent_name, event_start""",
        [data_start, latest],
    )
    book.table(
        "ACTIVITY_DETAIL", "Final Verint activity detail",
        "Exact classified intervals for investigation. Use component sheets or ABSENCE_DATA for totals.",
        activity_headers, activity_rows,
    )

    data_headers, data_rows = _query(
        conn,
        """SELECT business_date, lob, team_leader,
                  coalesce(agent_name,'Agent') || ' [' || agent_id || ']' AS agent_selector,
                  agent_id, agent_name, ops_manager, language, location,
                  scheduled_minutes/60.0 AS scheduled_hours,
                  planned_net_minutes/60.0 AS planned_net_hours,
                  final_absence_minutes/60.0 AS absence_hours,
                  final_vacation_minutes/60.0 AS vacation_hours,
                  final_unpaid_minutes/60.0 AS unpaid_hours,
                  final_shrinkage_minutes/60.0 AS shrinkage_hours,
                  final_unmapped_minutes/60.0 AS unmapped_hours,
                  final_absence_rate AS absence_rate, final_absence_day,
                  final_ledger_status, agent_day_key AS case_id
           FROM mart.verint_final_absence_agent_day
           WHERE business_date BETWEEN ? AND ?
           ORDER BY business_date, lob, team_leader, agent_name""",
        [data_start, latest],
    )
    book.table(
        "ABSENCE_DATA", "Absence clean data",
        "One agent per day. Use a personal Sheet View before filtering LOB, Team Leader or Agent Selector.",
        data_headers, data_rows,
    )
    book.table(
        "HELP", "How to use this report", "A short operating guide for the shared workbook.",
        ["Step", "What to do", "Why"],
        [
            (1, "Run WFM Hub refresh and confirm the latest source date.", "Updates the clean Absenteeism feeds."),
            (2, "Use TEAM_VIEW for period, LOB, Team Leader and Agent selection.", "Agent results, cases and components follow one selection."),
            (3, "For a permanent shared file, connect ABSENCE_DATA, ACTION_QUEUE and ACTIVITY_DETAIL once to the three fixed CSV feeds.", "Data > Refresh All updates facts without replacing the workbook."),
            (4, "Copy new Case IDs from ACTION_QUEUE into ACTIONS and fill only the blue fields.", "The permanent action log remains safe while facts refresh."),
            (5, "Review Finalized coverage and unresolved cases before sharing totals.", "Incomplete rows must not dilute the rate."),
            (6, "Use COMPONENT_VIEW for totals and ACTIVITY_DETAIL for exact intervals.", "Raw intervals may overlap; KPI components remain separate."),
        ],
    )
    book.definitions([
        ("Absence rate", "Final absence minutes / finalized planned net minutes", "Payroll and attendance result", "Incomplete cases are shown separately"),
        ("Shrinkage rate", "Final shrinkage minutes / finalized planned net minutes", "Capacity loss", "A parallel view; do not add to absence rate"),
        ("Finalized coverage", "Finalized planned minutes / all planned minutes", "Confidence in the headline", "Review when below 100%"),
        ("Uncoded empty shift", "Completed scheduled shift with no final code and no reliable login evidence", "Correction completeness", "Never treated as zero absence"),
        ("Component", "One exclusive activity classification inside its KPI view", "Management breakdown", "Raw overlapping intervals are counted once"),
    ])
    _add_absence_lookups(book)
    book.audit(_audit_rows(
        conn, config, "absence", start, end,
        (
            ("Shared feed", str(config.feed / "Absenteeism"), "Updated with this report"),
            ("Template version", "absence-2026.09.2", "Collaboration report contract"),
            ("All planned hours", all_planned / 60, f"{agent_days:,} agent-day row(s)"),
            ("All absence hours", all_absence / 60, "Includes review rows"),
            ("All shrinkage hours", all_shrinkage / 60, "Includes review rows"),
            ("All vacation hours", all_vacation / 60, "Includes review rows"),
            ("All unpaid hours", all_unpaid / 60, "Includes review rows"),
        ),
    ))
    return _finish(book, partial, target)


def build_attendance_corrections_workbook(
    conn: DatabaseConnection,
    config: Config,
    start: date,
    end: date,
    output: Path | None = None,
) -> Path:
    """Build the selected period's completed-day correction queue and timeline."""

    from .shift_view import add_shift_view

    today = date.today()
    completed_through = min(end, today - timedelta(days=1))
    book, partial, target = _atomic_book(
        config, "corrections", "ATTENDANCE REVIEW", start, end, output,
    )
    gap_count, gap_minutes, agents = conn.execute(
        """SELECT count(*), coalesce(sum(residual_minutes),0), count(DISTINCT agent_id)
           FROM mart.correction_residual_segment
           WHERE business_date BETWEEN ? AND ? AND business_date<?""",
        [start, end, today],
    ).fetchone()
    missing = conn.execute(
        """SELECT count(*) FROM mart.attendance_agent_day
           WHERE business_date BETWEEN ? AND ? AND business_date<?
             AND assignment_type NOT IN ('Off','Planned absence')
             AND attendance_result IN
               ('Schedule parse error','Data not loaded','Missing actual evidence',
                'Incomplete actual evidence','No schedule overlap')""",
        [start, end, today],
    ).fetchone()[0]
    status, status_text = _source_state(
        conn, ("fte", "start_end", "lilo", "agent_status", "activities"), completed_through, final=True,
    )
    if missing:
        status, status_text = "INCOMPLETE", f"{missing:,} scheduled row(s) lack complete observed evidence"
    book.dashboard(
        [
            KpiCard("Residual segments", gap_count, "integer"),
            KpiCard("Residual gap hours", gap_minutes / 60 if gap_minutes else 0, "decimal"),
            KpiCard("Agents to review", agents, "integer"),
            KpiCard("Missing evidence", missing, "integer"),
        ],
        status,
        status_text,
        ["Measure", "Value"],
        [("Residual segments", gap_count), ("Residual gap hours", gap_minutes / 60 if gap_minutes else 0), ("Agents to review", agents), ("Missing evidence", missing)],
        [
            f"The selected range is {start:%Y-%m-%d} to {end:%Y-%m-%d}; every completed date through {completed_through:%Y-%m-%d} is included, not only yesterday.",
            "Today is excluded from correction review, so an unfinished shift can never become an early-leave correction.",
            "VERINT_INJECTION contains one row per exact continuous residual interval; Start to inject and End to inject are never rounded to a 15-minute grid.",
            "After injection, export Activities again and refresh. Corrected coverage disappears automatically; no decision workbook is imported back into the Hub.",
            "SHIFT_VIEW is supporting evidence for the listed agents: schedule versus Agent Status/LILO over the full shift.",
            "Verint Activities verify whether an observed gap is corrected; they never create the original gap.",
        ],
    )
    headers, rows = _query(
        conn,
        """SELECT r.business_date, r.agent_id, c.agent_name, c.team_leader,
                  c.ops_manager, c.lob, d.language,
                  c.detected_issue AS incoherence,
                  r.residual_start AS start_to_inject,
                  r.residual_end AS end_to_inject,
                  r.residual_minutes AS minutes,
                  r.suggested_activity, c.confidence,
                  r.verint_reconciliation, r.observed_source,
                  c.scheduled_start, c.scheduled_end,
                  r.source_file, r.correction_id, r.residual_id AS segment_id
           FROM mart.correction_residual_segment r
           JOIN mart.correction_candidate c ON c.correction_id=r.correction_id
           LEFT JOIN core.dim_agent d ON d.agent_id=r.agent_id
           WHERE r.business_date BETWEEN ? AND ? AND r.business_date<?
           ORDER BY r.business_date, c.priority, r.residual_minutes DESC, r.agent_id, r.residual_start""",
        [start, end, today],
    )
    book.table(
        "VERINT_INJECTION", "Exact residual intervals ready for Verint",
        "One row is one continuous interval still uncovered in the latest Activities export. Use the exact timestamps shown; do not round or combine separate rows.",
        headers, rows,
    )

    timeline_headers, timeline_rows = _query(
        conn,
        """SELECT business_date, agent_id, agent_name, team_leader,
                  ops_manager, lob, language, scheduled_start, scheduled_end,
                  segment_start, segment_end, segment_minutes, planned_state,
                  actual_status, actual_category, mismatch_type, is_gap,
                  observed_source, source_file, evaluation_as_of
           FROM mart.shift_timeline_segment t
           WHERE business_date BETWEEN ? AND ? AND business_date<?
             AND EXISTS (
                 SELECT 1 FROM mart.correction_residual_segment r
                 WHERE r.business_date=t.business_date AND r.agent_id=t.agent_id
             )
           ORDER BY business_date, agent_id, segment_start""",
        [start, end, today],
    )
    timeline_dicts = [dict(zip(timeline_headers, row)) for row in timeline_rows]
    add_shift_view(book.report, timeline_dicts, start, completed_through)
    book.tables.append(ModelTable("TIMELINE", timeline_headers, timeline_rows))
    book.definitions([
        ("Observed gap", "Scheduled time minus unioned LILO/Agent Status evidence", "Correction candidate", "Activities cannot create the gap"),
        ("Residual gap", "Observed gap minus unioned corrected Verint Activities", "Minutes still requiring review", "Overlap-safe"),
        ("Start/End to inject", "Exact physical boundaries of one continuous residual interval", "Manual Verint entry", "Never rounded and never bridges a real return to service"),
        ("Current-day tail", "Future portion of an unfinished shift", "No action", "Never early leave"),
        ("Correction ID", "Stable detected-gap lineage key", "Audit and reconciliation", "One correction may have several residual segments"),
        ("Segment ID", "Stable exact residual-interval key", "Trace one injection row", "Changes only when its physical residual interval changes"),
    ])
    book.audit(_audit_rows(
        conn, config, "corrections", start, end,
        (("Completed-date cutoff", completed_through, "Today is excluded from correction review"),),
    ))
    return _finish(book, partial, target)
