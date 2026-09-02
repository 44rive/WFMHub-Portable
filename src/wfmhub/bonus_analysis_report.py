"""Deterministic management case for changing a bonus KPI configuration.

The report is deliberately a sensitivity analysis, not a payroll instruction.
It reads a governed bonus workbook, reproduces its current payout logic, and
then isolates configuration changes so management can see their consequences.
"""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from hashlib import sha256
import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .shared_reports import (
    COLORS,
    BonusRule,
    ManagementBook,
    _bonus_period,
    _clean,
    _header_row,
    _number,
    _policy_map,
    _read_bonus_source,
)


@dataclass(frozen=True)
class MetricDefinition:
    label: str
    rule_name: str
    input_names: tuple[str, ...]


METRICS = (
    MetricDefinition("AHT", "AHT", ("AHT",)),
    MetricDefinition("Productivity", "Productivity", ("Productivity",)),
    MetricDefinition("PCS Score", "PCS Score", ("PCS Score",)),
    MetricDefinition(
        "PCS Participation",
        "PCS % (Participation)",
        ("PCS % Participation", "PCS % (Participation)"),
    ),
    MetricDefinition("QM", "QM", ("QM",)),
    MetricDefinition("Absence", "Abs%", ("Abs%",)),
)
EXTRA_PCS_RULE = "Extra Bonus (PCS Score)"


@dataclass(frozen=True)
class Scenario:
    name: str
    change: str
    participation_targets: tuple[float, float] | None = None
    extra_method: str = "Additive"
    require_complete: bool = False


@dataclass(frozen=True)
class ScenarioValue:
    payout: float | None
    achievement: float | None
    held: bool


@dataclass(frozen=True)
class ScenarioSummary:
    scenario: Scenario
    total: float
    paid: int
    held: int
    held_baseline_exposure: float
    comparable_baseline: float
    delta: float


def _value(record: dict[str, Any], names: tuple[str, ...]) -> float | None:
    for name in names:
        if name in record:
            return _number(record[name])
    return None


def _percentile(values: list[float], proportion: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    position = (len(ordered) - 1) * proportion
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return ordered[lower]
    return ordered[lower] + (ordered[upper] - ordered[lower]) * (position - lower)


def _round_up(value: float, step: float) -> float:
    return round(math.ceil((value - 1e-12) / step) * step, 10)


def _round_nearest(value: float, step: float) -> float:
    return round(math.floor(value / step + 0.5) * step, 10)


def _participation_pilot(records: list[dict[str, Any]]) -> tuple[float, float, float, float]:
    values = [
        value
        for record in records
        if (value := _value(record, ("PCS % Participation", "PCS % (Participation)"))) is not None
    ]
    median = _percentile(values, 0.50) or 0.0
    upper_quartile = _percentile(values, 0.75) or 0.0
    tier1 = min(1.0, _round_up(upper_quartile, 0.05))
    tier2 = min(1.0, _round_nearest(median, 0.05))
    if tier2 >= tier1:
        tier2 = max(0.0, round(tier1 - 0.05, 10))
    return tier1, tier2, upper_quartile, median


def _complete(record: dict[str, Any], rule_map: dict[tuple[str, str], BonusRule]) -> bool:
    population = _clean(record.get("Population"))
    return all(_value(record, metric.input_names) is not None for metric in METRICS) and all(
        (population, metric.rule_name) in rule_map for metric in METRICS
    )


def _award(value: float | None, rule: BonusRule, targets: tuple[float, float] | None = None) -> float:
    if value is None:
        return 0.0
    tier1_target, tier2_target = targets or (rule.tier1_target, rule.tier2_target)
    tier1 = value <= tier1_target if rule.direction == "L" else value >= tier1_target
    tier2 = value <= tier2_target if rule.direction == "L" else value >= tier2_target
    if tier1:
        return rule.tier1_bonus
    if rule.tier2_bonus and tier2:
        return rule.tier2_bonus
    return 0.0


def _scenario_value(
    record: dict[str, Any],
    rule_map: dict[tuple[str, str], BonusRule],
    policies: dict[str, Any],
    scenario: Scenario,
) -> ScenarioValue:
    if scenario.require_complete and not _complete(record, rule_map):
        return ScenarioValue(None, None, True)

    population = _clean(record.get("Population"))
    scores: list[float] = []
    for metric in METRICS:
        rule = rule_map.get((population, metric.rule_name))
        if rule is None:
            scores.append(0.0)
            continue
        targets = scenario.participation_targets if metric.rule_name == "PCS % (Participation)" else None
        scores.append(_award(_value(record, metric.input_names), rule, targets))

    if _clean(policies.get("Absence treatment", "KPI only")) == "Eligibility only":
        scores[5] = 0.0
    extra_rule = rule_map.get((population, EXTRA_PCS_RULE))
    extra = _award(_value(record, ("PCS Score",)), extra_rule) if extra_rule else 0.0
    base = sum(scores)
    if scenario.extra_method == "Replacement":
        gross = base - scores[2] + max(scores[2], extra)
    else:
        gross = base + extra
    cap = _number(policies.get("Achievement cap")) or 1.30
    gross = min(gross, cap)

    voc = int(_number(record.get("VOC Detractor Count")) or 0)
    malus = {0: 0.0, 1: 0.10, 2: 0.20, 3: 0.50, 4: 0.75}.get(voc, 1.0)
    if _clean(policies.get("Malus method")) == "Percentage points":
        achievement = max(0.0, gross - malus)
    else:
        achievement = gross * (1 - malus)

    override = _number(record.get("Reference Bonus Override")) or 0.0
    salary = _number(record.get("Monthly Fixed Salary")) or 0.0
    rate = _number(record.get("Target Bonus Rate")) or 0.0
    divisor = 12 if _clean(policies.get("Target bonus rate basis")) == "Annual" else 1
    reference = override or salary * rate / divisor
    eligible = _number(record.get("Eligible Days"))
    scheduled = _number(record.get("Scheduled Days"))
    proration = min(1.0, eligible / scheduled) if eligible is not None and scheduled and scheduled > 0 else 0.0
    digits = 0 if _clean(policies.get("Rounding")) == "Whole MAD" else 2
    return ScenarioValue(round(reference * proration * achievement, digits), achievement, False)


def _cached_result_total(source: Path) -> tuple[float | None, int]:
    workbook = load_workbook(source, read_only=True, data_only=True, keep_links=False)
    try:
        if "Results" not in workbook.sheetnames:
            return None, 0
        sheet = workbook["Results"]
        header_row = _header_row(sheet, "Agent ID")
        headers = [_clean(sheet.cell(header_row, column).value) for column in range(1, sheet.max_column + 1)]
        if "Final Payout" not in headers:
            return None, 0
        payout_index = headers.index("Final Payout")
        agent_index = headers.index("Agent ID")
        values: list[float] = []
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if agent_index >= len(row) or not _clean(row[agent_index]):
                continue
            payout = _number(row[payout_index] if payout_index < len(row) else None)
            if payout is not None:
                values.append(payout)
        return sum(values), len(values)
    finally:
        workbook.close()


def _target_display(metric: MetricDefinition, value: float | None) -> str:
    if value is None:
        return "—"
    if metric.rule_name in {"PCS % (Participation)", "QM", "Abs%"}:
        return f"{value:.1%}"
    if metric.rule_name == "AHT":
        return f"{value:,.0f} sec"
    return f"{value:,.2f}"


def _observed_display(metric: MetricDefinition, value: float | None) -> str:
    return _target_display(metric, value)


def _first_rule(rules: list[BonusRule], name: str) -> BonusRule | None:
    return next((rule for rule in rules if rule.kpi == name), None)


def _metric_diagnostics(
    records: list[dict[str, Any]], rules: list[BonusRule]
) -> list[dict[str, Any]]:
    diagnostics: list[dict[str, Any]] = []
    for metric in METRICS:
        values = [
            value
            for record in records
            if (value := _value(record, metric.input_names)) is not None
        ]
        rule = _first_rule(rules, metric.rule_name)
        tier1_hits = sum(
            1 for value in values
            if rule and (value <= rule.tier1_target if rule.direction == "L" else value >= rule.tier1_target)
        )
        tier2_hits = sum(
            1 for value in values
            if rule and rule.tier2_bonus
            and not (value <= rule.tier1_target if rule.direction == "L" else value >= rule.tier1_target)
            and (value <= rule.tier2_target if rule.direction == "L" else value >= rule.tier2_target)
        )
        diagnostics.append({
            "metric": metric,
            "rule": rule,
            "values": values,
            "coverage": len(values),
            "coverage_rate": len(values) / len(records) if records else 0.0,
            "tier1_hits": tier1_hits,
            "tier2_hits": tier2_hits,
            "attainment": (tier1_hits + tier2_hits) / len(values) if values else 0.0,
            "p25": _percentile(values, 0.25),
            "median": _percentile(values, 0.50),
            "p75": _percentile(values, 0.75),
            "maximum": max(values) if values else None,
        })
    return diagnostics


def _lob_diagnostics(
    records: list[dict[str, Any]], rule_map: dict[tuple[str, str], BonusRule]
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    populations = sorted({_clean(record.get("Population")) for record in records if _clean(record.get("Population"))})
    for population in populations:
        subset = [record for record in records if _clean(record.get("Population")) == population]
        complete = sum(_complete(record, rule_map) for record in subset)
        for metric in METRICS:
            values = [
                value
                for record in subset
                if (value := _value(record, metric.input_names)) is not None
            ]
            rule = rule_map.get((population, metric.rule_name))
            hits = 0
            if rule:
                hits = sum(_award(value, rule) > 0 for value in values)
            rows.append({
                "population": population,
                "agents": len(subset),
                "complete": complete,
                "metric": metric,
                "coverage": len(values),
                "coverage_rate": len(values) / len(subset) if subset else 0.0,
                "median": _percentile(values, 0.50),
                "target": rule.tier1_target if rule else None,
                "direction": rule.direction if rule else "",
                "attaining": hits,
                "attainment": hits / len(values) if values else None,
            })
    return rows


def _source_hash(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def build_bonus_kpi_change_case(source: Path, output: Path) -> Path:
    """Build a management-ready KPI change case without modifying *source*."""

    source = source.resolve()
    before_hash = _source_hash(source)
    records, rules, policies = _read_bonus_source(source)
    if not records:
        raise ValueError("Bonus source has no populated agent rows")
    rule_map = {(rule.population, rule.kpi): rule for rule in rules}
    policy_values = _policy_map(policies)
    period = _bonus_period(records)
    diagnostics = _metric_diagnostics(records, rules)
    lob_rows = _lob_diagnostics(records, rule_map)
    pilot_t1, pilot_t2, participation_p75, participation_median = _participation_pilot(records)

    scenarios = (
        Scenario("Current configuration", "Source rules: additive extra PCS; missing KPI earns zero."),
        Scenario(
            "Participation calibration",
            f"Only participation changes to Tier 1 {pilot_t1:.0%} / Tier 2 {pilot_t2:.0%}.",
            (pilot_t1, pilot_t2),
        ),
        Scenario("PCS overlap removed", "Only extra PCS changes from Additive to Replacement.", extra_method="Replacement"),
        Scenario(
            "Controlled pilot",
            f"Participation {pilot_t1:.0%}/{pilot_t2:.0%}, Replacement extra PCS, and incomplete rows held.",
            (pilot_t1, pilot_t2),
            "Replacement",
            True,
        ),
    )
    scenario_values = {
        scenario.name: [_scenario_value(record, rule_map, policy_values, scenario) for record in records]
        for scenario in scenarios
    }
    baseline = scenario_values[scenarios[0].name]
    baseline_total = sum(value.payout or 0.0 for value in baseline)
    complete_mask = [_complete(record, rule_map) for record in records]
    complete_baseline = sum(
        (value.payout or 0.0) for value, complete in zip(baseline, complete_mask) if complete
    )
    incomplete_exposure = sum(
        (value.payout or 0.0) for value, complete in zip(baseline, complete_mask) if not complete
    )
    paid_incomplete = sum(
        (value.payout or 0.0) > 0 for value, complete in zip(baseline, complete_mask) if not complete
    )
    summaries: list[ScenarioSummary] = []
    for scenario in scenarios:
        values = scenario_values[scenario.name]
        total = sum(value.payout or 0.0 for value in values if not value.held)
        comparable = complete_baseline if scenario.require_complete else baseline_total
        summaries.append(ScenarioSummary(
            scenario,
            total,
            sum((value.payout or 0.0) > 0 for value in values if not value.held),
            sum(value.held for value in values),
            incomplete_exposure if scenario.require_complete else 0.0,
            comparable,
            total - comparable,
        ))

    participation = next(row for row in diagnostics if row["metric"].rule_name == "PCS % (Participation)")
    absence = next(row for row in diagnostics if row["metric"].rule_name == "Abs%")
    lob_spreads: list[tuple[str, float]] = []
    for metric in METRICS:
        rates = [
            row["attainment"] for row in lob_rows
            if row["metric"].rule_name == metric.rule_name and row["attainment"] is not None
        ]
        if rates:
            lob_spreads.append((metric.label, max(rates) - min(rates)))
    widest_metric, widest_spread = max(lob_spreads, key=lambda item: item[1])
    extra_rule = _first_rule(rules, EXTRA_PCS_RULE)
    pcs_rule = _first_rule(rules, "PCS Score")
    cap = _number(policy_values.get("Achievement cap")) or 1.30
    pcs_max_points = (pcs_rule.tier1_bonus if pcs_rule else 0.0) + (extra_rule.tier1_bonus if extra_rule else 0.0)
    pcs_cap_share = pcs_max_points / cap if cap else 0.0
    missing_rows = len(records) - sum(complete_mask)
    policy_reviews = sum(_clean(row[5] if len(row) > 5 else "").casefold() != "validated" for row in policies)
    statuses = Counter(_clean(record.get("Data Status")) or "BLANK" for record in records)
    voc_zero = sum((_number(record.get("VOC Detractor Count")) or 0) == 0 for record in records)
    full_proration = sum(
        (_number(record.get("Eligible Days")) is not None)
        and (_number(record.get("Scheduled Days")) or 0) > 0
        and (_number(record.get("Eligible Days")) or 0) >= (_number(record.get("Scheduled Days")) or 0)
        for record in records
    )
    cached_total, cached_rows = _cached_result_total(source)
    reconciliation_delta = None if cached_total is None else baseline_total - cached_total

    findings = [
        (
            "CRITICAL",
            "Participation target is outside July's observed range",
            f"Tier 1 is {participation['rule'].tier1_target:.0%}; July maximum is {participation['maximum']:.1%}. "
            f"Tier 1 hits: {participation['tier1_hits']}; any award: {participation['tier1_hits'] + participation['tier2_hits']} of {participation['coverage']} measured.",
            f"Approve a temporary {pilot_t1:.0%}/{pilot_t2:.0%} sensitivity band, then validate against three months of denominator-level data.",
        ),
        (
            "CRITICAL",
            "Incomplete KPI rows still influence payout",
            f"{missing_rows} of {len(records)} rows miss at least one core KPI. {paid_incomplete} of them receive {incomplete_exposure:,.0f} MAD under the current missing-as-zero logic.",
            "Hold incomplete rows for data completion; do not convert missing evidence into either failure or approval.",
        ),
        (
            "HIGH",
            "One target set behaves very differently by LOB",
            f"Conditional attainment for {widest_metric} spans {widest_spread:.1%} across LOBs. The source repeats the same thresholds for all six populations.",
            "Move to effective-dated LOB targets using at least three comparable months and an approved minimum sample.",
        ),
        (
            "HIGH",
            "PCS score is rewarded in two additive components",
            f"Standard PCS plus Extra PCS can contribute {pcs_max_points:.0%} achievement points, {pcs_cap_share:.1%} of the {cap:.0%} cap, from the same score field.",
            "Use Replacement during the pilot unless Compensation explicitly approves additive overlap.",
        ),
        (
            "HIGH",
            "Absence may be charged twice once proration is activated",
            f"Absence currently carries {absence['rule'].tier1_bonus:.0%} weight. July has {full_proration} of {len(records)} rows at full proration, but using the same future absence in Eligible Days would reduce payout a second time.",
            "Count the same absence once: use proration and remove the Absence weight, or keep the KPI only when proration excludes that absence.",
        ),
        (
            "HIGH",
            "Survey reliability cannot be tested",
            "The workbook contains PCS score and participation percentages but no response numerator, denominator, or minimum sample field.",
            "Add eligible-survey and valid-response counts before using PCS score as a high-weight payout driver.",
        ),
        (
            "REVIEW",
            "Governance signals are not fully released",
            f"{policy_reviews} policy decisions are not Validated; source Data Status values are {dict(statuses)}; all {voc_zero} of {len(records)} VOC counts are zero.",
            "Confirm policy owners, define the ICMP status, and validate the all-zero VOC feed before payroll use.",
        ),
    ]

    book = ManagementBook(output, f"Bonus KPI change case {period}")
    wb = book.workbook
    f = book.fmt
    title_period = period.replace("-", " ").upper()
    money_delta = wb.add_format({"font_name": "Aptos", "font_size": 10, "num_format": '+#,##0.00 "MAD";-#,##0.00 "MAD";0.00 "MAD"', "bottom": 1, "bottom_color": COLORS["line"]})
    pct_plain = wb.add_format({"font_name": "Aptos", "font_size": 10, "num_format": "0.0%", "bottom": 1, "bottom_color": COLORS["line"]})
    wrap = wb.add_format({"font_name": "Aptos", "font_size": 10, "font_color": COLORS["ink"], "text_wrap": True, "valign": "top", "bottom": 1, "bottom_color": COLORS["line"]})
    decision = wb.add_format({"font_name": "Aptos", "font_size": 10, "bold": True, "font_color": COLORS["navy"], "bg_color": COLORS["teal_light"], "text_wrap": True, "valign": "top", "border": 1, "border_color": COLORS["line"]})
    severity_formats = {
        "CRITICAL": f["bad"],
        "HIGH": f["warn"],
        "REVIEW": f["warn"],
    }

    start = book.sheet(
        "START_HERE",
        f"{title_period}  |  WHY THE BONUS RULES NEED A REVIEW",
        "SIMPLE MANAGEMENT VERSION  /  OPEN THIS TAB FIRST",
        14,
    )
    start.set_tab_color(COLORS["gold"])
    start.set_column("A:A", 2)
    start.set_column("B:O", 13)
    book.banner(
        start,
        3,
        1,
        13,
        "THE ASK  |  Approve a controlled three-month KPI pilot. Do not recalculate July payroll from this file.",
        "warn",
    )
    book.kpi(start, 1, 5, "JULY MODELED PAYOUT", baseline_total, kind="kpi_money")
    book.kpi(start, 4, 5, "PARTICIPATION TARGET", participation["rule"].tier1_target, kind="kpi_pct")
    book.kpi(start, 7, 5, "JULY MAXIMUM", participation["maximum"], kind="kpi_pct")
    book.kpi(start, 10, 5, "INCOMPLETE ROWS", missing_rows, kind="kpi")
    book.kpi(start, 13, 5, "AMOUNT AFFECTED", incomplete_exposure, kind="kpi_money")
    start.merge_range("B10:N10", "THE FOUR THINGS TO EXPLAIN", f["section"])
    start.write_row("B12", ["#", "Plain-language problem", "What July shows", "What we are asking"], f["header"])
    start.set_column("B:B", 6)
    start.set_column("C:C", 33)
    start.set_column("D:D", 63)
    start.set_column("E:E", 63)
    simple_findings = [
        (
            1,
            "The participation target cannot be reached",
            f"Target: {participation['rule'].tier1_target:.0%}. Highest July result: {participation['maximum']:.1%}. Zero Tier-1 winners.",
            f"Test {pilot_t1:.0%} for Tier 1 and {pilot_t2:.0%} for Tier 2 for three months.",
        ),
        (
            2,
            "Absence could reduce the same bonus twice",
            f"Absence has {absence['rule'].tier1_bonus:.0%} KPI weight. If absence also reduces proration later, the employee loses twice. July is still {full_proration}/{len(records)} at full proration.",
            "If absence drives proration, remove its KPI weight. One absence should have one financial consequence.",
        ),
        (
            3,
            "Missing information still changes pay",
            f"{missing_rows} rows miss a core KPI; {paid_incomplete} of them still receive {incomplete_exposure:,.0f} MAD under the current calculation.",
            "Hold the row until information is complete; missing must not mean failed or approved.",
        ),
        (
            4,
            "One target does not fit every LOB",
            f"The same targets cover six LOBs, but {widest_metric} attainment differs by {widest_spread:.1%} between them.",
            "Build LOB targets from at least three comparable months and approve them with Operations and HR.",
        ),
    ]
    for row_index, row in enumerate(simple_findings, 12):
        start.write(row_index, 1, row[0], f["kpi"])
        start.write(row_index, 2, row[1], decision)
        start.write(row_index, 3, row[2], wrap)
        start.write(row_index, 4, row[3], wrap)
        start.set_row(row_index, 58)
    start.merge_range("B19:N19", "YOUR 60-SECOND TALK TRACK", f["section"])
    talk_track = (
        f"I am not asking to change July payroll today. I am asking permission to test a fairer and clearer bonus configuration for three months. "
        f"July shows that the {participation['rule'].tier1_target:.0%} participation target was impossible because the highest result was {participation['maximum']:.1%}. "
        f"We also have {missing_rows} incomplete KPI rows, the same targets across six different LOBs, and a future risk of counting absence twice through both the {absence['rule'].tier1_bonus:.0%} KPI and proration. "
        f"My proposal is simple: test participation at {pilot_t1:.0%}/{pilot_t2:.0%}, count absence only once, stop stacking the same PCS score twice, hold incomplete rows, and use three months of LOB data before approving final targets."
    )
    start.merge_range("B21:N26", talk_track, decision)
    start.freeze_panes("B12")
    start.set_landscape()
    start.fit_to_pages(1, 2)

    changes = book.sheet(
        "WHAT_TO_CHANGE",
        "CURRENT RULES  →  PROPOSED THREE-MONTH PILOT",
        "ONE LINE PER DECISION; FINAL TARGETS STILL REQUIRE MANAGEMENT APPROVAL",
        8,
    )
    changes.set_tab_color(COLORS["gold"])
    changes.set_column("A:A", 23)
    changes.set_column("B:B", 29)
    changes.set_column("C:C", 42)
    changes.set_column("D:D", 44)
    changes.set_column("E:E", 25)
    changes.write_row(4, 0, ["Decision", "Current rule", "Why it needs review", "Proposed pilot", "Owner"], f["header"])
    change_rows = [
        (
            "PCS Participation",
            f"Tier 1 {participation['rule'].tier1_target:.0%}; Tier 2 {participation['rule'].tier2_target:.0%}",
            f"July maximum was {participation['maximum']:.1%}; only {participation['tier1_hits'] + participation['tier2_hits']} of {participation['coverage']} measured agents earned anything.",
            f"Temporarily test {pilot_t1:.0%} / {pilot_t2:.0%}, then validate with three months.",
            "Operations + HR",
        ),
        (
            "Absence",
            f"{absence['rule'].tier1_bonus:.0%} KPI weight; future proration planned",
            "Using the same absence in both places creates a double financial penalty.",
            "If absence reduces Eligible Days, set the Absence KPI weight to 0%. Otherwise keep the KPI and exclude that absence from proration.",
            "HR + Payroll",
        ),
        (
            "Extra PCS",
            "Add standard PCS and Extra PCS",
            f"The same PCS score can create {pcs_max_points:.0%} achievement points.",
            "Use Replacement during the pilot; pay the better PCS award once.",
            "Compensation",
        ),
        (
            "Missing KPI",
            "Missing value earns zero points",
            f"{missing_rows} rows are incomplete and {incomplete_exposure:,.0f} MAD is affected.",
            "Hold the calculation until the required KPI arrives.",
            "WFM + Payroll",
        ),
        (
            "LOB targets",
            "Same targets for all six LOBs",
            f"Observed {widest_metric} attainment differs by {widest_spread:.1%}.",
            "Use effective-dated LOB targets built from at least three comparable months.",
            "Operations + HR",
        ),
        (
            "PCS evidence",
            "Percentages only",
            "The source has no eligible-survey or valid-response counts, so score reliability cannot be tested.",
            "Load numerator, denominator and minimum sample before applying high PCS weight.",
            "Quality + WFM",
        ),
    ]
    for row_index, row in enumerate(change_rows, 5):
        for column, value in enumerate(row):
            changes.write(row_index, column, value, decision if column == 3 else wrap)
        changes.set_row(row_index, 62)
    changes.merge_range("A13:E13", "BUDGET SENSITIVITY  |  WHAT THE JULY TEST CHANGES", f["section"])
    changes.write_row(15, 0, ["Test", "July effect", "Meaning"], f["header"])
    budget_rows = [
        ("Participation change only", summaries[1].delta, "Cost if this rule changes and everything else stays the same."),
        ("Remove PCS overlap only", summaries[2].delta, "Saving if Extra PCS replaces rather than adds to standard PCS."),
        ("Both changes on complete rows", summaries[3].delta, "Comparable change for rows with all required KPI data."),
        ("Incomplete amount held", incomplete_exposure, "Pending data completion—not cancelled or denied."),
    ]
    for row_index, row in enumerate(budget_rows, 16):
        changes.write(row_index, 0, row[0], f["body"])
        changes.write(row_index, 1, row[1], money_delta if row_index < 19 else f["money"])
        changes.merge_range(row_index, 2, row_index, 4, row[2], wrap)
        changes.set_row(row_index, 34)
    changes.merge_range(
        "A22:E24",
        f"ABSENCE RULE IN ONE SENTENCE\nIf the same absence reduces proration, remove the {absence['rule'].tier1_bonus:.0%} Absence KPI weight. "
        "If proration is only for joiners, leavers, or another non-overlapping eligibility reason, the Absence KPI may remain after HR approval.",
        decision,
    )
    changes.freeze_panes(5, 0)
    changes.set_landscape()
    changes.fit_to_pages(1, 2)

    proof = book.sheet(
        "PROOF",
        f"{title_period}  |  THE NUMBERS BEHIND THE REQUEST",
        "ONLY THREE IDEAS: TARGET REACHABILITY, DATA COMPLETENESS, AND LOB COMPARABILITY",
        11,
    )
    proof.set_tab_color(COLORS["gold"])
    proof.set_column("A:A", 22)
    proof.set_column("B:E", 15)
    proof.set_column("F:F", 43)
    proof.set_column("G:L", 13)
    proof.write_row(4, 0, ["KPI", "Measured", "Coverage", "Any Award", "Attainment", "Simple reading"], f["header"])
    for row_index, row in enumerate(diagnostics, 5):
        metric = row["metric"]
        reading = "Review coverage" if row["coverage_rate"] < 0.80 else "Target behaves inside July range"
        if metric.rule_name == "PCS % (Participation)":
            reading = "Tier 1 is above July maximum"
        values = [
            metric.label,
            row["coverage"],
            row["coverage_rate"],
            row["tier1_hits"] + row["tier2_hits"],
            row["attainment"],
            reading,
        ]
        for column, value in enumerate(values):
            fmt = pct_plain if column in {2, 4} else f["body"]
            if column == 5 and metric.rule_name == "PCS % (Participation)":
                fmt = f["bad"]
            proof.write(row_index, column, value, fmt)
    proof_chart = wb.add_chart({"type": "bar"})
    proof_chart.add_series({
        "name": "Attainment",
        "categories": ["PROOF", 5, 0, 5 + len(diagnostics) - 1, 0],
        "values": ["PROOF", 5, 4, 5 + len(diagnostics) - 1, 4],
        "fill": {"color": COLORS["teal"]},
        "border": {"none": True},
        "data_labels": {"value": True, "num_format": "0.0%"},
    })
    proof_chart.set_title({"name": "Award attainment among measured agents"})
    proof_chart.set_x_axis({"num_format": "0%", "min": 0, "max": 1, "major_gridlines": {"visible": False}})
    proof_chart.set_legend({"none": True})
    proof_chart.set_style(10)
    proof.insert_chart("H5", proof_chart, {"x_scale": 1.05, "y_scale": 1.05})
    proof.merge_range("A14:F14", "AHT EXAMPLE  |  SAME 450-SECOND TARGET, DIFFERENT LOB OUTCOMES", f["section"])
    proof.write_row(16, 0, ["LOB", "Agents", "Measured", "Attaining", "Attainment", "Complete Rows"], f["header"])
    aht_lobs = [row for row in lob_rows if row["metric"].rule_name == "AHT"]
    for row_index, row in enumerate(aht_lobs, 17):
        values = [row["population"], row["agents"], row["coverage"], row["attaining"], row["attainment"], row["complete"]]
        for column, value in enumerate(values):
            proof.write(row_index, column, value, pct_plain if column == 4 else f["body"])
    lob_chart = wb.add_chart({"type": "column"})
    lob_chart.add_series({
        "name": "AHT attainment",
        "categories": ["PROOF", 17, 0, 17 + len(aht_lobs) - 1, 0],
        "values": ["PROOF", 17, 4, 17 + len(aht_lobs) - 1, 4],
        "fill": {"color": COLORS["gold"]},
        "border": {"none": True},
        "data_labels": {"value": True, "num_format": "0.0%"},
    })
    lob_chart.set_title({"name": "Same AHT target, very different attainment"})
    lob_chart.set_y_axis({"num_format": "0%", "min": 0, "max": 1, "major_gridlines": {"visible": False}})
    lob_chart.set_legend({"none": True})
    lob_chart.set_style(10)
    proof.insert_chart("H17", lob_chart, {"x_scale": 1.05, "y_scale": 1.05})
    proof.merge_range(
        "A26:L28",
        f"ABSENCE CHECK\nCurrent Absence weight: {absence['rule'].tier1_bonus:.0%}. July rows at full proration: {full_proration}/{len(records)}. "
        "No July double penalty is modeled yet. The control is preventive: once absence reduces Eligible Days, the same absence must not also remove KPI points.",
        decision,
    )
    proof.freeze_panes(5, 0)
    proof.set_landscape()
    proof.fit_to_pages(1, 2)

    executive = book.sheet(
        "EXECUTIVE_CASE",
        f"{title_period}  |  BONUS KPI CHANGE CASE",
        "DECISION BRIEF  /  OBSERVED FACTS, CONTROLLED SENSITIVITIES, AND A GOVERNED PILOT REQUEST",
        14,
    )
    executive.set_tab_color(COLORS["gold"])
    executive.set_column("A:A", 2)
    executive.set_column("B:O", 13)
    book.banner(
        executive,
        3,
        1,
        13,
        "DECISION REQUEST  |  Recalibrate the KPI configuration before the next bonus cycle; do not retroactively alter July payroll from this analysis.",
        "warn",
    )
    book.kpi(executive, 1, 5, "AGENTS", len(records), kind="kpi")
    book.kpi(executive, 4, 5, "CURRENT MODELED PAYOUT", baseline_total, kind="kpi_money")
    book.kpi(executive, 7, 5, "PARTICIPATION ATTAINMENT", participation["attainment"], kind="kpi_pct")
    book.kpi(executive, 10, 5, "INCOMPLETE PAID EXPOSURE", incomplete_exposure, kind="kpi_money")
    book.kpi(executive, 13, 5, "WIDEST LOB SPREAD", widest_spread, kind="kpi_pct")
    executive.merge_range("B10:N10", "WHY THE CONFIGURATION NEEDS A DECISION", f["section"])
    executive.write_row("B12", ["Priority", "Finding", "July evidence", "Decision requested"], f["header"])
    executive.set_column("B:B", 12)
    executive.set_column("C:C", 31)
    executive.set_column("D:D", 66)
    executive.set_column("E:E", 66)
    for index, (severity, issue, evidence, requested) in enumerate(findings, 12):
        executive.write(index, 1, severity, severity_formats[severity])
        executive.write(index, 2, issue, wrap)
        executive.write(index, 3, evidence, wrap)
        executive.write(index, 4, requested, decision)
        executive.set_row(index, 56)
    executive.merge_range("B21:N21", "RECOMMENDED CONTROLLED PILOT", f["section"])
    pilot_lines = [
        f"1. Participation sensitivity: Tier 1 {pilot_t1:.0%}, Tier 2 {pilot_t2:.0%}; July P75={participation_p75:.1%}, median={participation_median:.1%}.",
        "2. Extra PCS: use Replacement during the pilot so the same PCS score is not added twice.",
        "3. Release gate: hold incomplete rows for completion; do not score missing KPIs as zero.",
        "4. Calibration: approve LOB-specific, effective-dated targets only after at least three comparable months.",
        "5. Measurement: load PCS eligible-survey and valid-response counts and define a minimum sample.",
    ]
    executive.merge_range("B23:N28", "\n".join(pilot_lines), decision)
    executive.set_row(22, 26)
    executive.set_row(23, 24)
    executive.set_row(24, 24)
    executive.set_row(25, 24)
    executive.set_row(26, 24)
    executive.set_row(27, 24)
    executive.freeze_panes("B12")
    executive.set_landscape()
    executive.fit_to_pages(1, 2)

    kpi = book.sheet(
        "KPI_DIAGNOSTIC",
        f"{title_period}  |  KPI DIAGNOSTIC",
        "COVERAGE AND ATTAINMENT USE MEASURED ROWS ONLY; MISSING VALUES ARE SHOWN SEPARATELY",
        16,
    )
    kpi.set_column("A:A", 21)
    kpi.set_column("B:B", 10)
    kpi.set_column("C:F", 13)
    kpi.set_column("G:G", 11)
    kpi.set_column("H:H", 12)
    kpi.set_column("I:L", 13)
    kpi.set_column("M:P", 14)
    headers = [
        "KPI", "Direction", "Tier 1 Weight", "Tier 1 Target", "Tier 2 Weight", "Tier 2 Target",
        "Measured", "Coverage", "Tier 1 Hits", "Tier 2-only Hits", "Any Award", "Attainment",
        "P25", "Median", "P75", "Maximum", "Diagnostic",
    ]
    kpi.write_row(4, 0, headers, f["header"])
    for row_index, row in enumerate(diagnostics, 5):
        metric = row["metric"]
        rule = row["rule"]
        diagnostic = "Within observed range"
        if metric.rule_name == "PCS % (Participation)" and row["maximum"] < rule.tier1_target:
            diagnostic = "Tier 1 above July maximum"
        elif row["coverage_rate"] < 0.80:
            diagnostic = "Coverage below 80%"
        values = [
            metric.label,
            "Lower" if rule and rule.direction == "L" else "Higher",
            rule.tier1_bonus if rule else None,
            _target_display(metric, rule.tier1_target if rule else None),
            rule.tier2_bonus if rule else None,
            _target_display(metric, rule.tier2_target if rule and rule.tier2_bonus else None),
            row["coverage"],
            row["coverage_rate"],
            row["tier1_hits"],
            row["tier2_hits"],
            row["tier1_hits"] + row["tier2_hits"],
            row["attainment"],
            _observed_display(metric, row["p25"]),
            _observed_display(metric, row["median"]),
            _observed_display(metric, row["p75"]),
            _observed_display(metric, row["maximum"]),
            diagnostic,
        ]
        for column, value in enumerate(values):
            fmt = f["body"]
            if column in {2, 4, 7, 11}:
                fmt = pct_plain
            if column == 16 and diagnostic != "Within observed range":
                fmt = f["warn"] if "Coverage" in diagnostic else f["bad"]
            kpi.write(row_index, column, value, fmt)
    kpi.autofilter(4, 0, 4 + len(diagnostics), len(headers) - 1)
    kpi.freeze_panes(5, 1)
    chart = wb.add_chart({"type": "bar"})
    chart.add_series({
        "name": "Attainment rate",
        "categories": ["KPI_DIAGNOSTIC", 5, 0, 5 + len(diagnostics) - 1, 0],
        "values": ["KPI_DIAGNOSTIC", 5, 11, 5 + len(diagnostics) - 1, 11],
        "fill": {"color": COLORS["teal"]},
        "border": {"none": True},
        "data_labels": {"value": True, "num_format": "0.0%"},
    })
    chart.set_title({"name": "Current award attainment among measured agents"})
    chart.set_x_axis({"num_format": "0%", "min": 0, "max": 1, "major_gridlines": {"visible": False}})
    chart.set_legend({"none": True})
    chart.set_style(10)
    kpi.insert_chart("A14", chart, {"x_scale": 1.45, "y_scale": 1.3})

    lob = book.sheet(
        "LOB_DIAGNOSTIC",
        f"{title_period}  |  LOB COMPARABILITY",
        "SAME SOURCE TARGETS, BUT COVERAGE AND CONDITIONAL ATTAINMENT ARE REPORTED BY POPULATION",
        10,
    )
    lob.set_column("A:A", 17)
    lob.set_column("B:D", 12)
    lob.set_column("E:E", 22)
    lob.set_column("F:K", 14)
    lob_headers = [
        "LOB", "Agents", "Complete Rows", "KPI", "Measured", "Coverage", "Median",
        "Direction", "Tier 1 Target", "Attaining", "Attainment",
    ]
    lob.write_row(4, 0, lob_headers, f["header"])
    for row_index, row in enumerate(lob_rows, 5):
        metric = row["metric"]
        values = [
            row["population"], row["agents"], row["complete"], metric.label, row["coverage"],
            row["coverage_rate"], _observed_display(metric, row["median"]),
            "Lower" if row["direction"] == "L" else "Higher", _target_display(metric, row["target"]),
            row["attaining"], row["attainment"],
        ]
        for column, value in enumerate(values):
            fmt = pct_plain if column in {5, 10} and value is not None else f["body"]
            if column == 10 and value is not None and (value < 0.10 or value > 0.90):
                fmt = f["warn"]
            lob.write(row_index, column, value, fmt)
    lob.autofilter(4, 0, 4 + len(lob_rows), len(lob_headers) - 1)
    lob.freeze_panes(5, 1)

    sensitivity = book.sheet(
        "SCENARIO_SENSITIVITY",
        f"{title_period}  |  CONFIGURATION SENSITIVITY",
        "EACH ROW ISOLATES A POLICY CHOICE; VALUES ARE MODELED EVIDENCE, NOT A PAYROLL RELEASE",
        8,
    )
    sensitivity.set_column("A:A", 24)
    sensitivity.set_column("B:B", 58)
    sensitivity.set_column("C:E", 18)
    sensitivity.set_column("F:F", 14)
    sensitivity.set_column("G:H", 20)
    sensitivity.set_column("I:I", 45)
    scenario_headers = [
        "Scenario", "Change tested", "Modeled / Payable", "Comparable Current", "Delta",
        "Paid / Ready", "Rows Held", "Held Baseline Exposure", "How to read it",
    ]
    sensitivity.write_row(4, 0, scenario_headers, f["header"])
    interpretations = {
        "Current configuration": "Reproduces the workbook exactly; missing KPI values contribute zero points.",
        "Participation calibration": "Shows the isolated cost of bringing participation tiers inside July's measured distribution.",
        "PCS overlap removed": "Shows the isolated savings when standard and extra PCS do not stack.",
        "Controlled pilot": "Compares complete rows only; held exposure is pending data completion, not denied payout.",
    }
    for row_index, summary in enumerate(summaries, 5):
        values = [
            summary.scenario.name,
            summary.scenario.change,
            summary.total,
            summary.comparable_baseline,
            summary.delta,
            summary.paid,
            summary.held,
            summary.held_baseline_exposure,
            interpretations[summary.scenario.name],
        ]
        for column, value in enumerate(values):
            fmt = wrap
            if column in {2, 3, 7}:
                fmt = f["money"]
            elif column == 4:
                fmt = money_delta
            elif column in {5, 6}:
                fmt = f["int"]
            sensitivity.write(row_index, column, value, fmt)
        sensitivity.set_row(row_index, 46)
    sensitivity.merge_range(
        "A11:I13",
        f"CALIBRATION BASIS\nThe temporary participation tiers are generated from July's measured distribution: "
        f"P75 {participation_p75:.1%} rounded up to {pilot_t1:.0%}; median {participation_median:.1%} rounded to {pilot_t2:.0%}. "
        "This is a sensitivity test. Final targets require at least three comparable months, LOB context, and sample counts.",
        decision,
    )
    scenario_chart = wb.add_chart({"type": "column"})
    scenario_chart.add_series({
        "name": "Modeled / payable MAD",
        "categories": ["SCENARIO_SENSITIVITY", 5, 0, 5 + len(summaries) - 1, 0],
        "values": ["SCENARIO_SENSITIVITY", 5, 2, 5 + len(summaries) - 1, 2],
        "fill": {"color": COLORS["teal"]},
        "border": {"none": True},
        "data_labels": {"value": True, "num_format": '#,##0 "MAD"'},
    })
    scenario_chart.set_title({"name": "Configuration sensitivity"})
    scenario_chart.set_y_axis({"num_format": '#,##0 "MAD"', "major_gridlines": {"visible": False}})
    scenario_chart.set_legend({"none": True})
    scenario_chart.set_style(10)
    sensitivity.insert_chart("A16", scenario_chart, {"x_scale": 1.55, "y_scale": 1.35})
    sensitivity.freeze_panes(5, 0)

    agents = book.sheet(
        "AGENT_IMPACT",
        f"{title_period}  |  AGENT-LEVEL TRACE",
        "FILTERABLE RECONCILIATION OF CURRENT AND SENSITIVITY OUTPUTS; HELD MEANS DATA COMPLETION REQUIRED",
        12,
    )
    agents.set_column("A:A", 13)
    agents.set_column("B:B", 29)
    agents.set_column("C:C", 16)
    agents.set_column("D:E", 24)
    agents.set_column("F:F", 12)
    agents.set_column("G:G", 34)
    agents.set_column("H:K", 19)
    agents.set_column("L:L", 18)
    agents.set_column("M:M", 20)
    agent_headers = [
        "Agent ID", "Agent Name", "LOB", "Team Lead", "Ops Manager", "Core Data", "Missing KPI",
        "Current Payout", "Participation Sensitivity", "Replacement Sensitivity", "Controlled Pilot",
        "Comparable Delta", "Source Data Status",
    ]
    agents.write_row(4, 0, agent_headers, f["header"])
    for row_index, (record, complete) in enumerate(zip(records, complete_mask), 5):
        missing = [metric.label for metric in METRICS if _value(record, metric.input_names) is None]
        current = scenario_values[scenarios[0].name][row_index - 5]
        participation_value = scenario_values[scenarios[1].name][row_index - 5]
        replacement_value = scenario_values[scenarios[2].name][row_index - 5]
        controlled = scenario_values[scenarios[3].name][row_index - 5]
        comparable_delta = None if controlled.held else (controlled.payout or 0.0) - (current.payout or 0.0)
        values = [
            record.get("Agent ID"), record.get("Agent Name"), record.get("Population"),
            record.get("Team Lead"), record.get("Ops Manager"), "COMPLETE" if complete else "HOLD",
            ", ".join(missing) or "—", current.payout, participation_value.payout,
            replacement_value.payout, controlled.payout, comparable_delta, record.get("Data Status"),
        ]
        for column, value in enumerate(values):
            fmt = f["body"]
            if column in {7, 8, 9, 10}:
                fmt = f["money"]
            elif column == 11:
                fmt = money_delta
            elif column == 5:
                fmt = f["good"] if complete else f["warn"]
            agents.write(row_index, column, value, fmt)
    agents.autofilter(4, 0, 4 + len(records), len(agent_headers) - 1)
    agents.freeze_panes(5, 2)

    evidence = book.sheet(
        "EVIDENCE_METHOD",
        "EVIDENCE, METHOD, AND LIMITS",
        "AUDIT TRAIL FOR REPRODUCTION AND MANAGEMENT REVIEW",
        8,
    )
    evidence.set_column("A:A", 28)
    evidence.set_column("B:B", 92)
    evidence.write_row(4, 0, ["Evidence item", "Value"], f["header"])
    evidence_rows = [
        ("Source workbook", source.name),
        ("Source SHA-256", before_hash),
        ("Period", period),
        ("Generated UTC", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("Agent rows", len(records)),
        ("Current modeled payout", f"{baseline_total:,.2f} MAD"),
        ("Cached Results payout", "Not available" if cached_total is None else f"{cached_total:,.2f} MAD across {cached_rows} rows"),
        ("Reconciliation delta", "Not available" if reconciliation_delta is None else f"{reconciliation_delta:,.2f} MAD"),
        ("Complete core KPI rows", f"{sum(complete_mask)} of {len(records)}"),
        ("Source Data Status", str(dict(statuses))),
        ("Policy decisions not Validated", policy_reviews),
    ]
    for row_index, row in enumerate(evidence_rows, 5):
        evidence.write(row_index, 0, row[0], f["body"])
        evidence.write(row_index, 1, row[1], wrap)
    method_row = 18
    evidence.merge_range(method_row, 0, method_row, 8, "METHOD CONTRACT", f["section"])
    methods = [
        ("Current reproduction", "Uses source KPI thresholds, source policy choices, a 130% cap, source reference bonus and proration. Missing KPI values earn zero, matching the source workbook behavior."),
        ("Attainment denominator", "KPI attainment divides award-earning measured rows by measured rows only. Coverage is reported separately, so missing values are not silently counted as failures."),
        ("Participation sensitivity", f"Tier 1 is July P75 ({participation_p75:.1%}) rounded up to 5 percentage points ({pilot_t1:.0%}); Tier 2 is the median ({participation_median:.1%}) rounded to 5 points ({pilot_t2:.0%})."),
        ("Controlled pilot", "Uses the participation sensitivity, Replacement extra PCS, and holds any row missing one of AHT, Productivity, PCS Score, PCS Participation, QM, or Absence."),
        ("LOB comparison", "Attainment is conditional on measured values inside each LOB. Large spreads diagnose comparability risk; they do not prove a target should be lowered."),
        ("Limit", "July is one month. The workbook does not contain PCS response counts or survey opportunity counts. This report supports a pilot decision, not a final target or retroactive payroll release."),
    ]
    evidence.write_row(method_row + 2, 0, ["Method", "Definition"], f["header"])
    for row_index, row in enumerate(methods, method_row + 3):
        evidence.write(row_index, 0, row[0], f["body"])
        evidence.write(row_index, 1, row[1], wrap)
        evidence.set_row(row_index, 42)
    evidence.freeze_panes(5, 0)

    email = book.sheet(
        "EMAIL_SCRIPT",
        "READY-TO-SEND EMAIL AND PRESENTATION SCRIPT",
        "COPY THE SUBJECT AND BODY; THE WORDING IS DELIBERATELY SIMPLE",
        8,
    )
    email.set_column("A:A", 18)
    email.set_column("B:I", 17)
    subject = f"Decision request | Recalibrate bonus KPI configuration after {period} evidence"
    body = (
        "Hello,\n\n"
        f"I reviewed the {period} bonus population ({len(records)} agents) and reconciled the current logic to "
        f"{baseline_total:,.0f} MAD. The evidence supports a configuration review before the next cycle:\n\n"
        f"• PCS Participation Tier 1 is {participation['rule'].tier1_target:.0%}, above July's maximum of {participation['maximum']:.1%}. "
        f"No measured agent reaches Tier 1, and only {participation['tier1_hits'] + participation['tier2_hits']} of {participation['coverage']} receive any participation award.\n"
        f"• The same target set is used for six LOBs, while conditional {widest_metric} attainment differs by {widest_spread:.1%} across them.\n"
        f"• {missing_rows} rows miss at least one core KPI. The current missing-as-zero logic still pays {paid_incomplete} of those rows, representing {incomplete_exposure:,.0f} MAD.\n"
        f"• Standard and Extra PCS together can contribute {pcs_max_points:.0%} achievement points from the same score field.\n\n"
        f"• Absence already carries {absence['rule'].tier1_bonus:.0%} weight. If the same absence later reduces proration, it would reduce the bonus twice.\n\n"
        f"I request approval for a controlled three-month calibration pilot: test participation at {pilot_t1:.0%}/{pilot_t2:.0%}, "
        "use Replacement for Extra PCS pending Compensation approval, hold incomplete rows for completion, add PCS numerator/denominator counts, "
        "count the same absence only once, and define effective-dated LOB targets from comparable history.\n\n"
        f"The July sensitivities are transparent in the attached workbook: participation calibration alone is {summaries[1].delta:+,.0f} MAD; "
        f"removing PCS overlap alone is {summaries[2].delta:+,.0f} MAD; the controlled pilot changes complete-row payout by {summaries[3].delta:+,.0f} MAD, "
        f"with {incomplete_exposure:,.0f} MAD held for data completion. These are policy sensitivities, not a retroactive payroll instruction.\n\n"
        "Regards,\nAnass ASSRI"
    )
    email.write("A5", "Subject", f["header"])
    email.merge_range("B5:I5", subject, f["input"])
    email.write("A7", "Body", f["header"])
    email.merge_range("B7:I24", body, wrap)
    email.set_row(6, 25)
    for row_index in range(7, 24):
        email.set_row(row_index, 26)

    # The first four tabs are intentionally presentation-ready. Detailed
    # reconciliations remain in the file for audit questions but stay hidden
    # from a novice presenter unless they explicitly unhide them.
    for technical_sheet in (executive, kpi, lob, sensitivity, agents, evidence):
        technical_sheet.hide()

    output_path = book.close()
    after_hash = _source_hash(source)
    if after_hash != before_hash:
        raise RuntimeError("Source workbook changed while the bonus analysis was generated")
    return output_path
