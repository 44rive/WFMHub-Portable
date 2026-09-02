from __future__ import annotations

import tempfile
import unittest
from datetime import date
from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook

from wfmhub.bonus_analysis_report import build_bonus_kpi_change_case
from wfmhub.shared_reports import build_bonus_management, build_pcs_management


def make_bonus_source(path: Path) -> None:
    workbook = Workbook()
    raw = workbook.active
    raw.title = "Raw_Data"
    raw.append(["Agent ID", "Agent Name", "Population", "Period", "AHT", "Productivity",
                "PCS Score", "PCS % Participation", "QM", "Abs%", "VOC Detractor Count",
                "Currency", "Monthly Fixed Salary", "Target Bonus Rate", "Reference Bonus Override",
                "Eligible Days", "Scheduled Days", "Employment Status", "Data Status"])
    raw.append(["100", "Agent 100", "OEM", "2026-08", 440, 8, 4.5, .55, .95, .02, 1,
                "MAD", None, None, 2000, 31, 31, "Active", "SYNTHETIC"])
    config = workbook.create_sheet("KPI_Config")
    config.append(["Population", "KPI", "Direction", "Tier 1 Bonus %", "Tier 1 Target",
                   "Tier 2 Bonus %", "Tier 2 Target"])
    for values in (
        ("OEM", "AHT", "L", .15, 450, 0, 0),
        ("OEM", "Productivity", "H", .15, 7, 0, 0),
        ("OEM", "PCS Score", "H", .20, 4.3, .15, 4.1),
        ("OEM", "PCS % (Participation)", "H", .15, .5, .1, .4),
        ("OEM", "QM", "H", .2, .9, .15, .8),
        ("OEM", "Abs%", "L", .15, .05, 0, 0),
        ("OEM", "Extra Bonus (PCS Score)", "H", .3, 4.5, .2, 4.3),
    ):
        config.append(values)
    workbook.save(path)


def make_pcs_source(path: Path) -> None:
    workbook = Workbook()
    roster = workbook.active
    roster.title = "AGENT LIST"
    roster.append(["Client ID", "Status", "Name", "HR Start date", "Prod date", "Prod Month",
                   "Team leader", "Ops Manager", "LOB"])
    roster.append([100, "Active", "Agent 100", None, None, None, "TL 1", "Ops 1", "RSA"])
    overview = workbook.create_sheet("OverView")
    overview["B9"] = date(2026, 8, 31)
    workbook.save(path)


class SharedReportTests(unittest.TestCase):
    def test_bonus_change_case_reconciles_and_keeps_scenarios_explicit(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            source = root / "bonus.xlsx"
            output = root / "Bonus_KPI_Change_Case.xlsx"
            make_bonus_source(source)
            source_bytes = source.read_bytes()

            build_bonus_kpi_change_case(source, output)

            self.assertEqual(source.read_bytes(), source_bytes)
            workbook = load_workbook(output, data_only=True, keep_links=False)
            self.assertEqual(workbook.sheetnames[0], "START_HERE")
            self.assertEqual(workbook.sheetnames[1], "WHAT_TO_CHANGE")
            self.assertEqual(workbook.sheetnames[2], "PROOF")
            self.assertIn("SCENARIO_SENSITIVITY", workbook.sheetnames)
            self.assertIn("AGENT_IMPACT", workbook.sheetnames)
            self.assertIn("EMAIL_SCRIPT", workbook.sheetnames)
            self.assertEqual(workbook["SCENARIO_SENSITIVITY"]["A6"].value, "Current configuration")
            self.assertEqual(workbook["AGENT_IMPACT"]["F6"].value, "COMPLETE")
            self.assertIn("not a retroactive payroll instruction", workbook["EMAIL_SCRIPT"]["B7"].value)
            self.assertIn("count the same absence only once", workbook["EMAIL_SCRIPT"]["B7"].value)
            self.assertEqual(workbook["EXECUTIVE_CASE"].sheet_state, "hidden")
            self.assertEqual(workbook["KPI_DIAGNOSTIC"].sheet_state, "hidden")
            workbook.close()
            with ZipFile(output) as archive:
                names = archive.namelist()
                self.assertFalse(any("externalLink" in name for name in names))
                self.assertFalse(any("connections" in name for name in names))

    def test_bonus_management_is_gated_and_has_no_external_links(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            source = root / "bonus.xlsx"
            output = root / "Bonus_Management_Proposal.xlsx"
            make_bonus_source(source)
            build_bonus_management(source, output)

            workbook = load_workbook(output, data_only=False, keep_links=False)
            self.assertEqual(workbook.sheetnames[0], "SETUP")
            self.assertIn("MANAGEMENT", workbook.sheetnames)
            self.assertIn("FORMULA_LOGIC", workbook.sheetnames)
            self.assertEqual(workbook["MANAGEMENT"].freeze_panes, "B11")
            self.assertEqual(len(workbook["MANAGEMENT"]._charts), 3)
            self.assertIn("LOB DECOMPOSITION", workbook["MANAGEMENT"]["B11"].value)
            self.assertIn("RELEASE GOVERNANCE", workbook["MANAGEMENT"]["B54"].value)
            self.assertIn("SETUP!$B$9", workbook["AGENT_RESULTS"]["S5"].value)
            self.assertIn("VALIDATED", workbook["AGENT_RESULTS"]["T5"].value)
            workbook.close()
            cached = load_workbook(output, data_only=True, read_only=True)
            self.assertIsNone(cached["AGENT_RESULTS"]["S5"].value)
            cached.close()
            with ZipFile(output) as archive:
                names = archive.namelist()
                self.assertFalse(any("externalLink" in name for name in names))
                self.assertFalse(any("connections" in name for name in names))
                self.assertNotIn(b"#REF!", b"".join(archive.read(name) for name in names if name.endswith(".xml")))

    def test_pcs_template_preserves_original_participation_and_three_hour_window(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            source = root / "PCS Report.xlsx"
            output = root / "PCS_Management_3H_Template.xlsx"
            make_pcs_source(source)
            build_pcs_management(source, output, date(2026, 8, 31))

            workbook = load_workbook(output, data_only=False, keep_links=False)
            self.assertIn("PCS_REPORT", workbook.sheetnames)
            self.assertIn("INPUT_CALLS", workbook.sheetnames)
            participation = workbook["AGENT_DETAIL"]["L5"].value
            three_hour = workbook["AGENT_DETAIL"]["V5"].value
            self.assertIn("G5/F5", participation)
            self.assertIn("Q5/P5", three_hour)
            self.assertIn("SETUP!$B$6", workbook["AGENT_DETAIL"]["P5"].value)
            logic = " ".join(str(workbook["FORMULA_LOGIC"].cell(row, 2).value or "")
                             for row in range(1, workbook["FORMULA_LOGIC"].max_row + 1))
            self.assertIn("nonblank", logic)
            self.assertIn("external", logic)
            workbook.close()
            with ZipFile(output) as archive:
                names = archive.namelist()
                self.assertFalse(any("externalLink" in name for name in names))
                self.assertNotIn(b"#REF!", b"".join(archive.read(name) for name in names if name.endswith(".xml")))


if __name__ == "__main__":
    unittest.main()
