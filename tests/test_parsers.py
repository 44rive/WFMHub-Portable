from __future__ import annotations

import csv
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from wfmhub.ingestion import AgentScope, SourceSchemaError, parse_agent_status, parse_calls, parse_forecast, parse_fte, parse_lilo, parse_queue_actual, parse_schedule
from tools.build_fte_template import standardize_source


class ParserTests(unittest.TestCase):
    def test_leaver_cutoff_is_applied_inside_schedule_status_and_calls(self):
        scope = AgentScope(
            frozenset({"200"}),
            {"leaver agent": "200"},
            "scope",
            {"200": ("Leaver", date(2026, 8, 1))},
        )
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            schedule = root / "StartEndTimes.txt"
            with schedule.open("w", encoding="cp1252", newline="") as handle:
                writer = csv.writer(handle, delimiter="\t")
                writer.writerow(["Name", "Data Source IDs", "08/01/2026", "08/02/2026"])
                writer.writerow([
                    "Leaver Agent", "200",
                    ".ORG | Work 08/01/2026 8:00 AM-08/01/2026 4:00 PM",
                    ".ORG | Work 08/02/2026 8:00 AM-08/02/2026 4:00 PM",
                ])
            schedule_result = parse_schedule(schedule, "schedule", scope)
            self.assertEqual(
                [row["schedule_date"] for row in schedule_result.tables["raw.schedule_shift"]],
                [date(2026, 8, 1)],
            )
            self.assertEqual(schedule_result.scoped_out, 1)

            status = root / "Agent Status.csv"
            status.write_text(
                "[Serial Number],[Status],[Status Start Date and Time],[Agent],[Agent ID],[Status Duration],[Queue]\n"
                "one,Available,8/1/2026 9:00,Leaver Agent,200,0:15:00,Main\n"
                "two,Available,8/2/2026 9:00,Leaver Agent,200,0:15:00,Main\n",
                encoding="utf-8-sig",
            )
            status_result = parse_agent_status(status, "status", scope)
            self.assertEqual(
                [row["extract_date"] for row in status_result.tables["raw.agent_status"]],
                [date(2026, 8, 1)],
            )
            self.assertEqual(status_result.scoped_out, 1)

            calls = root / "Call by Call.csv"
            with calls.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow([
                    "[Call Date/Time]", "[Call End Date/Time]", "[Call ID]",
                    "[Call Reference Number]", "[Agent ID]", "[Agent]",
                    "[Call Direction]", "[Talk Time]", "[Hold Time]",
                    "[Total Wrap Time]",
                ])
                writer.writerow(["8/1/2026 9:00", "8/1/2026 9:05", "one", "one", "200", "Leaver Agent", "I", "0:04:00", "0:00:30", "0:00:30"])
                writer.writerow(["8/2/2026 9:00", "8/2/2026 9:05", "two", "two", "200", "Leaver Agent", "I", "0:04:00", "0:00:30", "0:00:30"])
            calls_result = parse_calls(calls, "calls", scope)
            self.assertEqual(
                [row["business_date"] for row in calls_result.tables["raw.call_leg"]],
                [date(2026, 8, 1)],
            )
            self.assertEqual(calls_result.scoped_out, 1)

    def test_shipped_fte_template_has_stable_contract(self):
        template = Path(__file__).resolve().parents[1] / "templates" / "FTE Count.xlsx"
        workbook = load_workbook(template, read_only=False, data_only=False)
        try:
            self.assertEqual(workbook.sheetnames, ["START_HERE", "Agent", "PTO", "Away"])
            headers = [cell.value for cell in workbook["Agent"][1]]
            self.assertEqual(headers, [
                "Client ID", "Status", "Name", "Team leader", "Ops Manager", "LOB",
                "Market", "Language", "Location", "City", "FTE", "End date if leaver",
            ])
            self.assertIn("tblFTEAgents", workbook["Agent"].tables)
            self.assertEqual([cell.value for cell in workbook["PTO"][1]], [
                "Client ID", "Name", "Start date", "End date", "Day coverage",
                "Start time", "End time", "PTO type", "Approval status", "Comment",
            ])
            self.assertIn("tblFTEPTO", workbook["PTO"].tables)
            self.assertEqual([cell.value for cell in workbook["Away"][1]], [
                "Client ID", "Name", "Start date", "End date", "Away type",
                "Case status", "Comment",
            ])
            self.assertIn("tblFTEAway", workbook["Away"].tables)
        finally:
            workbook.close()

    def test_fte_parses_governed_pto_and_away_registers(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "FTE Count.xlsx"
            workbook = Workbook()
            agent = workbook.active
            agent.title = "Agent"
            agent.append(["Client ID", "Status", "Name"])
            agent.append(["00123", "Active", "Jane Agent"])
            pto = workbook.create_sheet("PTO")
            pto.append([
                "Client ID", "Name", "Start date", "End date", "Day coverage",
                "Start time", "End time", "PTO type", "Approval status", "Comment",
            ])
            pto.append(["00123", "Jane Agent", date(2026, 8, 1), date(2026, 8, 1), "Partial day", "08:00", "10:00", "Vacation", "Approved", "OK"])
            pto.append(["00123", "Jane Agent", date(2026, 8, 2), date(2026, 8, 2), "Partial day", None, None, "Vacation", "Approved", "bad"])
            away = workbook.create_sheet("Away")
            away.append(["Client ID", "Name", "Start date", "End date", "Away type", "Case status", "Comment"])
            away.append(["00123", "Jane Agent", date(2026, 8, 3), None, "Long sickness", "Active", None])
            workbook.save(path)

            result = parse_fte(path, "file")
            rows = result.tables["raw.fte_time_off"]
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[0]["agent_id"], "00123")
            self.assertEqual(rows[0]["day_coverage"], "PARTIAL_DAY")
            self.assertEqual(rows[1]["source_kind"], "AWAY")
            self.assertIsNone(rows[1]["end_date"])
            self.assertEqual(len(result.rejected), 1)
            self.assertIn("Partial day PTO requires", result.rejected[0])

            standardized = Path(folder) / "Standardized FTE Count.xlsx"
            standardize_source(path, standardized)
            preserved = parse_fte(standardized, "standardized")
            self.assertEqual(len(preserved.tables["raw.fte_time_off"]), 2)
            self.assertEqual(
                {row["source_kind"] for row in preserved.tables["raw.fte_time_off"]},
                {"PTO", "AWAY"},
            )

    def test_fte_finds_renamed_sheet_offset_header_and_aliases(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "FTE Count.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "August Roster"
            sheet.append(["FTE population report"])
            sheet.append([])
            sheet.append(["Agent ID", "Agent Name", "Employment Status", "TL", "Line of Business", "Site"])
            sheet.append(["00123", "Jane Agent", "Active", "Team Lead", "Quality", "Berlin"])
            workbook.save(path)

            row = parse_fte(path, "file").tables["raw.fte_agent"][0]
            self.assertEqual(row["source_row"], 4)
            self.assertEqual(row["agent_id"], "00123")
            self.assertEqual(row["agent_name"], "Jane Agent")
            self.assertEqual(row["team_leader"], "Team Lead")
            self.assertEqual(row["lob"], "Quality")
            self.assertEqual(row["location"], "Berlin")
            self.assertIsNone(row["ops_manager"])

    def test_fte_error_names_searched_sheets_and_required_headers(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "FTE Count.xlsx"
            workbook = Workbook()
            workbook.active.title = "People"
            workbook.active.append(["Something", "Else"])
            workbook.save(path)
            with self.assertRaisesRegex(SourceSchemaError, "People.*Client ID/Agent ID"):
                parse_fte(path, "file")

    def test_fte_does_not_select_support_sheet_before_agent_roster(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "FTE Count.xlsx"
            workbook = Workbook()
            support = workbook.active
            support.title = "Support"
            support.append(["Client ID", "Name", "Status", "LOB", "Market", "Location", "End Date"])
            support.append(["SUP-1", "Support Person", "Active", "Support", "Market", "Site", None])
            agent = workbook.create_sheet("Agent")
            agent.append(["Client ID", "Name"])
            agent.append(["AGT-1", "Roster Person"])
            workbook.save(path)

            rows = parse_fte(path, "file").tables["raw.fte_agent"]
            self.assertEqual([row["agent_id"] for row in rows], ["AGT-1"])

    def test_fte_rejects_multiple_id_aliases_in_roster_header(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "FTE Count.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Agent"
            sheet.append(["Client ID", "Agent ID", "Name"])
            sheet.append(["CLIENT-1", "VERINT-1", "Person"])
            workbook.save(path)
            with self.assertRaisesRegex(SourceSchemaError, "multiple aliases.*agent_id"):
                parse_fte(path, "file")

    def test_fte_rejects_two_authoritative_roster_sheets(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "FTE Count.xlsx"
            workbook = Workbook()
            agent = workbook.active
            agent.title = "Agent"
            agent.append(["Client ID", "Name"])
            agent.append(["A-1", "First Person"])
            roster = workbook.create_sheet("Roster")
            roster.append(["Client ID", "Name"])
            roster.append(["A-2", "Second Person"])
            workbook.save(path)
            with self.assertRaisesRegex(SourceSchemaError, "multiple equally likely agent tables"):
                parse_fte(path, "file")

    def test_schedule_uses_quote_aware_tsv(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "schedule.txt"
            with path.open("w", encoding="cp1252", newline="") as handle:
                writer = csv.writer(handle, delimiter="\t")
                writer.writerow(["Name", "Data Source IDs", "Scheduling Period", "Before Overtime", "After Overtime", "Shift Assignment", "Shift Events"])
                writer.writerow(["08/01/2026", "", "", "", "", "", ""])
                writer.writerow(["Doe,\tJane", "123456", "August", "", "", "Off", ""])
            result = parse_schedule(path, "file")
            self.assertEqual(result.source_variant, "ACTIVITIES")
            self.assertEqual(len(result.tables["raw.schedule_shift"]), 1)
            self.assertEqual(result.tables["raw.schedule_shift"][0]["agent_id"], "123456")
            self.assertEqual(result.tables["raw.schedule_shift"][0]["agent_name"], "Doe,\tJane")

    def test_start_end_times_wide_extract_normalizes_every_date_column(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "VERINT_01082026_03082026_StartEndTimes.txt"
            with path.open("w", encoding="cp1252", newline="") as handle:
                writer = csv.writer(handle, delimiter="\t")
                writer.writerow(["Name", "Data Source IDs", "08/01/2026", "08/02/2026", "08/03/2026"])
                writer.writerow([
                    "Agent One", "123", ".AP | Short Sickness 08/01/2026 8:30 AM-08/01/2026 6:00 PM",
                    "Off", ".AP | Front-office 08/03/2026 9:00 AM-08/03/2026 6:30 PM",
                ])
            result = parse_schedule(path, "file")
            self.assertEqual(result.source_variant, "START_END")
            shifts = result.tables["raw.schedule_shift"]
            self.assertEqual(len(shifts), 3)
            self.assertEqual([str(row["schedule_date"]) for row in shifts], ["2026-08-01", "2026-08-02", "2026-08-03"])
            self.assertEqual(shifts[0]["agent_id"], "123")
            self.assertEqual(shifts[0]["assignment_type"], "Planned absence")
            self.assertEqual(shifts[1]["assignment_type"], "Off")
            self.assertEqual(result.tables["raw.schedule_event"], [])

    def test_schedule_rejects_an_unrecognized_tsv_in_schedule_folder(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "not-a-schedule.txt"
            path.write_text("Queue\tDate\tValue\nA\t2026-08-01\t1\n", encoding="cp1252")
            with self.assertRaisesRegex(SourceSchemaError, "neither a StartEndTimes export nor an Activities export"):
                parse_schedule(path, "file")

    def test_lilo_preserves_and_adjusts_overnight_boundary(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "AP-Historical-Report---Agent-Login 2026-08-01.csv"
            path.write_text(
                "[Agent],[Agent ID],[First Log-on Time],[Last Log-off Time]\n"
                "Jane,123,2026-08-01 21:55:00,2026-08-01 07:00:00\n",
                encoding="utf-8-sig",
            )
            row = parse_lilo(path, "file").tables["raw.lilo"][0]
            self.assertEqual(row["raw_last_logout"], datetime(2026, 8, 1, 7, 0))
            self.assertEqual(row["last_logout"], datetime(2026, 8, 2, 7, 0))
            self.assertTrue(row["overnight_adjusted"])

    def test_lilo_range_file_uses_row_date_and_preserves_blank_no_show(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "AP-Historical-Report---Agent-Login 2026-08-01 - 2026-08-02.csv"
            path.write_text(
                "[Date],[Agent],[Agent ID],[First Log-on Time],[Last Log-off Time]\n"
                "2026-08-01,Jane,123,,,\n"
                "2026-08-02,Jane,123,2026-08-02 08:00:00,2026-08-02 16:00:00\n",
                encoding="utf-8-sig",
            )
            rows = parse_lilo(path, "file").tables["raw.lilo"]
            self.assertEqual([str(row["extract_date"]) for row in rows], ["2026-08-01", "2026-08-02"])
            self.assertIsNone(rows[0]["first_login"])

    def test_lilo_multiday_blank_row_without_date_is_rejected_not_invented(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "AP-Historical-Report---Agent-Login 2026-08-01 - 2026-08-02.csv"
            path.write_text(
                "[Agent],[Agent ID],[First Log-on Time],[Last Log-off Time]\nJane,123,,\n",
                encoding="utf-8-sig",
            )
            result = parse_lilo(path, "file")
            self.assertEqual(result.tables["raw.lilo"], [])
            self.assertRegex(result.rejected[0], "no row date")

    def test_status_duration_can_exceed_24_hours(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "AP-Historical-Report---Agent-Status 2026-08-24.csv"
            path.write_text(
                "[Serial Number],[Status],[Status Start Date and Time],[Agent],[Agent ID],[Status Duration],[Queue]\n"
                "one,Logged Off,8/24/2026 0:00,Jane,123,37:59:56,\n",
                encoding="utf-8-sig",
            )
            row = parse_agent_status(path, "file").tables["raw.agent_status"][0]
            self.assertEqual(str(row["extract_date"]), "2026-08-24")
            self.assertEqual(row["status_end"], datetime(2026, 8, 25, 13, 59, 56))

    def test_status_range_filename_uses_each_rows_timestamp_date(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "AP-Historical-Report---Agent-Status 2026-08-24 - 2026-08-25.csv"
            path.write_text(
                "[Serial Number],[Status],[Status Start Date and Time],[Agent],[Agent ID],[Status Duration],[Queue]\n"
                "one,Available,8/24/2026 23:55,Jane,123,0:05:00,Main\n"
                "two,Available,8/25/2026 0:05,Jane,123,0:05:00,Main\n",
                encoding="utf-8-sig",
            )
            rows = parse_agent_status(path, "file").tables["raw.agent_status"]
            self.assertEqual([str(row["extract_date"]) for row in rows], ["2026-08-24", "2026-08-25"])

    def test_status_filename_does_not_need_a_date_when_rows_have_timestamps(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "Agent Status full history.csv"
            path.write_text(
                "[Serial Number],[Status],[Status Start Date and Time],[Agent],[Agent ID],[Status Duration],[Queue]\n"
                "one,Available,8/24/2026 12:00,Jane,123,0:05:00,Main\n",
                encoding="utf-8-sig",
            )
            row = parse_agent_status(path, "file").tables["raw.agent_status"][0]
            self.assertEqual(str(row["extract_date"]), "2026-08-24")

    def test_forecast_discovers_header_and_scales_service_level(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "forecast.txt"
            path.write_text(
                "DATE_TIME_FORMAT\nMM/DD/YYYY hh:mm A\n"
                "Queue Name\tDate\tTime\tTime Interval\tVolume (Absolute For)\tAbandons (Absolute For)\tService Level (Absolute For)\tService Level (Absolute Req)\tActivity Handling Time (Absolute For)\tHeadcount Staffing (Absolute For)\tNet Staffing (Absolute For)\tFull Time Equivalents (Absolute For)\tFull Time Equivalents (Absolute Req)\n"
                "All\t08/01/2026\t12:00 AM\t1:00\t10\t1\t79\t80\t200\t3\t-1\t2\t3\n",
                encoding="cp1252",
            )
            row = parse_forecast(path, "file").tables["raw.forecast_interval"][0]
            self.assertEqual(row["interval_minutes"], 60)
            self.assertEqual(row["sl_forecast"], 0.79)
            self.assertEqual(row["net_staffing_forecast"], -1)

    def test_forecast_accepts_reviewed_volume_only_extract(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "FORD_FR_08-2026.txt"
            path.write_text(
                "DATE_TIME_FORMAT\nMM/DD/YYYY hh:mm A\n"
                "Queue Name\tDate\tTime\tTime Interval\tVolume (Absolute For)\n"
                "Combined - All Media\t08/01/2026\t12:00 AM\t1:00\t8\n",
                encoding="cp1252",
            )
            row = parse_forecast(path, "file").tables["raw.forecast_interval"][0]
            self.assertEqual(row["volume_forecast"], 8)
            self.assertIsNone(row["fte_forecast"])
            self.assertIsNone(row["sl_forecast"])

    def test_apde_bracketed_csv_is_supported(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "APDE-Standard-KPIs---Inbound-Calls 2026-08-01.csv"
            path.write_text(
                "[Date],[BusinessPartnerID],[LineOfBusiness],[15 Minute Periods of Day],"
                "[Offered_calls (w/o short calls)],[Answered_Calls],[Abandoned_Calls (w/o short calls)],"
                "[Short_calls < 5s],[Answered_Calls <= 15s],[Answered_Calls <= 20s],"
                "[Answered_Calls <= 30s],[Abandoned_Calls (w/o s.c.) <= 20s],"
                "[Average_Speed_of_Answer],[Average_Handled_Time],[Average_Talk_Time],"
                "[Average_Hold_Time],[Average Total Wrap Time]\n"
                "2026/08/01,Ford,FORD,09:15,10,8,2,1,6,7,8,1,0:00:10,0:02:30,0:02:00,0:00:10,0:00:20\n",
                encoding="utf-8-sig",
            )
            row = parse_queue_actual(path, "file", "APDE").tables["raw.queue_actual"][0]
            self.assertEqual(row["language"], "DE")
            self.assertEqual(row["answered_20s"], 7)
            self.assertEqual(row["abandoned_20s"], 1)
            self.assertEqual(row["aht_seconds"], 150)

    def test_call_by_call_uses_row_dates_scores_and_fte_scope(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "Call by Call 2026-08-01 - 2026-08-31.csv"
            headers = [
                "[Call Date/Time]", "[Call End Date/Time]", "[Call ID]",
                "[Call Reference Number]", "[Agent ID]", "[Agent]",
                "[Talk Time]", "[Hold Time]", "[Total Wrap Time]",
                "[Call Direction]", "[PostCallSurveyMode]", "[Question 1]",
                "[Question 2]", "[Question 3]",
            ]
            with path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=headers)
                writer.writeheader()
                writer.writerow({
                    "[Call Date/Time]": "8/1/2026 8:00", "[Call End Date/Time]": "8/1/2026 8:05",
                    "[Call ID]": "call-1", "[Call Reference Number]": "ref-1",
                    "[Agent ID]": "123", "[Agent]": "Jane", "[Talk Time]": "0:04:00",
                    "[Hold Time]": "0:00:30", "[Total Wrap Time]": "0:00:30",
                    "[Call Direction]": "I", "[PostCallSurveyMode]": "2",
                    "[Question 1]": "5", "[Question 2]": "4", "[Question 3]": "Helpful",
                })
                writer.writerow({
                    "[Call Date/Time]": "8/2/2026 9:00", "[Call End Date/Time]": "8/2/2026 9:01",
                    "[Call ID]": "world", "[Call Reference Number]": "world",
                    "[Agent ID]": "999", "[Agent]": "Worldwide", "[Talk Time]": "0:01:00",
                    "[Hold Time]": "0:00:00", "[Total Wrap Time]": "0:00:00",
                })
            scope = AgentScope(frozenset({"123"}), {"jane": "123"}, "scope")
            result = parse_calls(path, "file", scope)
            self.assertEqual(result.scoped_out, 1)
            self.assertEqual(len(result.tables["raw.call_leg"]), 1)
            row = result.tables["raw.call_leg"][0]
            self.assertEqual(str(row["business_date"]), "2026-08-01")
            self.assertEqual(row["question_1_score"], 5)
            self.assertEqual(row["question_2_score"], 4)
            self.assertEqual(row["talk_seconds"], 240)


if __name__ == "__main__":
    unittest.main()
