from __future__ import annotations

import shutil
import tempfile
import unittest
from datetime import date
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

from wfmhub.bonus import import_bonus_matrix
from wfmhub.config import load_config
from wfmhub.database import connect, migrate
from wfmhub.excel_templates import (
    excel_template,
    materialize_pcs_power_queries,
    require_new_template,
)
from wfmhub.on_demand_analysis import build_analysis_workbook
from wfmhub.service_profiles import load_service_profiles
from wfmhub.sota_reports import build_kpi_catalog
from wfmhub.starter_templates import build_pcs_starter


REPO = Path(__file__).resolve().parents[1]


def _home(root: Path) -> Path:
    home = root / "hub"
    (home / "config").mkdir(parents=True)
    for source in (REPO / "config").glob("default*"):
        if source.is_file():
            shutil.copy2(source, home / "config" / source.name)
    shutil.copytree(REPO / "sql", home / "sql")
    return home


def _bonus_source(path: Path) -> None:
    workbook = Workbook()
    raw = workbook.active
    raw.title = "Raw_Data"
    raw.append([
        "Agent ID", "Agent Name", "Population", "Period", "AHT", "Productivity",
        "PCS Score", "PCS % Participation", "QM", "Abs%", "VOC Detractor Count",
        "Currency", "Monthly Fixed Salary", "Target Bonus Rate",
        "Reference Bonus Override", "Eligible Days", "Scheduled Days",
        "Employment Status", "Data Status", "Notes",
    ])
    raw.append([
        "007", "Agent Seven", "OEM", "2026-08", 440, 8, 4.5, .55, .95,
        .02, 0, "MAD", None, None, 2000, 31, 31, "Active", "VALIDATED", None,
    ])
    rules = workbook.create_sheet("KPI_Config")
    rules.append([
        "Population", "KPI", "Direction", "Tier 1 Bonus %", "Tier 1 Target",
        "Tier 2 Bonus %", "Tier 2 Target",
    ])
    for row in (
        ("OEM", "AHT", "L", .15, 450, 0, 0),
        ("OEM", "Productivity", "H", .15, 7, 0, 0),
        ("OEM", "PCS Score", "H", .20, 4.3, .15, 4.1),
        ("OEM", "PCS % (Participation)", "H", .15, .5, .1, .4),
        ("OEM", "QM", "H", .2, .9, .15, .8),
        ("OEM", "Abs%", "L", .15, .05, 0, 0),
        ("OEM", "Extra Bonus (PCS Score)", "H", .3, 4.5, .2, 4.3),
    ):
        rules.append(row)
    workbook.save(path)
    workbook.close()


class BonusImportTests(unittest.TestCase):
    def test_bonus_import_is_read_only_idempotent_and_preserves_text_id(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            home = _home(root)
            source = root / "Bonus_Matrix_v1.2.xlsx"
            _bonus_source(source)
            before = source.read_bytes()
            config = load_config(home)
            migrate(config)
            conn = connect(config)
            try:
                first = import_bonus_matrix(conn, source)
                second = import_bonus_matrix(conn, source)
                self.assertFalse(first.unchanged)
                self.assertTrue(second.unchanged)
                self.assertEqual(first.period, "2026-08")
                self.assertEqual(
                    conn.execute("SELECT agent_id FROM mart.bonus_agent_month").fetchone()[0],
                    "007",
                )
                self.assertEqual(conn.execute("SELECT count(*) FROM raw.bonus_import").fetchone()[0], 1)
                self.assertEqual(conn.execute("SELECT count(*) FROM mart.bonus_agent_month").fetchone()[0], 1)
            finally:
                conn.close()
            self.assertEqual(source.read_bytes(), before)


class ExcelTemplateTests(unittest.TestCase):
    def test_public_pcs_starter_is_data_free_and_has_setup_contract(self):
        with tempfile.TemporaryDirectory() as folder:
            path = build_pcs_starter(Path(folder) / "pcs.xlsx")
            workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["START_HERE", "SETUP", "PCS_REPORT", "COACHING_LOG", "PIVOT_AREA", "FORMULAS"],
                )
                self.assertNotIn("pFeedFolder", workbook.defined_names)
                self.assertEqual(len(workbook._external_links), 0)
                self.assertIn("_system", workbook["SETUP"]["B5"].value)
            finally:
                workbook.close()

    def test_power_query_assigns_pcs_numeric_types_before_data_model_load(self):
        for filename, query_name in (
            ("PCS_AgentDay.pq", "PCS Agent Day"),
            ("PCS_Calls.pq", "PCS Calls"),
            ("PCS_Agents.pq", "PCS Agents"),
            ("PCS_Dates.pq", "PCS Dates"),
        ):
            typed_query = (REPO / "templates" / "power_query" / filename).read_text(encoding="utf-8")
            self.assertIn(f"Query name: {query_name}", typed_query)
            self.assertIn("__PCS_FEED_FOLDER__", typed_query)
            self.assertNotIn("Excel.CurrentWorkbook", typed_query)
            self.assertIn("Table.TransformColumnTypes", typed_query)
            self.assertTrue(typed_query.rstrip().endswith("Typed"))

    def test_setup_materializes_firewall_safe_pcs_queries(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            config = SimpleNamespace(home=REPO, system=root / "_system")
            paths = materialize_pcs_power_queries(config)
            self.assertEqual(len(paths), 4)
            for path in paths:
                text = path.read_text(encoding="utf-8")
                self.assertIn(str(root / "_system" / "feeds" / "pcs" / "current"), text)
                self.assertNotIn("__PCS_FEED_FOLDER__", text)
                self.assertNotIn("Excel.CurrentWorkbook", text)

    def test_governance_workbook_exposes_service_and_mapping_controls(self):
        with tempfile.TemporaryDirectory() as folder:
            home = _home(Path(folder))
            config = load_config(home)
            path = build_kpi_catalog(config)
            workbook = load_workbook(path, read_only=True)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    [
                        "METRIC_METHODS", "ACTIVITY_RULES", "ANALYTICS_RULES",
                        "REPORT_CONTRACTS", "SERVICE_PROFILES", "SERVICE_GROUPS",
                        "QUEUE_MAPPING",
                    ],
                )
                self.assertEqual(len(workbook._external_links), 0)
            finally:
                workbook.close()

    def test_master_location_is_stable_and_existing_file_is_protected(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            config = SimpleNamespace(
                home=root,
                output=root / "output",
                reports=root / "Reports",
                system=root / "_system",
            )
            template = require_new_template(config, "PCS")
            self.assertEqual(template.path, root / "Reports" / "PCS Team.xlsx")
            self.assertEqual(template.model_folder, root / "_system" / "feeds" / "pcs" / "current")
            self.assertEqual(template.feed_folder, root / "_system" / "feeds" / "pcs" / "current")
            template.path.write_bytes(b"excel-master")
            with self.assertRaises(FileExistsError):
                require_new_template(config, "PCS")
            self.assertEqual(template.path.read_bytes(), b"excel-master")
            self.assertEqual(require_new_template(config, "PCS", force=True), excel_template(config, "PCS"))

    def test_default_ford_profile_matches_original_flash_gross_sl(self):
        catalog = load_service_profiles(REPO, REPO / "config" / "default_service_profiles.toml")
        ford = catalog.select("ford_oem_fr", date(2026, 8, 25))
        self.assertEqual(ford.service_level_metric, "service_level_gross")
        self.assertEqual(ford.availability_metric, "service_availability")
        self.assertEqual([group.label for group in ford.groups], ["Toyota", "Chery", "Ford"])

    def test_forecast_analysis_uses_materialized_forecast_and_service_tables(self):
        with tempfile.TemporaryDirectory() as folder:
            home = _home(Path(folder))
            config = load_config(home)
            migrate(config)
            conn = connect(config)
            try:
                path = build_analysis_workbook(
                    conn, config, "forecast", date(2026, 8, 1), date(2026, 8, 31),
                    "target", home / "output" / "forecast_analysis.xlsx",
                )
            finally:
                conn.close()
            self.assertTrue(path.exists())


if __name__ == "__main__":
    unittest.main()
