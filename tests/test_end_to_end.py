from __future__ import annotations

import csv
import shutil
import tempfile
import unittest
import zipfile
from dataclasses import replace
from datetime import date, datetime, timezone
from pathlib import Path
from unittest.mock import patch
from zoneinfo import ZoneInfoNotFoundError

from openpyxl import Workbook, load_workbook

from wfmhub.config import ensure_user_config, load_config, write_source_root
from wfmhub.database import write_session
from wfmhub.ingestion import ingest_all
from wfmhub.models import _evaluation_time, refresh_models, resolve_period
from wfmhub.on_demand_analysis import build_analysis_workbook
from wfmhub.exports import export_dataset
from wfmhub.report_packs import build_report_pack
from wfmhub.reports import build_report


REPO = Path(__file__).resolve().parents[1]


def make_fte(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Agent"
    sheet.append(["Client ID", "Status", "Name", "Team leader", "Ops Manager", "LOB", "Market", "Language", "Location", "City", "FTE", "End date if leaver"])
    for agent_id in ("100", "200", "300"):
        sheet.append([agent_id, "Active", f"Agent {agent_id}", "TL 1", "Ops 1", "RSA", "RSA", "EN", "Onsite", "City", 1, None])
    workbook.save(path)


def make_schedule(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = ["Name", "Data Source IDs", "Scheduling Period", "Before Overtime", "After Overtime", "Shift Assignment", "Shift Events"]
    with path.open("w", encoding="cp1252", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerow(headers)
        writer.writerow(["08/01/2026", "", "", "", "", "", ""])
        writer.writerow(["Agent 100", "100", "August", "", "", ".ORG | Work 08/01/2026 8:00 AM-08/01/2026 4:00 PM", ".ORG | Late 08/01/2026 8:00 AM-08/01/2026 8:10 AM;"])
        writer.writerow(["Agent 200", "200", "August", "", "", ".ORG | Work 08/01/2026 8:00 AM-08/01/2026 4:00 PM", ""])
        writer.writerow(["Agent 300", "300", "August", "", "", ".ORG | Work 08/01/2026 8:00 AM-08/01/2026 4:00 PM", ".ORG | Late 08/01/2026 8:00 AM-08/01/2026 8:10 AM;"])
        writer.writerow(["Worldwide Agent", "999", "August", "", "", ".ORG | Work 08/01/2026 8:00 AM-08/01/2026 4:00 PM", ""])
        writer.writerow(["08/02/2026", "", "", "", "", "", ""])
        writer.writerow(["Agent 300", "300", "August", "", "", ".ORG | Work 08/02/2026 8:00 AM-08/02/2026 4:00 PM", ""])


def make_start_end_schedule(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="cp1252", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerow(["Name", "Data Source IDs", "08/01/2026", "08/02/2026"])
        writer.writerow(["Agent 100", "100", ".ORG | Work 08/01/2026 8:00 AM-08/01/2026 4:00 PM", ""])
        writer.writerow(["Agent 200", "200", ".ORG | Work 08/01/2026 8:00 AM-08/01/2026 4:00 PM", ""])
        writer.writerow(["Agent 300", "300", ".ORG | Work 08/01/2026 8:00 AM-08/01/2026 4:00 PM", ".ORG | Work 08/02/2026 8:00 AM-08/02/2026 4:00 PM"])


def make_partial_start_end_schedule(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="cp1252", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerow(["Name", "Data Source IDs", "08/01/2026"])
        writer.writerow(["Agent 100", "100", ".ORG | Work 08/01/2026 8:00 AM-08/01/2026 4:00 PM"])


def make_lilo(path: Path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["[Agent]", "[Agent ID]", "[First Log-on Time]", "[Last Log-off Time]"])
        writer.writerows(rows)


def make_status(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["[Serial Number]", "[Status]", "[Status Start Date and Time]", "[Agent]", "[Agent ID]", "[Status Duration]", "[Queue]"])
        writer.writerow(["one", "Available", "8/1/2026 8:20", "Agent 100", "100", "3:40:00", "Queue"])
        writer.writerow(["two", "Logged Off", "8/1/2026 12:00", "Agent 100", "100", "1:00:00", "Queue"])
        writer.writerow(["three", "Available", "8/1/2026 13:00", "Agent 100", "100", "3:00:00", "Queue"])
        writer.writerow(["world", "Available", "8/1/2026 9:00", "Worldwide Agent", "999", "1:00:00", "Queue"])


def make_logged_off_status(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["[Serial Number]", "[Status]", "[Status Start Date and Time]", "[Agent]", "[Agent ID]", "[Status Duration]", "[Queue]"])
        writer.writerow(["off", "Logged Off", "8/1/2026 8:00", "Agent 200", "200", "8:00:00", "Queue"])


def make_forecast(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "DATE_TIME_FORMAT\nMM/DD/YYYY hh:mm A\n"
        "Queue Name\tDate\tTime\tTime Interval\tVolume (Absolute For)\tAbandons (Absolute For)\tService Level (Absolute For)\tService Level (Absolute Req)\tActivity Handling Time (Absolute For)\tHeadcount Staffing (Absolute For)\tNet Staffing (Absolute For)\tFull Time Equivalents (Absolute For)\tFull Time Equivalents (Absolute Req)\n"
        "All\t08/01/2026\t09:00 AM\t1:00\t10\t1\t79\t80\t200\t3\t-1\t2\t3\n",
        encoding="cp1252",
    )


def make_apbe(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Report Name:", "APBE"])
    sheet.append([])
    sheet.append(["Period:", "test"])
    sheet.append([])
    sheet.append(["Date", "Language", "Queue ID", "Queue", "BusinessPartnerID", "LineOfBusiness", "15 Minute Periods of Day", "Offered_calls (w/o short calls)", "Answered_Calls", "Abandoned_Calls (w/o short calls)", "Short_calls < 5s", "Answered_Calls <= 15s", "Answered_Calls <= 20s", "Answered_Calls <= 30s", "Average_Speed_of_Answer", "Average_Talk_Time", "Average_Hold_Time", "Average Total Wrap Time"])
    sheet.append(["2026/08/01", "EN", 1, "Queue", "Partner", "RSA", "09:00", 10, 8, 2, 0, 7, 8, 8, 10, 100, 20, 10])
    workbook.save(path)


def make_apfr(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Report Name:", "APFR"])
    sheet.append([])
    sheet.append(["Date", "BusinessPartnerID", "LineOfBusiness", "15 Minute Periods of Day", "APPELS ENTRANTS", "APPELS RÉP", "APPELS ABAN", "Short_calls < 5s", "APPELS RÉP <= 15s", "APPELS RÉP <= 20s", "APPELS RÉP <= 30s", "Average_Speed_of_Answer", "Average_Talk_Time", "Average_Hold_Time", "Average Total Wrap Time"])
    sheet.append(["2026/08/01", "Partner FR", "RSA", "09:15", 12, 10, 2, 0, 8, 9, 10, 12, 110, 15, 10])
    workbook.save(path)


def make_calls(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = [
        "[Call Date/Time]", "[Call End Date/Time]", "[Call ID]",
        "[Call Reference Number]", "[Agent ID]", "[Agent]",
        "[Talk Time]", "[Hold Time]", "[Total Wrap Time]",
        "[Call Direction]", "[PostCallSurveyMode]", "[PCSStatus]",
        "[Question 1]", "[Question 2]", "[Question 3]", "[Queue]",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerow({
            "[Call Date/Time]": "8/1/2026 9:00", "[Call End Date/Time]": "8/1/2026 9:05",
            "[Call ID]": "call-100", "[Call Reference Number]": "ref-100",
            "[Agent ID]": "100", "[Agent]": "Agent 100", "[Talk Time]": "0:04:00",
            "[Hold Time]": "0:00:30", "[Total Wrap Time]": "0:00:30",
            "[Call Direction]": "I", "[PostCallSurveyMode]": "2", "[PCSStatus]": "1",
            "[Question 1]": "5", "[Question 2]": "4", "[Question 3]": "Good",
            "[Queue]": "Queue",
        })
        writer.writerow({
            "[Call Date/Time]": "8/1/2026 9:00", "[Call End Date/Time]": "8/1/2026 9:02",
            "[Call ID]": "world", "[Call Reference Number]": "world",
            "[Agent ID]": "999", "[Agent]": "Worldwide Agent", "[Talk Time]": "0:02:00",
            "[Hold Time]": "0:00:00", "[Total Wrap Time]": "0:00:00",
            "[Call Direction]": "I",
        })
        writer.writerow({
            "[Call Date/Time]": "8/1/2026 11:00", "[Call End Date/Time]": "8/1/2026 11:04",
            "[Call ID]": "low-200", "[Call Reference Number]": "ref-low-200",
            "[Agent ID]": "200", "[Agent]": "Agent 200", "[Talk Time]": "0:03:00",
            "[Hold Time]": "0:00:30", "[Total Wrap Time]": "0:00:30",
            "[Call Direction]": "I", "[PostCallSurveyMode]": "2", "[PCSStatus]": "1",
            "[Question 1]": "2", "[Question 2]": "3", "[Question 3]": "Needs follow-up",
            "[Queue]": "Queue",
        })
        for call_id, direction, status, q1, q2 in (
            ("half-score", "I", "1", "4.5", "5"),
            ("outbound-score", "O", "1", "1", ""),
            ("q2-only", "I", "1", "", "1"),
            ("answer-no-status", "I", "0", "*", ""),
            ("zero-score", "I", "1", "0", ""),
        ):
            writer.writerow({
                "[Call Date/Time]": "8/1/2026 10:00",
                "[Call End Date/Time]": "8/1/2026 10:00",
                "[Call ID]": call_id, "[Call Reference Number]": call_id,
                "[Agent ID]": "100", "[Agent]": "Agent 100",
                "[Call Direction]": direction, "[PostCallSurveyMode]": "2",
                "[PCSStatus]": status, "[Question 1]": q1,
                "[Question 2]": q2, "[Queue]": "Queue",
            })


class EndToEndTests(unittest.TestCase):
    def test_evaluation_time_uses_named_zone_and_survives_missing_tzdata(self):
        self.assertEqual(
            _evaluation_time("Europe/Berlin", datetime(2026, 8, 1, 12, tzinfo=timezone.utc)),
            datetime(2026, 8, 1, 14),
        )
        supplied = datetime(2026, 8, 1, 14)
        with patch("wfmhub.models.ZoneInfo", side_effect=ZoneInfoNotFoundError("missing")):
            self.assertEqual(_evaluation_time("Europe/Berlin", supplied), supplied)

    def test_activities_shift_assignment_is_safe_schedule_fallback(self):
        with tempfile.TemporaryDirectory() as folder:
            home = Path(folder) / "hub"
            source = Path(folder) / "source"
            (home / "config").mkdir(parents=True)
            for name in (
                "default.toml", "default_rules.toml", "default_metrics.toml",
                "default_analytics.toml", "default_reports.toml",
            ):
                shutil.copy2(REPO / "config" / name, home / "config" / name)
            shutil.copytree(REPO / "sql", home / "sql")
            make_fte(source / "FTE/FTE Count.xlsx")
            make_schedule(source / "Verint/Schedules & Activities/Activities.txt")
            make_lilo(source / "Storm/LILO/LILO 2026-08-01.csv", [
                ["Agent 100", "100", "2026-08-01 08:00:00", "2026-08-01 16:00:00"],
                ["Agent 200", "200", "", ""],
                ["Agent 300", "300", "", ""],
            ])

            config_file = ensure_user_config(home)
            write_source_root(config_file, source)
            config = load_config(home)
            with write_session(config) as conn:
                self.assertEqual(ingest_all(conn, config).failed, 0)
                refresh_models(
                    conn, config, "activities-fallback",
                    date(2026, 8, 1), date(2026, 8, 1),
                    as_of=datetime(2026, 8, 1, 17, 0),
                )
                self.assertEqual(
                    conn.execute(
                        "SELECT count(*) FROM meta.source_file "
                        "WHERE source_family='schedule' AND active=true AND source_variant='START_END'"
                    ).fetchone()[0],
                    0,
                )
                self.assertEqual(
                    conn.execute("SELECT count(*) FROM mart.attendance_agent_day").fetchone()[0],
                    3,
                )
                self.assertEqual(
                    conn.execute(
                        "SELECT attendance_result FROM mart.attendance_agent_day WHERE agent_id='100'"
                    ).fetchone()[0],
                    "Present",
                )
                fallback_issue = conn.execute(
                    "SELECT severity, details FROM meta.quality_issue "
                    "WHERE issue_type='Dedicated StartEndTimes schedule not loaded'"
                ).fetchone()
                self.assertEqual(fallback_issue[0], "REVIEW")
                self.assertIn("Activities Shift Assignment", fallback_issue[1])

    def test_activities_fills_start_end_coverage_per_agent_day(self):
        with tempfile.TemporaryDirectory() as folder:
            home = Path(folder) / "hub"
            source = Path(folder) / "source"
            (home / "config").mkdir(parents=True)
            for name in (
                "default.toml", "default_rules.toml", "default_metrics.toml",
                "default_analytics.toml", "default_reports.toml",
            ):
                shutil.copy2(REPO / "config" / name, home / "config" / name)
            shutil.copytree(REPO / "sql", home / "sql")
            make_fte(source / "FTE/FTE Count.xlsx")
            schedule_folder = source / "Verint/Schedules & Activities"
            make_partial_start_end_schedule(schedule_folder / "StartEndTimes.txt")
            make_schedule(schedule_folder / "Activities.txt")
            make_lilo(source / "Storm/LILO/LILO 2026-08-01.csv", [
                ["Agent 100", "100", "2026-08-01 08:00:00", "2026-08-01 16:00:00"],
                ["Agent 200", "200", "2026-08-01 08:00:00", "2026-08-01 16:00:00"],
                ["Agent 300", "300", "2026-08-01 08:00:00", "2026-08-01 16:00:00"],
            ])

            config_file = ensure_user_config(home)
            write_source_root(config_file, source)
            config = load_config(home)
            with write_session(config) as conn:
                self.assertEqual(ingest_all(conn, config).failed, 0)
                refresh_models(
                    conn, config, "partial-start-end",
                    date(2026, 8, 1), date(2026, 8, 1),
                    as_of=datetime(2026, 8, 1, 17, 0),
                )
                self.assertEqual(
                    conn.execute("SELECT count(*) FROM mart.attendance_agent_day").fetchone()[0],
                    3,
                )
                self.assertEqual(
                    conn.execute(
                        "SELECT count(*) FROM meta.quality_issue "
                        "WHERE issue_type='StartEndTimes coverage incomplete' AND severity='REVIEW'"
                    ).fetchone()[0],
                    1,
                )

    def test_full_day_pto_drives_expected_work_and_verint_completeness(self):
        with tempfile.TemporaryDirectory() as folder:
            home = Path(folder) / "hub"
            source = Path(folder) / "source"
            (home / "config").mkdir(parents=True)
            for name in (
                "default.toml", "default_rules.toml", "default_metrics.toml",
                "default_analytics.toml", "default_reports.toml",
            ):
                shutil.copy2(REPO / "config" / name, home / "config" / name)
            shutil.copytree(REPO / "sql", home / "sql")

            fte = source / "FTE/FTE Count.xlsx"
            make_fte(fte)
            workbook = load_workbook(fte)
            pto = workbook.create_sheet("PTO")
            pto.append([
                "Client ID", "Name", "Start date", "End date", "Day coverage",
                "Start time", "End time", "PTO type", "Approval status", "Comment",
            ])
            pto.append(["200", "Agent 200", date(2026, 8, 1), date(2026, 8, 1), "Full day", None, None, "Vacation", "Approved", "Approved request"])
            pto.append(["100", "Agent 100", date(2026, 8, 1), date(2026, 8, 1), "Partial day", "12:00", "13:00", "Personal leave", "Approved", "Appointment"])
            pto.append(["300", "Agent 300", date(2026, 8, 1), date(2026, 8, 1), "Partial day", "10:00", "11:00", "Personal leave", "Approved", "Appointment"])
            away = workbook.create_sheet("Away")
            away.append(["Client ID", "Name", "Start date", "End date", "Away type", "Case status", "Comment"])
            away.append(["300", "Agent 300", date(2026, 8, 1), date(2026, 8, 1), "Long sickness", "Planned", "Future planning only"])
            workbook.save(fte)
            workbook.close()
            make_start_end_schedule(source / "Verint/Schedules & Activities/StartEndTimes.txt")
            make_lilo(source / "Storm/LILO/LILO 2026-08-01.csv", [
                ["Agent 100", "100", "2026-08-01 08:00:00", "2026-08-01 16:00:00"],
                ["Agent 200", "200", "", ""],
                ["Agent 300", "300", "", ""],
            ])
            status_path = source / "Storm/Agent Status/Status 2026-08-01.csv"
            status_path.parent.mkdir(parents=True, exist_ok=True)
            with status_path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(["[Serial Number]", "[Status]", "[Status Start Date and Time]", "[Agent]", "[Agent ID]", "[Status Duration]", "[Queue]"])
                writer.writerow(["pto-gap", "Logged Off", "8/1/2026 12:00", "Agent 100", "100", "1:00:00", "Queue"])

            config_file = ensure_user_config(home)
            write_source_root(config_file, source)
            config = load_config(home)
            with write_session(config) as conn:
                self.assertEqual(ingest_all(conn, config).failed, 0)
                refresh_models(
                    conn, config, "pto-test", date(2026, 8, 1), date(2026, 8, 1),
                    as_of=datetime(2026, 8, 1, 17, 0),
                )
                attendance = conn.execute(
                    """SELECT attendance_result, planned_work_minutes,
                              planning_overlay_minutes, requires_call
                       FROM mart.attendance_agent_day WHERE agent_id='200'"""
                ).fetchone()
                self.assertEqual(attendance, ("PTO", 0, 480, 0))
                self.assertEqual(
                    conn.execute("SELECT count(*) FROM mart.correction_candidate WHERE agent_id='200'").fetchone()[0],
                    0,
                )
                partial = conn.execute(
                    """SELECT planned_work_minutes, planning_overlay_minutes
                       FROM mart.attendance_agent_day WHERE agent_id='100'"""
                ).fetchone()
                self.assertEqual(partial, (420, 60))
                planned_past = conn.execute(
                    """SELECT attendance_result, planning_overlay_minutes, no_show_minutes
                       FROM mart.attendance_agent_day WHERE agent_id='300'"""
                ).fetchone()
                self.assertNotEqual(planned_past[0], "Away")
                self.assertEqual(planned_past, ("No show - partial time off", 60, 420))
                split_gaps = conn.execute(
                    """SELECT time(gap_start), time(gap_end), gap_minutes
                       FROM mart.correction_candidate WHERE agent_id='300'
                       ORDER BY gap_start"""
                ).fetchall()
                self.assertEqual(split_gaps, [
                    ("08:00:00", "10:00:00", 120),
                    ("11:00:00", "16:00:00", 300),
                ])
                self.assertEqual(
                    conn.execute("SELECT count(*) FROM mart.correction_candidate WHERE agent_id='100'").fetchone()[0],
                    0,
                )
                staffing = conn.execute(
                    """SELECT gross_scheduled_fte, planned_time_off_fte, scheduled_fte
                       FROM mart.staffing_interval WHERE time(interval_start)='08:00:00'
                         AND lob='RSA' AND language='EN'"""
                ).fetchone()
                self.assertEqual(staffing, (3.0, 1.0, 2.0))
                noon_staffing = conn.execute(
                    """SELECT gross_scheduled_fte, planned_time_off_fte, scheduled_fte
                       FROM mart.staffing_interval WHERE time(interval_start)='12:00:00'
                         AND lob='RSA' AND language='EN'"""
                ).fetchone()
                self.assertEqual(noon_staffing, (3.0, 2.0, 1.0))
                self.assertEqual(
                    conn.execute(
                        "SELECT final_ledger_status FROM mart.verint_final_absence_agent_day WHERE agent_id='200'"
                    ).fetchone()[0],
                    "PLANNED_TIME_OFF_NOT_IN_VERINT",
                )

    def test_refresh_builds_safe_attendance_gaps_and_excel(self):
        with tempfile.TemporaryDirectory() as folder:
            home = Path(folder) / "hub"
            source = Path(folder) / "source"
            (home / "config").mkdir(parents=True)
            for name in (
                "default.toml", "default_rules.toml", "default_metrics.toml",
                "default_analytics.toml", "default_reports.toml",
            ):
                shutil.copy2(REPO / "config" / name, home / "config" / name)
            shutil.copytree(REPO / "sql", home / "sql")
            make_fte(source / "FTE/FTE Count.xlsx")
            make_start_end_schedule(source / "Verint/Schedules & Activities/StartEndTimes.txt")
            make_schedule(source / "Verint/Schedules & Activities/Activities.txt")
            make_lilo(source / "Storm/LILO/AP-Historical-Report---Agent-Login 2026-08-01.csv", [
                ["Agent 100", "100", "2026-08-01 08:20:00", "2026-08-01 12:00:00"],
                ["Agent 200", "200", "", ""],
                ["Worldwide Agent", "999", "2026-08-01 08:00:00", "2026-08-01 16:00:00"],
            ])
            make_lilo(source / "Storm/LILO/AP-Historical-Report---Agent-Login 2026-08-02.csv", [
                ["Agent 300", "300", "2026-08-02 08:00:00", "2026-08-02 16:00:00"],
            ])
            make_status(source / "Storm/Agent Status/AP-Historical-Report---Agent-Status 2026-08-01.csv")
            make_logged_off_status(source / "Storm/Agent Status/AP-Historical-Report---Agent-Status logged-off 2026-08-01.csv")
            make_forecast(source / "Verint/Forecast/forecast.txt")
            make_apbe(source / "Storm/APBE ALL WFM/apbe.xlsx")
            make_apfr(source / "Storm/APFR KPI SUIVI JOUR/apfr.xlsx")
            make_calls(source / "Storm/Call by Call/AP-Historical-Report---Call-by-Call 2026-08-01 - 2026-08-02.csv")
            make_calls(source / "Storm/Call by Call/AP-Historical-Report---Call-by-Call full history.csv")

            config_file = ensure_user_config(home)
            write_source_root(config_file, source)
            config = load_config(home)
            with write_session(config) as conn:
                ingest_progress = []
                ingest = ingest_all(
                    conn, config,
                    progress=lambda current, total, label: ingest_progress.append((current, total, label)),
                )
                self.assertEqual(ingest.failed, 0)
                self.assertEqual(ingest.loaded, 12)
                self.assertEqual(ingest.scoped_out, 5)
                self.assertEqual(ingest_progress[-1][:2], (12, 12))
                self.assertTrue(any(total == 0 and "Call by Call" in label for _, total, label in ingest_progress))
                for table in ("raw.schedule_shift", "raw.lilo", "raw.agent_status", "raw.call_leg"):
                    self.assertEqual(conn.execute(f"SELECT count(*) FROM {table} WHERE agent_id='999'").fetchone()[0], 0)
                model_progress = []
                model = refresh_models(
                    conn, config, "test", date(2026, 8, 1), date(2026, 8, 2),
                    progress=lambda current, total, label: model_progress.append((current, total, label)),
                )
                self.assertEqual(model_progress[-1], (22, 22, "Models ready"))
                self.assertEqual(
                    dict(conn.execute(
                        "SELECT source_variant, count(*) FROM meta.source_file WHERE source_family='schedule' AND active=true GROUP BY source_variant"
                    ).fetchall()),
                    {"ACTIVITIES": 1, "START_END": 1},
                )
                saved_period = replace(config, period_start=date(2026, 8, 1), period_end=date(2026, 8, 1))
                self.assertEqual(resolve_period(conn, saved_period, None, None, True), (date(2026, 8, 1), date(2026, 8, 1)))
                self.assertEqual(resolve_period(conn, saved_period, None, None, False), (date(2026, 8, 1), date(2026, 8, 2)))
                attendance = {row[0]: row[1] for row in conn.execute("SELECT agent_day_key, attendance_result FROM mart.attendance_agent_day").fetchall()}
                self.assertEqual(attendance["20260801-200"], "No show")
                self.assertEqual(attendance["20260801-300"], "Missing actual evidence")
                late = conn.execute("SELECT gap_start, gap_end, gap_minutes FROM mart.correction_candidate WHERE agent_id='100' AND detected_issue='Late'").fetchone()
                self.assertEqual(str(late[0]), "2026-08-01 08:00:00")
                self.assertEqual(str(late[1]), "2026-08-01 08:20:00")
                self.assertEqual(late[2], 20)
                self.assertEqual(
                    conn.execute("SELECT verint_reconciliation FROM mart.correction_candidate WHERE agent_id='100' AND detected_issue='Late'").fetchone()[0],
                    "PARTIAL",
                )
                status_gap = conn.execute(
                    "SELECT gap_minutes, observed_source, verint_reconciliation FROM mart.correction_candidate WHERE agent_id='100' AND detected_issue='Mid-shift logged off'"
                ).fetchone()
                self.assertEqual(status_gap, (60, "AGENT_STATUS", "NOT_CORRECTED"))
                self.assertEqual(
                    conn.execute(
                        "SELECT uncoded_early_leave_minutes FROM mart.attendance_agent_day WHERE agent_day_key='20260801-100'"
                    ).fetchone()[0],
                    0,
                )
                self.assertGreater(
                    conn.execute(
                        "SELECT count(*) FROM mart.staffing_interval WHERE staffing_state='DATA_PARTIAL'"
                    ).fetchone()[0],
                    0,
                )
                self.assertEqual(conn.execute("SELECT count(*) FROM mart.conformance_agent_day").fetchone()[0], 0)
                self.assertEqual(model.forecast_rows, 1)
                self.assertEqual(model.intraday_rows, 2)
                self.assertEqual(model.pcs_rows, 2)
                self.assertEqual(model.absence_rows, 4)
                self.assertGreater(model.absence_event_rows, 0)
                self.assertEqual(model.service_rows, 2)
                self.assertGreater(model.metric_rows, 0)
                self.assertGreater(model.finding_rows, 0)
                self.assertEqual(
                    conn.execute("SELECT count(*) FROM meta.metric_application WHERE run_id='test'").fetchone()[0],
                    1,
                )
                service_metric = conn.execute(
                    """SELECT sum(numerator), sum(denominator)
                       FROM mart.metric_value WHERE metric_id='service_level'"""
                ).fetchone()
                service_components = conn.execute(
                    """SELECT sum(answered_within_target), sum(offered-short_abandoned)
                       FROM mart.service_interval"""
                ).fetchone()
                self.assertAlmostEqual(
                    service_metric[0] / service_metric[1],
                    service_components[0] / service_components[1],
                )
                final_absence_max = conn.execute(
                    "SELECT max(final_absence_rate) FROM mart.verint_final_absence_agent_day"
                ).fetchone()[0]
                self.assertLessEqual(final_absence_max or 0, 1.0)
                absence_100 = conn.execute(
                    "SELECT absence_minutes, absence_rate FROM mart.absence_agent_day WHERE agent_day_key='20260801-100'"
                ).fetchone()
                self.assertEqual(absence_100[0], 80)
                self.assertAlmostEqual(absence_100[1], 80 / 480)
                self.assertEqual(
                    conn.execute(
                        "SELECT final_ledger_status FROM mart.verint_final_absence_agent_day WHERE agent_day_key='20260801-200'"
                    ).fetchone()[0],
                    "UNCODED_EMPTY_SHIFT",
                )
                self.assertEqual(
                    conn.execute(
                        "SELECT final_ledger_status FROM mart.verint_final_absence_agent_day WHERE agent_day_key='20260801-100'"
                    ).fetchone()[0],
                    "PARTIAL_CORRECTION_REVIEW",
                )
                self.assertEqual(
                    conn.execute("SELECT count(*) FROM mart.absence_event WHERE evidence_type IN ('SHIFT_EVENT','SHIFT_ASSIGNMENT')").fetchone()[0],
                    0,
                )
                self.assertEqual(
                    conn.execute("SELECT absence_minutes FROM mart.absence_agent_day WHERE agent_day_key='20260802-300'").fetchone()[0],
                    0,
                )
                self.assertEqual(
                    conn.execute("SELECT exception_type FROM mart.verint_final_exception WHERE agent_id='300'").fetchone()[0],
                    "VERINT_FINAL_WITHOUT_OBSERVED_GAP",
                )
                self.assertEqual(
                    conn.execute(
                        "SELECT final_ledger_status FROM mart.verint_final_absence_agent_day WHERE agent_day_key='20260801-300'"
                    ).fetchone()[0],
                    "VERINT_WITHOUT_OBSERVED_GAP",
                )
                service = conn.execute(
                    "SELECT sum(answered), sum(offered), sum(handled_seconds) FROM mart.service_interval"
                ).fetchone()
                self.assertEqual(service[:2], (18, 22))
                self.assertEqual(service[2], 2390)
                self.assertEqual(conn.execute("SELECT count(*) FROM raw.call_leg").fetchone()[0], 14)
                self.assertEqual(conn.execute("SELECT count(*) FROM core.clean_call_leg").fetchone()[0], 7)
                pcs = conn.execute(
                    """SELECT survey_responses, pcs_average, average_handle_seconds,
                              pcs_status_calls, pcs_participation_responses,
                              pcs_invalid_responses, pcs_status_blank_responses,
                              pcs_response_without_status, low_score_responses,
                              top_box_responses
                       FROM mart.agent_pcs_day WHERE agent_id='100'"""
                ).fetchone()
                self.assertEqual(pcs[0], 1)
                self.assertEqual(pcs[1], 5.0)
                self.assertEqual(pcs[2], 300)
                self.assertEqual(pcs[3:], (4, 4, 3, 1, 1, 0, 1))
                self.assertEqual(
                    conn.execute(
                        "SELECT metric_value FROM mart.metric_value WHERE metric_id='pcs_average'"
                    ).fetchone()[0],
                    pcs[1],
                )

                refresh_models(
                    conn, config, "pcs-today-only", date(2026, 8, 2), date(2026, 8, 2),
                    as_of=datetime(2026, 8, 2, 17, 0),
                )
                self.assertEqual(
                    str(conn.execute(
                        "SELECT min(business_date) FROM mart.agent_pcs_day"
                    ).fetchone()[0]),
                    "2026-08-01",
                )

                # A refresh during an active shift is provisional: the future
                # part of that shift must not become an early-leave or no-show
                # correction. Once the shift has ended, normal final logic
                # applies again.
                refresh_models(
                    conn, config, "in-progress", date(2026, 8, 1), date(2026, 8, 1),
                    as_of=datetime(2026, 8, 1, 12, 0),
                )
                provisional = dict(conn.execute(
                    "SELECT agent_id, attendance_result FROM mart.attendance_agent_day"
                ).fetchall())
                self.assertEqual(provisional["100"], "Late - shift in progress")
                self.assertEqual(provisional["200"], "Not seen - shift in progress")
                self.assertEqual(
                    conn.execute("SELECT call_action FROM mart.attendance_agent_day WHERE agent_id='200'").fetchone()[0],
                    "CALL_NOT_SEEN_NOW",
                )
                self.assertEqual(
                    conn.execute(
                        "SELECT final_ledger_status FROM mart.verint_final_absence_agent_day WHERE agent_id='200'"
                    ).fetchone()[0],
                    "PROVISIONAL_DAY",
                )
                self.assertEqual(
                    conn.execute(
                        """SELECT count(*) FROM mart.correction_candidate
                           WHERE detected_issue IN ('Early leave','No show')"""
                    ).fetchone()[0],
                    0,
                )
                self.assertIsNone(conn.execute(
                    "SELECT attendance_percent FROM mart.attendance_agent_day WHERE agent_id='100'"
                ).fetchone()[0])
                refresh_models(
                    conn, config, "completed-today", date(2026, 8, 1), date(2026, 8, 2),
                    as_of=datetime(2026, 8, 2, 17, 0),
                )
                self.assertEqual(
                    conn.execute(
                        "SELECT attendance_result FROM mart.attendance_agent_day WHERE agent_day_key='20260801-200'"
                    ).fetchone()[0],
                    "No show",
                )
                before_failure = conn.execute(
                    "SELECT agent_day_key, attendance_result FROM mart.attendance_agent_day ORDER BY agent_day_key"
                ).fetchall()
                with patch("wfmhub.models._build_absence", side_effect=RuntimeError("injected model failure")):
                    with self.assertRaises(RuntimeError):
                        refresh_models(conn, config, "failed", date(2026, 8, 1), date(2026, 8, 2))
                self.assertEqual(
                    conn.execute("SELECT agent_day_key, attendance_result FROM mart.attendance_agent_day ORDER BY agent_day_key").fetchall(),
                    before_failure,
                )
                report = build_report(conn, config, model.start, model.end)
                self.assertEqual(report.parent, home / "_system" / "legacy_reports")
                self.assertEqual(report.name, "Legacy Daily Operations.xlsx")
                corrections_report = build_report_pack("corrections", conn, config, model.start, model.end)
                pcs_report = build_report_pack("quality_pcs", conn, config, model.start, model.end)
                focused_pcs_report = build_report_pack("pcs", conn, config, model.start, model.end)
                service_report = build_report_pack(
                    "service", conn, config, model.start, model.end,
                    service_profile="ford_oem_fr",
                )
                realisations_report = build_report_pack(
                    "realisations", conn, config, model.start, model.end,
                    service_profile="ford_oem_fr",
                )
                all_realisations_report = build_report_pack(
                    "realisations", conn, config, model.start, model.end,
                    home / "output" / "all-realisations.xlsx",
                )
                staffing_report = build_report_pack(
                    "staffing", conn, config, model.start, model.end,
                )
                attendance_report = build_report_pack("attendance", conn, config, model.start, model.end)
                absence_report = build_report_pack("absence", conn, config, model.start, model.end)
                analysis_report = build_analysis_workbook(
                    conn, config, "pcs", model.start, model.end,
                )
                export_progress = []
                clean_calls = export_dataset(
                    conn, config, "calls", model.start, model.end,
                    progress=lambda current, total, label: export_progress.append((current, total, label)),
                )
                self.assertEqual(clean_calls.rows, 7)
                self.assertTrue(clean_calls.manifest.exists())
                self.assertEqual(export_progress[-1], (7, 0, "Exported calls: 7 rows"))
                governed_service = export_dataset(
                    conn, config, "daily_service_lob", model.start, model.end,
                )
                governed_pcs = export_dataset(
                    conn, config, "pcs_team_day", model.start, model.end,
                )
                governed_absence = export_dataset(
                    conn, config, "final_absence_lob_month", model.start, model.end,
                )
                self.assertIn(
                    "service_level",
                    governed_service.path.read_text(encoding="utf-8-sig").splitlines()[0],
                )
                self.assertEqual(governed_pcs.rows, 1)
                self.assertGreater(governed_absence.rows, 0)
                self.assertIn(
                    "Metric catalog SHA-256:",
                    governed_service.manifest.read_text(encoding="utf-8"),
                )

                filtered_report = build_report(
                    conn, config, date(2026, 8, 1), date(2026, 8, 1),
                    home / "output" / "filtered.xlsx",
                )
                filtered = load_workbook(filtered_report, read_only=True, data_only=True)
                try:
                    filtered_dates = [row[0] for row in filtered["ATTENDANCE_CALLS"].iter_rows(min_row=5, values_only=True) if row[0] is not None]
                    self.assertGreaterEqual(len(filtered_dates), 1)
                    self.assertTrue(all(value.date() == date(2026, 8, 1) for value in filtered_dates))
                finally:
                    filtered.close()

                pcs_edit = load_workbook(focused_pcs_report)
                pcs_actions = pcs_edit["COACHING"]
                pcs_headers = {cell.value: cell.column for cell in pcs_actions[4]}
                self.assertEqual(pcs_actions.max_row, 5)
                self.assertEqual(pcs_actions.cell(5, pcs_headers["Agent ID"]).value, "200")
                pcs_actions.cell(5, pcs_headers["Coaching Status"], "Completed")
                pcs_actions.cell(5, pcs_headers["Coaching Date"], date(2026, 8, 2))
                pcs_actions.cell(5, pcs_headers["Coach"], "TL 1")
                pcs_actions.cell(5, pcs_headers["Coaching Comment"], "Reviewed")
                pcs_edit.save(focused_pcs_report)
                pcs_edit.close()
                self.assertEqual(conn.execute("SELECT count(*) FROM core.pcs_coaching_action").fetchone()[0], 0)

            self.assertTrue(report.exists())
            workbook = load_workbook(report, read_only=True, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, [
                    "DAILY_SUMMARY", "FINDINGS", "ATTENDANCE_CALLS", "STAFFING_GAPS",
                    "SERVICE_LEVEL", "DATA_QUALITY", "SOURCE_HEALTH", "SCHEDULE_SOURCES",
                    "DOMAIN_RULES", "METHODS", "PROVENANCE",
                ])
            finally:
                workbook.close()
            corrections_book = load_workbook(corrections_report, read_only=True, data_only=True)
            try:
                self.assertEqual(corrections_book.sheetnames, [
                    "DASHBOARD", "VERINT_INJECTION", "SHIFT_VIEW", "DEFINITIONS", "_AUDIT",
                ])
                self.assertIn("2026-08-01 to 2026-08-02", corrections_book["DASHBOARD"]["A2"].value)
                injection = corrections_book["VERINT_INJECTION"]
                injection_headers = [cell.value for cell in injection[4]]
                self.assertIn("Start To Inject", injection_headers)
                self.assertIn("End To Inject", injection_headers)
                self.assertNotIn("Validation Status", injection_headers)
                start_column = injection_headers.index("Start To Inject")
                end_column = injection_headers.index("End To Inject")
                exact_intervals = {
                    (row[start_column], row[end_column])
                    for row in injection.iter_rows(min_row=5, values_only=True)
                    if row[start_column] is not None
                }
                self.assertIn(
                    (datetime(2026, 8, 1, 12, 0), datetime(2026, 8, 1, 13, 0)),
                    exact_intervals,
                )
                self.assertEqual(corrections_book["_AUDIT"].sheet_state, "hidden")
            finally:
                corrections_book.close()
            pcs_book = load_workbook(pcs_report, read_only=True, data_only=True)
            try:
                self.assertIn("AGENT_MONTH", pcs_book.sheetnames)
                self.assertIn("RESPONSE_DETAIL", pcs_book.sheetnames)
                self.assertIn("METHODS", pcs_book.sheetnames)
            finally:
                pcs_book.close()
            attendance_book = load_workbook(attendance_report, read_only=True, data_only=True)
            try:
                self.assertIn("2026-08-01 to 2026-08-02", attendance_book["DASHBOARD"]["A2"].value)
                action_dates = [
                    row[0].date() if isinstance(row[0], datetime) else row[0]
                    for row in attendance_book["ACTIONS"].iter_rows(min_row=5, values_only=True)
                    if row[0] is not None
                ]
                self.assertIn(date(2026, 8, 1), action_dates)
                self.assertEqual(attendance_book["_AUDIT"].sheet_state, "hidden")
            finally:
                attendance_book.close()
            focused_pcs_book = load_workbook(focused_pcs_report, read_only=False, data_only=False)
            try:
                self.assertEqual(focused_pcs_book.sheetnames, [
                    "DASHBOARD", "TEAM_VIEW", "AGENT_RESULTS", "COACHING",
                    "COACHING_QUEUE", "PCS_DATA", "HELP", "DEFINITIONS",
                    "_LOOKUPS", "_AUDIT",
                ])
                self.assertIn("SUMPRODUCT", focused_pcs_book["DASHBOARD"]["A11"].value)
                self.assertIn("tblPcsData", focused_pcs_book["PCS_DATA"].tables)
                self.assertIn("tblCoaching", focused_pcs_book["COACHING"].tables)
                self.assertIn("tblCoachingQueue", focused_pcs_book["COACHING_QUEUE"].tables)
                self.assertIn("tblPcsData", focused_pcs_book["AGENT_RESULTS"]["I5"].value)
                team_formula = focused_pcs_book["TEAM_VIEW"]["A11"].value
                self.assertIn("tblPcsData", getattr(team_formula, "text", ""))
                self.assertEqual(focused_pcs_book["_LOOKUPS"].sheet_state, "hidden")
                pcs_table_headers = [cell.value for cell in focused_pcs_book["PCS_DATA"][4]]
                defined = {
                    item.name: item.attr_text
                    for item in focused_pcs_book.defined_names.values()
                }
                self.assertIn("Current week", defined["PCS_From"])
            finally:
                focused_pcs_book.close()
            service_book = load_workbook(service_report, read_only=True, data_only=True)
            try:
                self.assertEqual(service_report, home / "Reports" / "OEM Flash.xlsx")
                self.assertEqual(service_book.sheetnames, [
                    "FLASH", "HOURLY", "QUEUES", "EXCEPTIONS", "DATA_STATUS",
                    "DEFINITIONS", "_AUDIT",
                ])
                self.assertEqual(service_book["FLASH"]["A1"].value, "OEM FLASH  /  FORD OEM FRANCE")
            finally:
                service_book.close()
            absence_book = load_workbook(absence_report, read_only=False, data_only=False)
            try:
                self.assertEqual(absence_book.sheetnames, [
                    "DASHBOARD", "TEAM_VIEW", "TEAM_SUMMARY", "AGENT_RESULTS",
                    "ACTIONS", "ACTION_QUEUE", "ABSENCE_COMPONENTS",
                    "SHRINKAGE_COMPONENTS", "COMPONENT_VIEW", "ACTIVITY_DETAIL",
                    "ABSENCE_DATA", "HELP", "DEFINITIONS", "_LOOKUPS", "_AUDIT",
                ])
                self.assertIn("tblAbsenceData", absence_book["ABSENCE_DATA"].tables)
                self.assertIn("tblActions", absence_book["ACTIONS"].tables)
                self.assertIn("tblActionQueue", absence_book["ACTION_QUEUE"].tables)
                absence_team_formula = absence_book["TEAM_VIEW"]["A17"].value
                self.assertIn("tblAbsenceData", getattr(absence_team_formula, "text", ""))
                component_formula = absence_book["COMPONENT_VIEW"]["A7"].value
                self.assertIn("tblActivityDetail", getattr(component_formula, "text", ""))
                self.assertEqual(absence_book["_LOOKUPS"].sheet_state, "hidden")
                self.assertEqual(absence_book["_AUDIT"].sheet_state, "hidden")
                absence_table_headers = [cell.value for cell in absence_book["ABSENCE_DATA"][4]]
            finally:
                absence_book.close()
            self.assertTrue((home / "Feed").is_dir())
            pcs_feed = home / "Feed" / "PCS" / "PCS_AGENT_DAY_CURRENT.csv"
            absence_feed = home / "Feed" / "Absenteeism" / "ABSENCE_AGENT_DAY_CURRENT.csv"
            self.assertTrue(pcs_feed.is_file())
            self.assertTrue(absence_feed.is_file())
            with pcs_feed.open("r", encoding="utf-8-sig", newline="") as handle:
                self.assertEqual(
                    next(csv.reader(handle)),
                    pcs_table_headers,
                )
            with absence_feed.open("r", encoding="utf-8-sig", newline="") as handle:
                self.assertEqual(
                    next(csv.reader(handle)),
                    absence_table_headers,
                )
            realisations_book = load_workbook(realisations_report, read_only=False, data_only=False)
            try:
                self.assertEqual(realisations_book.sheetnames, [
                    "DASHBOARD", "LOB_RESULTS", "TREND", "DATA", "DEFINITIONS", "_AUDIT",
                ])
                self.assertIn("Processing Hours", [cell.value for cell in realisations_book["LOB_RESULTS"][4]])
                self.assertGreaterEqual(len(realisations_book["DASHBOARD"]._charts), 1)
            finally:
                realisations_book.close()
            all_realisations_book = load_workbook(
                all_realisations_report, read_only=True, data_only=True,
            )
            try:
                reporting_lobs = {
                    row[4]
                    for row in all_realisations_book["LOB_RESULTS"].iter_rows(
                        min_row=5, values_only=True,
                    )
                    if row[1]
                }
                self.assertEqual(reporting_lobs, {
                    "Ford OEM France", "Ford Netherlands", "RSA Belgium",
                    "RSA Netherlands",
                })
            finally:
                all_realisations_book.close()
            staffing_book = load_workbook(staffing_report, read_only=True, data_only=True)
            try:
                self.assertIn("WEEKLY_PLAN", staffing_book.sheetnames)
                staffing_dates = {
                    row[0].date() if isinstance(row[0], datetime) else row[0]
                    for row in staffing_book["INTRADAY"].iter_rows(
                        min_row=5, values_only=True,
                    )
                    if row[0] is not None
                }
                self.assertEqual(staffing_dates, {date(2026, 8, 1), date(2026, 8, 2)})
            finally:
                staffing_book.close()
            self.assertEqual(analysis_report.parent, home / "Reports" / "Analysis")
            self.assertTrue(analysis_report.is_file())
            for generated_report in (
                report, corrections_report, pcs_report, focused_pcs_report,
                service_report, realisations_report, all_realisations_report,
                staffing_report, attendance_report, absence_report, analysis_report,
            ):
                with zipfile.ZipFile(generated_report) as archive:
                    self.assertFalse(any("externalLinks" in name for name in archive.namelist()))


if __name__ == "__main__":
    unittest.main()
